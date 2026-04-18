"""# RC DOMOTIC - Cotizador App v3.0
Flask + SQLite3 | Sin dependencias de nube | Funciona en local o servidor
v3: Fix Excel, Selector categorÃ­as, Merge duplicados, Logo+Watermark PDF,
    Dashboard APROBADA, CatÃ¡logo Pro, Comandos seguros
"""
import sqlite3, os, json, datetime, re, uuid, traceback, secrets, time, ipaddress, socket, unicodedata, shutil, threading
from functools import wraps
from flask import (Flask, request, jsonify, g, send_file, session,
                   render_template_string, send_from_directory, abort)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from io import BytesIO
from PIL import Image
import urllib.request
import urllib.parse
# â”€â”€â”€ Seguridad opcional (si estÃ¡n instaladas) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
try:
    from argon2 import PasswordHasher
    from argon2.exceptions import VerifyMismatchError
    _ARGON2 = PasswordHasher()
except Exception:
    PasswordHasher = None
    VerifyMismatchError = Exception
    _ARGON2 = None

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
except Exception:
    Limiter = None
    get_remote_address = None

_DEFAULT_SECRET_KEY = 'rcdomotic-dev-secret'
_DEFAULT_BOT_KEY = 'test-key-local-123'

app = Flask(__name__)
# Soporte de proxy (Render/NGINX): respeta X-Forwarded-Proto/Host
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.secret_key = os.environ.get('SECRET_KEY', _DEFAULT_SECRET_KEY)

# â”€â”€â”€ Seguridad base â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
IS_PROD = bool(os.environ.get('RENDER')) or (os.environ.get('FLASK_ENV','').lower() == 'production') or (os.environ.get('ENV','').lower() == 'production')
# Cookies de sesiÃ³n
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=IS_PROD,  # en Render normalmente es HTTPS
    PERMANENT_SESSION_LIFETIME=datetime.timedelta(hours=10),
)
# LÃ­mite de carga (uploads)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR_ENV = (os.environ.get('DATA_DIR') or '').strip()
_RENDER_DATA_DIR = '/var/data'
if _DATA_DIR_ENV:
    DATA_DIR = _DATA_DIR_ENV
elif os.path.isdir(_RENDER_DATA_DIR):
    DATA_DIR = _RENDER_DATA_DIR
else:
    DATA_DIR = BASE_DIR
DB_PATH = os.environ.get('DB_PATH', os.path.join(DATA_DIR, 'rc_domotic.db'))

# ── Abstracción de base de datos ─────────────────────────────────────────────
# Cuando DATABASE_URL esté definido (Postgres en producción), get_db() usa
# psycopg2. Para migrar: define DATABASE_URL=postgres://... en el entorno.
_DATABASE_URL = os.environ.get('DATABASE_URL', '').strip()
_USE_POSTGRES = _DATABASE_URL.startswith('postgres')
if _USE_POSTGRES:
    try:
        import psycopg2, psycopg2.extras
        _PG_AVAILABLE = True
    except ImportError:
        _PG_AVAILABLE = False
        print("[WARN] DATABASE_URL definido pero psycopg2 no está instalado. Usando SQLite.")
        _USE_POSTGRES = False
else:
    _PG_AVAILABLE = False

UPLOADS_DIR = os.environ.get('UPLOADS_DIR', os.path.join(DATA_DIR, 'uploads'))
UPLOAD_FOLDER = os.path.join(UPLOADS_DIR, 'products')
FALLBACK_UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads', 'products')
BACKUP_DIR = os.environ.get('BACKUP_DIR', os.path.join(DATA_DIR, 'backups'))
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)
AUDIT_LOG_PATH = os.environ.get('AUDIT_LOG_PATH') or os.path.join(DATA_DIR, 'security_audit.jsonl')
os.makedirs(os.path.dirname(AUDIT_LOG_PATH) or DATA_DIR, exist_ok=True)

def _env_flag(name: str, default=False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return bool(default)
    return str(raw).strip().lower() in ('1', 'true', 'yes', 'on', 'si', 'sí')

def _env_int(name: str, default: int, min_v=None, max_v=None) -> int:
    try:
        value = int(str(os.environ.get(name, default)).strip())
    except Exception:
        value = int(default)
    if min_v is not None:
        value = max(int(min_v), value)
    if max_v is not None:
        value = min(int(max_v), value)
    return value

def _parse_ip_networks(raw: str, source: str = ''):
    raw = (raw or '').strip()
    if not raw:
        return []
    out = []
    for piece in raw.split(','):
        token = piece.strip()
        if not token:
            continue
        try:
            if '/' in token:
                out.append(ipaddress.ip_network(token, strict=False))
            else:
                ip_obj = ipaddress.ip_address(token)
                prefix = 32 if ip_obj.version == 4 else 128
                out.append(ipaddress.ip_network(f'{token}/{prefix}', strict=False))
        except Exception:
            if source:
                print(f"[SECURITY] WARNING: IP/CIDR invalido en {source}: {token}")
    return out

def _configured_ip_networks(env_name: str):
    return _parse_ip_networks((os.environ.get(env_name) or '').strip(), source=env_name)

ADMIN_IP_ALLOWLIST_DEFAULT = _configured_ip_networks('ADMIN_IP_ALLOWLIST')
BOT_IP_ALLOWLIST_DEFAULT = _configured_ip_networks('BOT_IP_ALLOWLIST')
ADMIN_REAUTH_SECONDS_DEFAULT = _env_int('ADMIN_REAUTH_SECONDS', 900 if IS_PROD else 0, min_v=0, max_v=86400)
AUTO_BACKUP_ENABLED_DEFAULT = _env_flag('AUTO_BACKUP_ENABLED', IS_PROD)
AUTO_BACKUP_EVERY_MIN_DEFAULT = _env_int('AUTO_BACKUP_EVERY_MIN', 360, min_v=15, max_v=10080)
AUTO_BACKUP_KEEP_DEFAULT = _env_int('AUTO_BACKUP_KEEP', 40 if IS_PROD else 12, min_v=3, max_v=500)
SECURITY_BACKUP_MAX_AGE_HOURS_DEFAULT = _env_int('SECURITY_BACKUP_MAX_AGE_HOURS', 24 if IS_PROD else 168, min_v=1, max_v=720)
SECURITY_ALERT_WEBHOOK_DEFAULT = (os.environ.get('SECURITY_ALERT_WEBHOOK') or '').strip()
BACKUP_WEBHOOK_URL_DEFAULT = (os.environ.get('BACKUP_WEBHOOK_URL') or '').strip()

_AUTO_BACKUP_LOCK_PATH = os.path.join(BACKUP_DIR, '.auto_backup.lock')
_AUTO_BACKUP_THREAD = None
_AUTO_BACKUP_THREAD_LOCK = threading.Lock()
_ALERT_LOCK = threading.Lock()
_ALERT_LAST_SENT = {}
_SEC_SETTINGS_LOCK = threading.Lock()
_SEC_SETTINGS_CACHE = {'ts': 0.0, 'data': None}

_SECURITY_PARAM_KEYS = (
    'security_admin_ip_allowlist',
    'security_bot_ip_allowlist',
    'security_admin_reauth_seconds',
    'security_auto_backup_enabled',
    'security_auto_backup_every_min',
    'security_auto_backup_keep',
    'security_backup_max_age_hours',
    'security_alert_webhook',
    'security_backup_webhook',
)

def _coerce_int(value, default: int, min_v=None, max_v=None) -> int:
    try:
        out = int(str(value).strip())
    except Exception:
        out = int(default)
    if min_v is not None:
        out = max(int(min_v), out)
    if max_v is not None:
        out = min(int(max_v), out)
    return out

def _coerce_bool(value, default=False) -> bool:
    if value is None:
        return bool(default)
    s = str(value).strip().lower()
    if s in ('1', 'true', 'yes', 'on', 'si', 'sí'):
        return True
    if s in ('0', 'false', 'no', 'off'):
        return False
    return bool(default)

def _load_security_param_overrides(force=False, max_age_seconds=20):
    now = time.time()
    if (not force) and _SEC_SETTINGS_CACHE.get('data') is not None and (now - float(_SEC_SETTINGS_CACHE.get('ts') or 0) < max_age_seconds):
        return dict(_SEC_SETTINGS_CACHE.get('data') or {})
    with _SEC_SETTINGS_LOCK:
        if (not force) and _SEC_SETTINGS_CACHE.get('data') is not None and (time.time() - float(_SEC_SETTINGS_CACHE.get('ts') or 0) < max_age_seconds):
            return dict(_SEC_SETTINGS_CACHE.get('data') or {})
        out = {}
        try:
            if os.path.isfile(DB_PATH):
                conn = sqlite3.connect(DB_PATH)
                try:
                    cur = conn.cursor()
                    marks = ','.join(['?'] * len(_SECURITY_PARAM_KEYS))
                    rows = cur.execute(f"SELECT clave, valor FROM parametros WHERE clave IN ({marks})", _SECURITY_PARAM_KEYS).fetchall()
                    for k, v in rows:
                        out[str(k)] = '' if v is None else str(v)
                finally:
                    conn.close()
        except Exception:
            out = {}
        _SEC_SETTINGS_CACHE['ts'] = time.time()
        _SEC_SETTINGS_CACHE['data'] = dict(out)
        return dict(out)

def _runtime_security_settings(force=False):
    over = _load_security_param_overrides(force=force)

    def env_text(name):
        if name in os.environ:
            return str(os.environ.get(name) or '').strip(), True
        return '', False

    def pick_text(env_name, param_key, default=''):
        e, has_env = env_text(env_name)
        if has_env:
            return e
        p = str(over.get(param_key) or '').strip()
        return p if p else str(default or '').strip()

    def pick_int(env_name, param_key, default, min_v=None, max_v=None):
        if env_name in os.environ:
            return _coerce_int(os.environ.get(env_name), default, min_v=min_v, max_v=max_v)
        p = over.get(param_key)
        if p is None or str(p).strip() == '':
            return _coerce_int(default, default, min_v=min_v, max_v=max_v)
        return _coerce_int(p, default, min_v=min_v, max_v=max_v)

    def pick_bool(env_name, param_key, default=False):
        if env_name in os.environ:
            return _coerce_bool(os.environ.get(env_name), default=default)
        p = over.get(param_key)
        if p is None or str(p).strip() == '':
            return bool(default)
        return _coerce_bool(p, default=default)

    admin_ip_raw = pick_text('ADMIN_IP_ALLOWLIST', 'security_admin_ip_allowlist', '')
    bot_ip_raw = pick_text('BOT_IP_ALLOWLIST', 'security_bot_ip_allowlist', '')
    admin_ip_allowlist = _parse_ip_networks(admin_ip_raw, source='runtime_admin_ip_allowlist') if admin_ip_raw else list(ADMIN_IP_ALLOWLIST_DEFAULT)
    bot_ip_allowlist = _parse_ip_networks(bot_ip_raw, source='runtime_bot_ip_allowlist') if bot_ip_raw else list(BOT_IP_ALLOWLIST_DEFAULT)

    admin_reauth_seconds = pick_int('ADMIN_REAUTH_SECONDS', 'security_admin_reauth_seconds', ADMIN_REAUTH_SECONDS_DEFAULT, min_v=0, max_v=86400)
    auto_backup_enabled = pick_bool('AUTO_BACKUP_ENABLED', 'security_auto_backup_enabled', AUTO_BACKUP_ENABLED_DEFAULT)
    auto_backup_every_min = pick_int('AUTO_BACKUP_EVERY_MIN', 'security_auto_backup_every_min', AUTO_BACKUP_EVERY_MIN_DEFAULT, min_v=15, max_v=10080)
    auto_backup_keep = pick_int('AUTO_BACKUP_KEEP', 'security_auto_backup_keep', AUTO_BACKUP_KEEP_DEFAULT, min_v=3, max_v=500)
    backup_max_age_hours = pick_int('SECURITY_BACKUP_MAX_AGE_HOURS', 'security_backup_max_age_hours', SECURITY_BACKUP_MAX_AGE_HOURS_DEFAULT, min_v=1, max_v=720)
    security_alert_webhook = pick_text('SECURITY_ALERT_WEBHOOK', 'security_alert_webhook', SECURITY_ALERT_WEBHOOK_DEFAULT)
    backup_webhook = pick_text('BACKUP_WEBHOOK_URL', 'security_backup_webhook', BACKUP_WEBHOOK_URL_DEFAULT)

    return {
        'admin_ip_allowlist': admin_ip_allowlist,
        'admin_ip_allowlist_count': len(admin_ip_allowlist),
        'bot_ip_allowlist': bot_ip_allowlist,
        'bot_ip_allowlist_count': len(bot_ip_allowlist),
        'admin_reauth_seconds': int(admin_reauth_seconds),
        'auto_backup_enabled': bool(auto_backup_enabled),
        'auto_backup_every_min': int(auto_backup_every_min),
        'auto_backup_keep': int(auto_backup_keep),
        'backup_max_age_hours': int(backup_max_age_hours),
        'security_alert_webhook': str(security_alert_webhook or '').strip(),
        'backup_webhook': str(backup_webhook or '').strip(),
    }

def _maybe_migrate_legacy_db():
    """Si cambió DB_PATH a ruta persistente, copia la BD legacy una sola vez."""
    try:
        legacy = os.path.join(BASE_DIR, 'rc_domotic.db')
        if os.path.abspath(legacy) == os.path.abspath(DB_PATH):
            return
        if os.path.isfile(DB_PATH):
            return
        if os.path.isfile(legacy):
            os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
            shutil.copy2(legacy, DB_PATH)
            print(f"[DB] Migrada BD legacy -> {DB_PATH}")
    except Exception as e:
        print("[WARN] No se pudo migrar BD legacy:", e)

_maybe_migrate_legacy_db()

def snapshot_db(reason: str = 'manual'):
    """Crea copia de seguridad sqlite en BACKUP_DIR."""
    if not os.path.isfile(DB_PATH):
        return None
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_reason = re.sub(r'[^a-zA-Z0-9_-]+', '_', reason or 'manual').strip('_') or 'manual'
    out = os.path.join(BACKUP_DIR, f'rc_domotic_{safe_reason}_{ts}.db')
    # Backup consistente aun con escrituras concurrentes.
    src = sqlite3.connect(DB_PATH)
    try:
        dst = sqlite3.connect(out)
        try:
            src.backup(dst)
            dst.commit()
        finally:
            dst.close()
    finally:
        src.close()
    try:
        shutil.copystat(DB_PATH, out)
    except Exception:
        pass
    return out

def _db_counts_from_file(path: str):
    conn = sqlite3.connect(path)
    try:
        cur = conn.cursor()
        c = cur.execute("SELECT COUNT(*) FROM cotizaciones").fetchone()[0]
        i = cur.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        return int(c), int(i)
    finally:
        conn.close()

def _client_ip() -> str:
    try:
        route = list(request.access_route or [])
        if route:
            return str(route[0]).strip()
    except Exception:
        pass
    try:
        return str(request.remote_addr or '').strip()
    except Exception:
        return ''

def _ip_allowed(ip_text: str, networks) -> bool:
    if not networks:
        return True
    try:
        ip_obj = ipaddress.ip_address((ip_text or '').strip())
    except Exception:
        return False
    for net in networks:
        try:
            if ip_obj in net:
                return True
        except Exception:
            continue
    return False

def _list_backups(limit=0):
    if not os.path.isdir(BACKUP_DIR):
        return []
    rows = []
    for fn in os.listdir(BACKUP_DIR):
        fp = os.path.join(BACKUP_DIR, fn)
        if not (os.path.isfile(fp) and fn.lower().endswith('.db')):
            continue
        rows.append({
            'path': os.path.abspath(fp),
            'name': fn,
            'size': int(os.path.getsize(fp)),
            'mtime': int(os.path.getmtime(fp)),
        })
    rows.sort(key=lambda x: x['mtime'], reverse=True)
    if limit and limit > 0:
        return rows[:limit]
    return rows

def _latest_backup_info():
    rows = _list_backups(limit=1)
    return rows[0] if rows else None

def _verify_sqlite_file(path: str):
    result = {'ok': False, 'integrity': '', 'cotizaciones': 0, 'items': 0}
    if not path or not os.path.isfile(path):
        result['integrity'] = 'missing'
        return result
    conn = sqlite3.connect(path)
    try:
        row = conn.execute('PRAGMA integrity_check').fetchone()
        msg = str((row or [''])[0] or '').strip().lower()
        result['integrity'] = msg or 'empty'
        result['ok'] = (msg == 'ok')
    except Exception as e:
        result['integrity'] = f'error:{e}'
    finally:
        conn.close()
    try:
        c, i = _db_counts_from_file(path)
        result['cotizaciones'] = int(c)
        result['items'] = int(i)
    except Exception:
        pass
    return result

def _prune_backup_files(keep_n: int):
    keep_n = max(1, int(keep_n or 1))
    rows = _list_backups(limit=0)
    removed = 0
    for row in rows[keep_n:]:
        try:
            os.remove(row['path'])
            removed += 1
        except Exception:
            continue
    return removed

def _post_json_webhook_async(url: str, payload: dict, timeout=6):
    if not url:
        return
    if str(url).strip().lower().startswith('internal://'):
        return
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    headers = {'Content-Type': 'application/json; charset=utf-8'}
    def _worker():
        req = urllib.request.Request(url, data=body, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                resp.read(64)
        except Exception:
            pass
    threading.Thread(target=_worker, daemon=True).start()

def _acquire_file_lock(lock_path: str, stale_seconds=7200):
    os.makedirs(os.path.dirname(lock_path) or BACKUP_DIR, exist_ok=True)
    try:
        fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(int(time.time())).encode('utf-8'))
        return fd
    except FileExistsError:
        try:
            age = time.time() - os.path.getmtime(lock_path)
            if age > stale_seconds:
                os.remove(lock_path)
        except Exception:
            pass
        return None
    except Exception:
        return None

def _release_file_lock(lock_path: str, fd):
    try:
        if fd is not None:
            os.close(fd)
    except Exception:
        pass
    try:
        if os.path.isfile(lock_path):
            os.remove(lock_path)
    except Exception:
        pass

def _run_auto_backup_once(force=False):
    sec = _runtime_security_settings()
    if (not sec.get('auto_backup_enabled')) and (not force):
        return {'ok': False, 'skipped': 'disabled'}

    every_min = int(sec.get('auto_backup_every_min') or AUTO_BACKUP_EVERY_MIN_DEFAULT)
    keep_n = int(sec.get('auto_backup_keep') or AUTO_BACKUP_KEEP_DEFAULT)
    webhook = str(sec.get('backup_webhook') or '').strip()

    lock_fd = _acquire_file_lock(_AUTO_BACKUP_LOCK_PATH, stale_seconds=max(3600, every_min * 120))
    if lock_fd is None:
        return {'ok': True, 'skipped': 'locked'}

    try:
        latest = _latest_backup_info()
        if latest and not force:
            age_s = time.time() - float(latest.get('mtime') or 0)
            if age_s < (every_min * 60):
                return {'ok': True, 'skipped': 'not_due', 'age_seconds': int(age_s)}

        out = snapshot_db('auto')
        if not out:
            _audit_event('auto_backup', outcome='error', reason='db_missing')
            return {'ok': False, 'error': 'db_missing'}

        verify = _verify_sqlite_file(out)
        removed = _prune_backup_files(keep_n)
        meta = {
            'path': os.path.abspath(out),
            'name': os.path.basename(out),
            'size': int(os.path.getsize(out)) if os.path.isfile(out) else 0,
            'mtime': int(os.path.getmtime(out)) if os.path.isfile(out) else int(time.time()),
        }
        _audit_event(
            'auto_backup',
            outcome='ok' if verify.get('ok') else 'error',
            actor={'type': 'system', 'username': 'scheduler'},
            backup=meta['name'],
            size=meta['size'],
            integrity=verify.get('integrity'),
            removed_old=removed,
        )
        if webhook:
            _post_json_webhook_async(webhook, {
                'event': 'auto_backup',
                'ok': bool(verify.get('ok')),
                'backup': meta,
                'verify': verify,
                'removed_old': removed,
            })
        return {'ok': True, 'backup': meta, 'verify': verify, 'removed_old': removed}
    finally:
        _release_file_lock(_AUTO_BACKUP_LOCK_PATH, lock_fd)

def _auto_backup_worker():
    while True:
        try:
            _run_auto_backup_once(force=False)
        except Exception:
            pass
        time.sleep(60)

def _ensure_auto_backup_thread():
    global _AUTO_BACKUP_THREAD
    if _AUTO_BACKUP_THREAD and _AUTO_BACKUP_THREAD.is_alive():
        return
    with _AUTO_BACKUP_THREAD_LOCK:
        if _AUTO_BACKUP_THREAD and _AUTO_BACKUP_THREAD.is_alive():
            return
        _AUTO_BACKUP_THREAD = threading.Thread(target=_auto_backup_worker, name='auto-backup-worker', daemon=True)
        _AUTO_BACKUP_THREAD.start()
# Rate limiting (si flask-limiter estÃ¡ disponible)
limiter = None
if Limiter and get_remote_address:
    limiter = Limiter(get_remote_address, app=app,
                      default_limits=["200 per hour", "60 per minute"],
                      storage_uri=os.environ.get("LIMITER_STORAGE_URI") or "memory://")
def limit(rule: str):
    """Decorator de rate-limit que no rompe si flask-limiter no estÃ¡ instalado."""
    def deco(fn):
        if limiter:
            return limiter.limit(rule)(fn)
        return fn
    return deco

@app.after_request
def _set_security_headers(resp):
    try:
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        resp.headers['X-Frame-Options'] = 'DENY'
        resp.headers['X-Permitted-Cross-Domain-Policies'] = 'none'
        resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        resp.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
        # CSP conservadora para no romper el frontend y permitir Google Fonts
        resp.headers['Content-Security-Policy'] = (
            "default-src 'self'; "
            "base-uri 'self'; "
            "object-src 'none'; "
            "form-action 'self'; "
            "img-src 'self' data: https:; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com data:; "
            "script-src 'self' 'unsafe-inline' https://unpkg.com; "
            "connect-src 'self'; "
            "frame-ancestors 'none'"
        )
        # Evita que páginas, APIs y PDFs sensibles queden cacheados en el navegador.
        if request.path != '/health' and not request.path.startswith('/static/'):
            if (
                request.path.startswith(('/api/', '/print/', '/q/', '/cotizacion/', '/export/'))
                or resp.mimetype in ('text/html', 'application/json', 'application/pdf', 'text/plain')
            ):
                resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
                resp.headers['Pragma'] = 'no-cache'
                resp.headers['Expires'] = '0'
        if IS_PROD:
            resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        try:
            path = request.path or ''
            should_audit = (
                path.startswith('/api/admin/')
                or path.startswith('/api/bot/')
                or (path.startswith('/api/') and request.method in ('POST', 'PUT', 'PATCH', 'DELETE')
                    and path not in ('/api/login', '/api/logout', '/api/change_password', '/api/admin/db_backup', '/api/admin/db_restore', '/api/admin/security_events'))
            )
            if should_audit:
                _audit_event(
                    'http_request',
                    outcome='ok' if getattr(resp, 'status_code', 0) < 400 else 'error',
                    status_code=int(getattr(resp, 'status_code', 0) or 0),
                )
        except Exception:
            pass
    except Exception:
        pass
    return resp

# â”€â”€â”€ Auth helpers (roles: admin | vendedor | lector) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PUBLIC_API_WHITELIST = set(['/api/login', '/api/me'])

CSRF_EXEMPT = set(['/api/login'])

if IS_PROD and app.secret_key == _DEFAULT_SECRET_KEY:
    print('[SECURITY] WARNING: SECRET_KEY sigue con el valor por defecto en producciÃ³n.')

def _configured_bot_keys():
    raw = (os.environ.get('BOT_KEY') or '').strip()
    keys = {k.strip() for k in raw.split(',') if k.strip()}
    if not keys and not IS_PROD:
        keys.add(_DEFAULT_BOT_KEY)
    return keys

BOT_KEYS = _configured_bot_keys()
if IS_PROD and not BOT_KEYS:
    print('[SECURITY] WARNING: BOT_KEY no configurado en producciÃ³n; /api/bot/* quedarÃ¡ deshabilitado.')

def _csrf_token():
    tok = session.get('csrf_token')
    if not tok:
        tok = secrets.token_urlsafe(32)
        session['csrf_token'] = tok
    return tok

def hash_password(pw: str) -> str:
    pw = pw or ''
    if _ARGON2:
        return _ARGON2.hash(pw)
    return generate_password_hash(pw)

def verify_password(stored_hash: str, pw: str) -> bool:
    if not stored_hash:
        return False
    if stored_hash.startswith('$argon2'):
        if not _ARGON2:
            return False
        try:
            return _ARGON2.verify(stored_hash, pw or '')
        except VerifyMismatchError:
            return False
        except Exception:
            return False
    # fallback werkzeug (pbkdf2)
    return check_password_hash(stored_hash, pw or '')



def current_user():
    try:
        return session.get('user')
    except Exception:
        return None

def _current_user_db_row():
    u = current_user() or {}
    uid = u.get('id')
    if not uid:
        return None
    return query("SELECT * FROM users WHERE id=?", (uid,), one=True)

def _verify_current_user_password(pw: str) -> bool:
    row = _current_user_db_row()
    if not row:
        return False
    return verify_password(row.get('password_hash') or '', pw or '')

def _admin_reauth_is_recent(valid_for_seconds=None) -> bool:
    sec = _runtime_security_settings()
    max_age = int(sec.get('admin_reauth_seconds') or 0) if valid_for_seconds is None else int(valid_for_seconds)
    if max_age <= 0:
        return True
    try:
        ts = float(session.get('admin_reauth_at') or 0)
    except Exception:
        ts = 0
    if ts <= 0:
        return False
    return (time.time() - ts) <= max_age

def _mark_admin_reauth_ok():
    try:
        session['admin_reauth_at'] = float(time.time())
        session.modified = True
    except Exception:
        pass

def login_required(fn):
    @wraps(fn)
    def _wrap(*args, **kwargs):
        if not current_user():
            return jsonify({'ok': False, 'error': 'No autenticado'}), 401
        return fn(*args, **kwargs)
    return _wrap

def role_required(*roles):
    roles = set([r.lower() for r in roles])
    def deco(fn):
        @wraps(fn)
        def _wrap(*args, **kwargs):
            u = current_user()
            if not u:
                return jsonify({'ok': False, 'error': 'No autenticado'}), 401
            if (u.get('role') or '').lower() not in roles:
                return jsonify({'ok': False, 'error': 'Sin permisos'}), 403
            return fn(*args, **kwargs)
        return _wrap
    return deco

def _audit_mask_text(value, keep=3):
    text = str(value or '')
    if len(text) <= (keep * 2) + 2:
        return '*' * len(text)
    return f"{text[:keep]}...{text[-keep:]}"

def _audit_sanitize(value, key=''):
    key_l = str(key or '').lower()
    if isinstance(value, dict):
        return {str(k): _audit_sanitize(v, k) for k, v in value.items()}
    if isinstance(value, list):
        return [_audit_sanitize(v, key_l) for v in value]
    if isinstance(value, tuple):
        return tuple(_audit_sanitize(v, key_l) for v in value)
    if any(tok in key_l for tok in ('password', 'secret', 'token', 'csrf', 'authorization', 'cookie', 'bot_key', 'audio_b64', 'prompt', 'texto_original')):
        return '[redacted]'
    if key_l in {'chat_id', 'telefono', 'nit_cc', 'nit', 'email', 'email_cliente'}:
        return _audit_mask_text(value, keep=2)
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, (int, float, bool)) or value is None:
        return value
    if isinstance(value, str):
        return value[:250] + ('...' if len(value) > 250 else '')
    return str(value)

def _should_send_security_alert(payload: dict) -> bool:
    action = str(payload.get('action') or '').lower()
    outcome = str(payload.get('outcome') or '').lower()
    path = str(payload.get('path') or '').lower()
    if action in {'admin_ip_blocked', 'bot_ip_blocked', 'bot_key_failed', 'db_restore_reauth_failed'}:
        return True
    if action == 'login_failed':
        return True
    if action in {'db_restore', 'auto_backup'}:
        return outcome in {'ok', 'error', 'rejected'}
    if path.startswith('/api/admin/') and outcome in {'error', 'rejected'}:
        return True
    return False

def _maybe_send_security_alert(payload: dict):
    sec = _runtime_security_settings()
    webhook = str(sec.get('security_alert_webhook') or '').strip()
    if not webhook:
        return
    if not _should_send_security_alert(payload):
        return
    key = f"{payload.get('action')}|{payload.get('outcome')}"
    now = time.time()
    with _ALERT_LOCK:
        last = float(_ALERT_LAST_SENT.get(key) or 0)
        # Evita spam masivo de alertas iguales.
        if now - last < 15:
            return
        _ALERT_LAST_SENT[key] = now
    _post_json_webhook_async(webhook, {'event': 'security_alert', 'payload': payload}, timeout=5)

def _audit_event(action, outcome='ok', actor=None, **details):
    try:
        payload = {
            'ts': datetime.datetime.utcnow().replace(microsecond=0).isoformat() + 'Z',
            'action': action,
            'outcome': outcome,
        }
        try:
            payload['path'] = request.path or ''
            payload['method'] = request.method or ''
            payload['ip'] = _client_ip()
            ua = request.headers.get('User-Agent') or ''
            if ua:
                payload['ua'] = _audit_sanitize(ua, 'user_agent')
        except Exception:
            pass

        if actor is None:
            actor = current_user()
        if isinstance(actor, dict) and actor:
            payload['actor'] = {
                'type': actor.get('type') or 'user',
                'id': actor.get('id'),
                'username': actor.get('username'),
                'role': actor.get('role'),
            }
        elif actor:
            payload['actor'] = {'type': 'user', 'username': str(actor)}
        else:
            payload['actor'] = {'type': 'bot' if str(payload.get('path') or '').startswith('/api/bot/') else 'anonymous'}

        if details:
            payload['details'] = _audit_sanitize(details)

        with open(AUDIT_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(json.dumps(payload, ensure_ascii=False, separators=(',', ':')) + '\n')
        _maybe_send_security_alert(payload)
    except Exception:
        pass

def _read_audit_tail(limit=50):
    if not os.path.isfile(AUDIT_LOG_PATH):
        return []
    rows = []
    try:
        with open(AUDIT_LOG_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except Exception:
                    continue
    except Exception:
        return []
    return rows[-limit:]

_MOJIBAKE_HINTS = ('Ã', 'Â', 'â', 'ð', '\ufffd')

def _mojibake_score(text: str) -> int:
    return sum(text.count(h) for h in _MOJIBAKE_HINTS)

def _repair_mojibake_text(value: str) -> str:
    """Best-effort repair for mojibake strings (UTF-8 seen as latin-1/cp1252)."""
    if not isinstance(value, str) or not value:
        return value
    current = value
    for _ in range(2):
        if _mojibake_score(current) == 0:
            break
        candidate = None
        for enc in ('latin-1', 'cp1252'):
            try:
                fixed = current.encode(enc).decode('utf-8')
            except Exception:
                continue
            if fixed != current and _mojibake_score(fixed) < _mojibake_score(current):
                candidate = fixed
                break
        if not candidate:
            break
        current = candidate
    return current

def _repair_mojibake_obj(value):
    if isinstance(value, str):
        return _repair_mojibake_text(value)
    if isinstance(value, list):
        return [_repair_mojibake_obj(v) for v in value]
    if isinstance(value, tuple):
        return tuple(_repair_mojibake_obj(v) for v in value)
    if isinstance(value, dict):
        return {k: _repair_mojibake_obj(v) for k, v in value.items()}
    return value

def _upload_folders():
    seen = set()
    for folder in (UPLOAD_FOLDER, FALLBACK_UPLOAD_FOLDER):
        f = os.path.abspath(folder)
        if f in seen:
            continue
        seen.add(f)
        if os.path.isdir(f):
            yield f

def _resolve_upload_filepath(filename: str):
    if not filename:
        return None
    safe = os.path.basename(str(filename))
    if not safe:
        return None
    for folder in _upload_folders():
        fp = os.path.join(folder, safe)
        if os.path.isfile(fp):
            return fp
    return None

def _resolve_upload_from_url(url: str):
    raw = str(url or '').strip()
    if not raw.startswith('/uploads/products/'):
        return None
    fn = raw.split('/uploads/products/', 1)[1]
    return _resolve_upload_filepath(fn)

def _delete_primary_upload_from_url(url: str):
    raw = str(url or '').strip()
    if not raw.startswith('/uploads/products/'):
        return
    fn = os.path.basename(raw.split('/uploads/products/', 1)[1])
    if not fn:
        return
    target = os.path.join(os.path.abspath(UPLOAD_FOLDER), fn)
    if os.path.isfile(target):
        try:
            os.remove(target)
        except Exception:
            pass

# â”€â”€â”€ Imagen utils (fondo transparente simple: blanco -> alpha) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _remove_white_bg_to_png(src_path: str, dst_path: str, thr: int = 245):
    """Convierte fondo blanco/casi blanco a transparente.

    Nota: no usa IA; es un mÃ©todo prÃ¡ctico para fotos sobre fondo blanco.
    """
    im = Image.open(src_path).convert('RGBA')
    px = im.getdata()
    new = []
    for r, g, b, a in px:
        if r >= thr and g >= thr and b >= thr:
            new.append((r, g, b, 0))
        else:
            new.append((r, g, b, a))
    im.putdata(new)
    im.save(dst_path, format='PNG', optimize=True)

# â”€â”€â”€ Arranque seguro (Render/Gunicorn/Flask run) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# La BD antes se inicializaba solo en "python app.py" (bloque __main__).
# En servidores (gunicorn / flask run) eso no se ejecuta y el frontend se queda
# sin datos (y el JS se cae al intentar parsear HTML como JSON).
_DB_INIT_DONE = False
def ensure_db_ready():
    global _DB_INIT_DONE
    if _DB_INIT_DONE:
        return
    try:
        init_db()
        print(f"[DB] Usando base: {DB_PATH}")
        _DB_INIT_DONE = True
        _ensure_auto_backup_thread()
    except Exception as e:
        # No tumbar el proceso; las rutas devolverÃ¡n error controlado.
        print("[WARN] No se pudo inicializar la BD:", e)

# â”€â”€â”€ DB helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_db():
    if _USE_POSTGRES:
        if 'db' not in g:
            g.db = psycopg2.connect(_DATABASE_URL)
            g.db.autocommit = False
        return g.db
    ensure_db_ready()
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def query(sql, params=(), one=False):
    cur = get_db().execute(sql, params)
    rv = cur.fetchone() if one else cur.fetchall()
    if one:
        return _repair_mojibake_obj(dict(rv)) if rv else None
    return [_repair_mojibake_obj(dict(r)) for r in rv]

def execute(sql, params=()):
    db = get_db()
    cur = db.execute(sql, params)
    db.commit()
    return cur.lastrowid

def _enforce_ip_allowlist(scope: str, networks, actor=None):
    if not networks:
        return None
    ip = _client_ip()
    if _ip_allowed(ip, networks):
        return None
    _audit_event(f'{scope}_ip_blocked', outcome='rejected', actor=actor, reason='ip_not_allowed', ip=ip)
    return jsonify({'ok': False, 'error': 'IP no autorizada para esta ruta'}), 403

def _bot_key_from_headers() -> str:
    return (request.headers.get('X-Bot-Key') or request.headers.get('x-bot-key') or '').strip()

def _request_has_valid_bot_key() -> bool:
    return bool(BOT_KEYS) and (_bot_key_from_headers() in BOT_KEYS)

def _require_admin_reauth_for_action(data, target_action='admin_action'):
    sec = _runtime_security_settings()
    reauth_seconds = int(sec.get('admin_reauth_seconds') or 0)
    if reauth_seconds <= 0:
        return None
    if _admin_reauth_is_recent(reauth_seconds):
        return None
    data = data or {}
    pw = str(data.get('admin_password') or data.get('password') or '').strip()
    if not pw:
        _audit_event('db_restore_reauth_failed', outcome='rejected', actor=current_user(), reason='password_required', action_name=target_action)
        return jsonify({
            'ok': False,
            'error': 'Se requiere confirmacion de contrasena admin para esta accion.',
            'code': 'admin_reauth_required',
            'valid_for_seconds': reauth_seconds,
        }), 401
    if not _verify_current_user_password(pw):
        _audit_event('db_restore_reauth_failed', outcome='rejected', actor=current_user(), reason='invalid_password', action_name=target_action)
        return jsonify({
            'ok': False,
            'error': 'Contrasena admin invalida.',
            'code': 'admin_reauth_invalid',
            'valid_for_seconds': reauth_seconds,
        }), 401
    _mark_admin_reauth_ok()
    _audit_event('admin_reauth', actor=current_user(), source='inline', action_name=target_action, valid_for_seconds=reauth_seconds)
    return None


@app.before_request
def _enforce_auth_for_api():
    """Protege /api/* con sesiÃ³n y valida CSRF en mÃ©todos de escritura.
    Se dejan abiertos /api/login y /api/me (para bootstrap/login).
    """
    p = request.path or ''
    if not p.startswith('/api/'):
        return
    sec = _runtime_security_settings()
    if p.startswith('/api/admin/'):
        ip_err = _enforce_ip_allowlist('admin', sec.get('admin_ip_allowlist') or [], actor=current_user() or {'type': 'anonymous'})
        if ip_err:
            return ip_err
    if p in PUBLIC_API_WHITELIST:
        return
    # endpoints pÃºblicos por token (si en el futuro se crean en /api/public/...)
    if p.startswith('/api/public/'):
        return
    # bot endpoints se autentican con X-Bot-Key, no con sesiÃ³n
    if p.startswith('/api/bot/'):
        bot_networks = sec.get('bot_ip_allowlist') or []
        has_valid_bot_key = _request_has_valid_bot_key()
        if bot_networks and not has_valid_bot_key:
            ip_err = _enforce_ip_allowlist('bot', bot_networks, actor={'type': 'bot'})
            if ip_err:
                return ip_err
        elif bot_networks and has_valid_bot_key:
            ip = _client_ip()
            if not _ip_allowed(ip, bot_networks):
                _audit_event('bot_ip_bypass_key', actor={'type': 'bot'}, reason='valid_bot_key', ip=ip)
        return

    if not current_user():
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    # CSRF: aplica a POST/PUT/PATCH/DELETE dentro de /api/*
    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE') and (p not in CSRF_EXEMPT):
        sent = request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token') or ''
        tok = session.get('csrf_token') or ''
        if not tok or not sent or sent != tok:
            return jsonify({'ok': False, 'error': 'CSRF invÃ¡lido. Recarga e intenta de nuevo.'}), 403


# â”€â”€â”€ DB INIT (SEGURO â€” nunca DROP) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS parametros (
    clave TEXT PRIMARY KEY,
    valor TEXT
);

CREATE TABLE IF NOT EXISTS catalogo (
    id_producto     TEXT PRIMARY KEY,
    categoria       TEXT NOT NULL,
    nombre          TEXT NOT NULL,
    descripcion     TEXT,
    unidad          TEXT DEFAULT 'Und',
    precio          REAL DEFAULT 0,
    aplica_iva      INTEGER DEFAULT 0,
    pct_iva         REAL DEFAULT 0,
    inst_default    REAL DEFAULT 0,
    config_default  REAL DEFAULT 0,
    costo_unitario REAL DEFAULT 0,
    stock_qty      REAL DEFAULT 0,
    stock_min      REAL DEFAULT 0,
    imagen_url      TEXT DEFAULT '',
    activo          INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS cotizaciones (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    no_cotizacion   TEXT UNIQUE,
    fecha           TEXT DEFAULT (date('now')),
    vendedor        TEXT DEFAULT 'Admin',
    cliente         TEXT NOT NULL,
    empresa         TEXT,
    nit_cc          TEXT,
    telefono        TEXT,
    email_cliente   TEXT,
    direccion       TEXT,
    ciudad          TEXT,
    proyecto        TEXT,
    tipo_cotizacion TEXT DEFAULT 'MIXTA',
    forma_pago      TEXT DEFAULT '70% - 30%',
    anticipo_pct    REAL DEFAULT 0.70,
    descuento_pct   REAL DEFAULT 0,
    descuento_val   REAL DEFAULT 0,
    notas           TEXT,
    notas_internas  TEXT DEFAULT '',
    checklist_json  TEXT DEFAULT '',
    price_list_code TEXT DEFAULT 'PUBLICO',
    price_list_desc_pct REAL DEFAULT 0,
    public_token    TEXT DEFAULT '',
    accepted        INTEGER DEFAULT 0,
    accepted_at     TEXT DEFAULT '',
    accepted_name   TEXT DEFAULT '',
    accepted_ip     TEXT DEFAULT '',
    estado          TEXT DEFAULT 'BORRADOR',
    created_at      TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS users (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    username      TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role          TEXT NOT NULL DEFAULT 'admin',
    must_change_password INTEGER DEFAULT 0,
    created_at    TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS price_lists (
    code        TEXT PRIMARY KEY,
    name        TEXT NOT NULL,
    desc_pct    REAL DEFAULT 0,
    active      INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    cot_id          INTEGER NOT NULL REFERENCES cotizaciones(id) ON DELETE CASCADE,
    linea           INTEGER,
    id_producto     TEXT NOT NULL REFERENCES catalogo(id_producto),
    cantidad        REAL DEFAULT 1,
    precio_manual   REAL DEFAULT 0,
    inst_manual     REAL DEFAULT 0,
    cfg_manual      REAL DEFAULT 0,
    notas_item      TEXT
);

CREATE TABLE IF NOT EXISTS commands_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at  TEXT DEFAULT (datetime('now')),
    payload     TEXT,
    status      TEXT DEFAULT 'OK',
    error_msg   TEXT DEFAULT '',
    cot_id      INTEGER
);


-- INVENTARIO / KARDEX
CREATE TABLE IF NOT EXISTS movimientos_inventario (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at  TEXT DEFAULT (datetime('now')),
    id_producto TEXT NOT NULL REFERENCES catalogo(id_producto),
    bodega      TEXT DEFAULT 'PRINCIPAL',
    tipo        TEXT NOT NULL, -- ENTRADA/SALIDA/AJUSTE/RESERVA/LIBERACION
    cantidad    REAL NOT NULL,
    nota        TEXT DEFAULT '',
    ref         TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS bodegas (
    nombre TEXT PRIMARY KEY
);

CREATE TABLE IF NOT EXISTS stock_bodega (
    bodega      TEXT NOT NULL REFERENCES bodegas(nombre),
    id_producto TEXT NOT NULL REFERENCES catalogo(id_producto),
    stock_qty   REAL DEFAULT 0,
    PRIMARY KEY (bodega, id_producto)
);

-- PROVEEDORES / COMPRAS
CREATE TABLE IF NOT EXISTS proveedores (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre      TEXT NOT NULL,
    whatsapp    TEXT DEFAULT '',
    email       TEXT DEFAULT '',
    condiciones TEXT DEFAULT '',
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS compras (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    no_compra   TEXT UNIQUE,
    fecha       TEXT DEFAULT (date('now')),
    proveedor_id INTEGER REFERENCES proveedores(id),
    notas       TEXT DEFAULT '',
    total       REAL DEFAULT 0,
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS compra_items (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    compra_id   INTEGER NOT NULL REFERENCES compras(id) ON DELETE CASCADE,
    id_producto TEXT NOT NULL REFERENCES catalogo(id_producto),
    cantidad    REAL DEFAULT 1,
    costo_unit  REAL DEFAULT 0,
    subtotal    REAL DEFAULT 0
);

-- HISTORIAL DE PRECIOS/COSTOS
CREATE TABLE IF NOT EXISTS historial_precios (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at  TEXT DEFAULT (datetime('now')),
    id_producto TEXT NOT NULL REFERENCES catalogo(id_producto),
    precio      REAL DEFAULT 0,
    costo_unit  REAL DEFAULT 0,
    nota        TEXT DEFAULT ''
);

-- PAQUETES / PLANTILLAS
CREATE TABLE IF NOT EXISTS paquetes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre      TEXT NOT NULL,
    categoria   TEXT DEFAULT 'MIXTA',
    notas       TEXT DEFAULT '',
    activo      INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS paquete_items (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    paquete_id  INTEGER NOT NULL REFERENCES paquetes(id) ON DELETE CASCADE,
    id_producto TEXT NOT NULL REFERENCES catalogo(id_producto),
    cantidad    REAL DEFAULT 1
);

-- CRM / ACTIVIDADES
CREATE TABLE IF NOT EXISTS crm_actividades (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    cot_id        INTEGER NOT NULL REFERENCES cotizaciones(id) ON DELETE CASCADE,
    created_at    TEXT DEFAULT (datetime('now')),
    tipo          TEXT DEFAULT 'NOTA',
    nota          TEXT DEFAULT '',
    proxima_fecha TEXT DEFAULT ''
);
"""

PRODUCTOS = [
    ('DOM-001','DOMOTICA','Interruptor Orvibo Gris','Interruptor WiFi color gris Orvibo 2-3 circuitos','Und',280000,0,0,80000,0,'',1),
    ('DOM-002','DOMOTICA','Interruptor Orvibo Negro','Interruptor WiFi color negro Orvibo','Und',270000,0,0,80000,0,'',1),
    ('DOM-003','DOMOTICA','Pantalla Mixpad Mini','Orvibo Mixpad mini WiFi BT ZB Alexa 2 circuitos','Und',980000,0,0,0,200000,'',1),
    ('DOM-004','DOMOTICA','Pantalla Mixpad Grande','Orvibo Mixpad WiFi BT pantalla grande Alexa','Und',3250000,0,0,0,400000,'',1),
    ('DOM-005','DOMOTICA','Toma USB Tipo C','Toma corriente con puerto USB A y tipo C','Und',75000,0,0,90000,0,'',1),
    ('DOM-006','DOMOTICA','Toma GFCI Gris','Toma corriente GFCI protecciÃ³n diferencial','Und',85000,0,0,90000,0,'',1),
    ('DOM-007','DOMOTICA','Toma GFCI Negro','Toma corriente GFCI negra','Und',105000,0,0,90000,0,'',1),
    ('DOM-008','DOMOTICA','Toma Naranja','Toma corriente naranja','Und',65000,0,0,90000,0,'',1),
    ('DOM-009','DOMOTICA','Relay Micro Orvibo','Micro relay multifuncional cortinas contacto seco','Und',345000,0,0,0,0,'',1),
    ('DOM-010','DOMOTICA','Control TV Magic Dot','WiFi Remote Control Magic Dot 5V IR/RF','Und',85000,0,0,0,0,'',1),
    ('DOM-011','DOMOTICA','Hub Control Orvibo','Hub central del sistema Orvibo','Und',380000,0,0,0,150000,'',1),
    ('DOM-012','DOMOTICA','Enchufe InalÃ¡mbrico','Cargador inalÃ¡mbrico 15W 4 salidas CA USB A/C 20W','Und',450000,0,0,0,0,'',1),
    ('DOM-013','DOMOTICA','Control Universal Pro','IR y RF Control remoto universal todo en uno Hub','Und',420000,0,0,0,0,'',1),
    ('DOM-014','DOMOTICA','Dimmer WiFi D1','US Dimmer WiFi Switch D1','Und',380000,0,0,80000,0,'',1),
    ('DOM-015','DOMOTICA','Sensor Presencia Orvibo','Orvibo sensor detecciÃ³n de presencia humana','Und',290000,0,0,0,0,'',1),
    ('DOM-016','DOMOTICA','Controlador Cortinas RM4','Hub RM4pro para control de cortinas WiFi','Und',220000,0,0,80000,0,'',1),
    ('DOM-017','DOMOTICA','Alexa Echo Spot','Amazon Alexa Echo Spot con pantalla 5"','Und',340000,0,0,0,0,'',1),
    ('DOM-018','DOMOTICA','BaterÃ­a Echo Spot','Base baterÃ­a 5000mAh para Alexa Echo Spot','Und',210000,0,0,0,0,'',1),
    ('DOM-019','DOMOTICA','Cargador BaterÃ­as AA/AAA','Cargador USB alta velocidad AA/AAA','Und',75000,0,0,0,0,'',1),
    ('DOM-020','DOMOTICA','BaterÃ­as Litio Recargable','Paquete baterÃ­as litio recargables AA/AAA','Und',170000,0,0,0,0,'',1),
    ('CEK-001','CERRADURAS','Cerradura Digital Huella','Cerradura con huella digital y llaves','Und',250000,0,0,80000,0,'',1),
    ('CEK-002','CERRADURAS','Cerradura C220','Cerradura digital C220 app inteligente','Und',945000,0,0,80000,0,'',1),
    ('CEK-003','CERRADURAS','Cerradura Gerencia Venas','Cerradura reconocimiento de venas biomÃ©trico','Und',1050000,0,0,80000,0,'',1),
    ('CEK-004','CERRADURAS','BotÃ³n Salida No Touch','BotÃ³n de salida sin contacto','Und',120000,0,0,50000,0,'',1),
    ('CEK-005','CERRADURAS','Terminal Reconocimiento Facial','Terminal biomÃ©trica facial acceso','Und',650000,0,0,100000,0,'',1),
    ('CEK-006','CERRADURAS','Video Portero','Sistema de Video Portero IP completo','Servicio',4562930,0,0,150000,0,'',1),
    ('CAM-001','CCTV','CÃ¡mara 360 Interior 2K','CÃ¡mara inalÃ¡mbrica 360Â° 2K+6MP 2.4/5GHz interior','Und',880000,0,0,100000,0,'',1),
    ('CAM-002','CCTV','CÃ¡mara 360 Ojo de Pez','CÃ¡mara 360 ojo de pez WiFi interior','Und',660000,0,0,100000,0,'',1),
    ('CAM-003','CCTV','CÃ¡mara Exterior 16MP','CÃ¡mara WiFi 16MP ultra gran angular 180Â° Duo','Und',1250000,0,0,100000,0,'',1),
    ('CAM-004','CCTV','CÃ¡mara Bala 4K WiFi 6','CÃ¡mara exterior bala 4K WiFi 6 IP66','Und',380000,0,0,100000,0,'',1),
    ('CAM-005','CCTV','CÃ¡mara Lumios 2K WiFi 6','CÃ¡mara Lumios 1080P exterior foco WiFi','Und',280000,0,0,100000,0,'',1),
    ('CAM-006','CCTV','NVR 8 Canales WiFi','Sistema NVR grabaciÃ³n cÃ¡maras WiFi 8 canales','Und',1400000,0,0,0,200000,'',1),
    ('CAM-007','CCTV','NVR 12 Canales WiFi 6 2TB','Reolink NVR 12CH WiFi 6 disco 2TB 24/7','Und',1450000,0,0,0,200000,'',1),
    ('CAM-008','CCTV','CÃ¡mara de doble lente de 180Â° de 8MP','CÃ¡maras de seguridad solar 4K inalÃ¡mbricas para exteriores, cÃ¡mara de doble lente de 180Â°','Und',889000,0,0,100000,50000,'https://m.media-amazon.com/images/I/615MkukhHIL._AC_UY218_.jpg',1),
    ('RED-001','REDES','Deco BE3000 WiFi 7 Interior','TP-Link Deco BE3000 WiFi 7 unidad interior','Und',860000,0,0,50000,100000,'',1),
    ('RED-002','REDES','Deco x55 WiFi 7 M5','Deco x55 WiFi 7 M5 interior alta velocidad','Und',425000,0,0,50000,100000,'https://m.media-amazon.com/images/I/61D7vtclcsL._AC_UY218_.jpg',1),
    ('RED-003','REDES','Deco WiFi 6 Doble Banda','Unidad extensora WiFi 6 doble banda mesh','Und',290000,0,0,50000,100000,'',1),
    ('RED-004','REDES','Switch Internet','Switch de red ethernet 8 puertos','Und',140000,0,0,50000,0,'',1),
    ('RED-005','REDES','Soporte Pared Deco Interior','Soporte pared para Decos internet interior','Und',180000,0,0,0,0,'https://m.media-amazon.com/images/I/313fppQuz6L._AC_SR250,250_QL65_.jpg',1),
    ('RED-006','REDES','Soporte Deco Interior (small)','Soporte pequeÃ±o para Decos internet interior','Und',75000,0,0,0,0,'',1),
    ('RED-007','REDES','Extensor Red Internet','Extensor de red internet WiFi','Und',75000,0,0,50000,0,'',1),
    ('RED-008','REDES','Soporte para Antena Starlink','Starlink - Mini soporte para versiÃ³n 2025, adaptador redondo de aleaciÃ³n de aluminio ajustable de 360Â°','Und',938000,0,0,100000,50000,'https://m.media-amazon.com/images/I/41qa9-id8tL._AC_SR250,250_QL65_.jpg',1),
    ('AV-001','AUDIOVISUAL','Amplificador WiiM','WiiM Amp multihabitaciÃ³n AirPlay Alexa','Und',2550000,0,0,0,150000,'',1),
    ('AV-002','AUDIOVISUAL','Amplificador WiiM Amp Ultra','WiiM Amp Ultra 100W voz AirPlay Spotify','Und',3150000,0,0,0,200000,'',1),
    ('AV-003','AUDIOVISUAL','Parlante Klipsch Empotrado','Parlante empotrado techo Klipsch 5.25"','Und',770000,0,0,150000,0,'',1),
    ('AV-004','AUDIOVISUAL','Parlante Pyle Sala Junta','Parlante Pyle para sala de juntas','Und',850000,0,0,150000,0,'',1),
    ('AV-005','AUDIOVISUAL','Parlante Sala Central','Parlante sala central','Und',535000,0,0,150000,0,'',1),
    ('AV-006','AUDIOVISUAL','Parlante Exterior Impermeable','Altavoz exterior 4.5" 80W impermeable IP65','Und',657000,0,0,150000,0,'',1),
    ('AV-007','AUDIOVISUAL','Bocina Techo 5.25','Bocina techo 5.25" audio ambiente','Und',568000,0,0,150000,0,'',1),
    ('AV-008','AUDIOVISUAL','Subwoofer Polk','Subwoofer 10" clase D 100W Dolby Atmos','Und',2490000,0,0,0,0,'',1),
    ('AV-009','AUDIOVISUAL','Receptor AV 7.2 Canales','Receptor AV 7.2ch 80W 8K HDMI Dolby HEOS','Und',4350000,0,0,0,200000,'',1),
    ('AV-010','AUDIOVISUAL','Receptor Sala de Juntas','Receptor estÃ©reo M19-BT amplificador alta fidelidad','Und',0,0,0,0,200000,'',1),
    ('AV-011','AUDIOVISUAL','MicrÃ³fono Conferencia 8en1','MicrÃ³fono altavoz conferencia 8 en 1','Und',745000,0,0,0,0,'',1),
    ('AV-012','AUDIOVISUAL','TelÃ³n Proyector 100"','Pantalla proyector motorizada 100 pulgadas','Und',1050000,0,0,200000,0,'',1),
    ('AV-013','AUDIOVISUAL','Proyector Optoma 4K','Optoma proyector 4K HD alto brillo 3D HDR','Und',5325100,0,0,200000,0,'',1),
    ('AV-014','AUDIOVISUAL','Ascensor Proyector','Accesorio ascensor motorizado para proyector','Und',3870000,0,0,200000,0,'',1),
    ('AV-015','AUDIOVISUAL','Cable RCA 50pies','Cable subwoofer 50ft RCA a RCA','Und',265000,0,0,0,0,'',1),
    ('AV-016','AUDIOVISUAL','Cable HDMI 8K 25ft','Cable HDMI 8K 48Gbps versiÃ³n 2.1','Und',280000,0,0,0,0,'',1),
    ('AV-017','AUDIOVISUAL','Fire TV 4K','Amazon Fire TV Stick 4K Max','Und',205000,0,0,0,0,'',1),
    ('OTR-001','OTROS','Tapita Cat6 Negra','Placa pared Ethernet negra Cat6 certificaciÃ³n UL','Und',29700,0,0,0,0,'',1),
    ('MO-001','SERVICIOS','InstalaciÃ³n CÃ¡maras','InstalaciÃ³n fÃ­sica de cÃ¡mara (incluye fijaciÃ³n)','Und',100000,0,0,0,0,'',1),
    ('MO-002','SERVICIOS','InstalaciÃ³n Interruptores','InstalaciÃ³n interruptor domÃ³tico','Und',80000,0,0,0,0,'',1),
    ('MO-003','SERVICIOS','InstalaciÃ³n Tomas','InstalaciÃ³n toma corriente','Und',90000,0,0,0,0,'',1),
    ('MO-004','SERVICIOS','InstalaciÃ³n Decos/Red','InstalaciÃ³n dispositivo de red/deco','Und',50000,0,0,0,0,'',1),
    ('MO-005','SERVICIOS','ConfiguraciÃ³n Sistema','ConfiguraciÃ³n y programaciÃ³n del sistema completo','Servicio',0,0,0,0,0,'',1),
    ('MO-006','SERVICIOS','Montaje y Cableado Completo','Montaje, cableado, instalaciÃ³n y configuraciÃ³n total','Servicio',0,0,0,0,0,'',1),
]

PARAMS_DATA = [
    ('empresa','RC DOMOTIC'),('nit','1102809561'),
    ('direccion','Calle 31-8-88'),('ciudad','Sincelejo / Sucre'),
    ('telefono','3123042156'),('email',''),('web',''),
    ('contacto','RAUL CUELLO'),('banco','Bancolombia Ahorros'),
    ('cuenta','506-826941-20'),('titular','RAUL CUELLO GONZALEZ'),
    ('iva_general','0.19'),('vigencia_dias','30'),
    ('forma_pago_default','70% - 30%'),('anticipo_default','0.70'),
    ('prefijo_cot','A05'),('consecutivo','60'),
    ('garantia','6 meses en mano de obra / segun fabricante en equipos'),
    ('plazo_entrega','A convenir segun proyecto'),
    ('condiciones','Precios en COP incluyen equipos. La mano de obra se detalla por separado.'),
    ('logo_path', '/static/brand_logo.png'),
    ('watermark_path', '/static/watermark.png'),
    ('brand_primary', '#0F0F0F'),
    ('brand_surface', '#111111'),
    ('brand_accent', '#25D366'),
    # Seguridad configurable desde la app (si no existe variable de entorno).
    ('security_admin_ip_allowlist', ''),
    ('security_bot_ip_allowlist', ''),
    ('security_admin_reauth_seconds', str(ADMIN_REAUTH_SECONDS_DEFAULT)),
    ('security_auto_backup_enabled', '1' if AUTO_BACKUP_ENABLED_DEFAULT else '0'),
    ('security_auto_backup_every_min', str(AUTO_BACKUP_EVERY_MIN_DEFAULT)),
    ('security_auto_backup_keep', str(AUTO_BACKUP_KEEP_DEFAULT)),
    ('security_backup_max_age_hours', str(SECURITY_BACKUP_MAX_AGE_HOURS_DEFAULT)),
    ('security_alert_webhook', ''),
    ('security_backup_webhook', ''),
]

RECOVERY_GILBERTO_QUOTE = {
    'no_cotizacion': 'A05-00060',
    'fecha': '2026-04-06',
    'created_at': '2026-04-06 04:45:00',
    'vendedor': 'Admin',
    'cliente': 'Sr. Gilberto',
    'empresa': 'Finca',
    'nit_cc': '0',
    'telefono': '0',
    'email_cliente': '',
    'direccion': '',
    'ciudad': '',
    'proyecto': 'Finca',
    'tipo_cotizacion': 'MIXTA',
    'forma_pago': '70% - 30%',
    'anticipo_pct': 0.70,
    'anticipo_val_manual': 0,
    'abonado_val': 0,
    'descuento_pct': 0,
    'descuento_val': 0,
    'notas': '',
    'notas_internas': '',
    'checklist_json': json.dumps([
        {"label":"Visita técnica", "done": False},
        {"label":"Medidas confirmadas", "done": False},
        {"label":"Cliente aprueba equipos", "done": False},
        {"label":"Anticipo recibido", "done": False},
        {"label":"Programación agenda", "done": False},
        {"label":"Instalación completa", "done": False},
        {"label":"Entrega y capacitación", "done": False},
    ], ensure_ascii=False),
    'price_list_code': 'PUBLICO',
    'price_list_desc_pct': 0.0,
    'public_token': '',
    'public_expires_at': '',
    'public_revoked': 0,
    'accepted': 0,
    'accepted_at': '',
    'accepted_name': '',
    'accepted_ip': '',
    'estado': 'BORRADOR',
    'etapa': 'COTIZADA',
    'items': [
        {'linea': 1, 'id_producto': 'CAM-008', 'cantidad': 1, 'precio_manual': 1550000, 'inst_manual': 200000, 'cfg_manual': 50000, 'notas_item': ''},
        {'linea': 2, 'id_producto': 'RED-005', 'cantidad': 1, 'precio_manual': 98000,  'inst_manual': 80000,  'cfg_manual': 50000, 'notas_item': ''},
        {'linea': 3, 'id_producto': 'RED-002', 'cantidad': 2, 'precio_manual': 425000, 'inst_manual': 150000, 'cfg_manual': 50000, 'notas_item': ''},
        {'linea': 4, 'id_producto': 'RED-008', 'cantidad': 1, 'precio_manual': 165000, 'inst_manual': 60000,  'cfg_manual': 50000, 'notas_item': ''},
    ],
}

def init_db():
    fresh = not os.path.exists(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    # Usuario admin por defecto (para BD nueva o migrada). Puedes sobreescribir en Render con env vars.
    admin_user = os.environ.get('DEFAULT_ADMIN_USER', 'admin')
    _ap_env = os.environ.get('DEFAULT_ADMIN_PASS')
    admin_pass = _ap_env or 'admin123'
    must_change_admin = 1 if (_ap_env is None or admin_pass == 'admin123') else 0
    if fresh:
        conn.executemany(
            "INSERT OR IGNORE INTO catalogo (id_producto,categoria,nombre,descripcion,unidad,precio,aplica_iva,pct_iva,inst_default,config_default,imagen_url,activo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            [_repair_mojibake_obj(p) for p in PRODUCTOS]
        )
        conn.executemany(
            "INSERT OR IGNORE INTO parametros (clave, valor) VALUES (?,?)",
            [_repair_mojibake_obj(p) for p in PARAMS_DATA]
        )
        # En BD fresca: crear admin
        try:
            conn.execute("INSERT OR IGNORE INTO users(username,password_hash,role,must_change_password) VALUES (?,?,?,?)",
                         (admin_user, hash_password(admin_pass), 'admin', must_change_admin))
        except Exception:
            pass
        # Listas de precios por defecto
        defaults = [
            ('PUBLICO', 'PÃºblico', 0.0, 1),
            ('FRECUENTE', 'Cliente frecuente', 0.05, 1),
            ('ALIADO', 'Aliado / referido', 0.10, 1),
            ('MAYORISTA', 'Mayorista / constructor', 0.15, 1),
        ]
        try:
            conn.executemany("INSERT OR IGNORE INTO price_lists(code,name,desc_pct,active) VALUES (?,?,?,?)", defaults)
        except Exception:
            pass
        print("âœ“ Base de datos inicializada con 65 productos")

    # Si la BD ya existÃ­a (o viene de una versiÃ³n anterior), puede no tener usuarios.
    # Aseguramos al menos 1 admin para que el login no quede bloqueado.
    try:
        n = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if int(n) == 0:
            conn.execute("INSERT INTO users(username,password_hash,role,must_change_password) VALUES (?,?,?,?)",
                         (admin_user, hash_password(admin_pass), 'admin', must_change_admin))
            print("  âœ“ Admin creado automÃ¡ticamente (BD existente)")
    except Exception:
        pass
    # Migraciones seguras
    safe_cols = [
        ("users", "must_change_password", "INTEGER DEFAULT 0"),
        ("catalogo", "costo_unitario", "REAL DEFAULT 0"),
        ("catalogo", "stock_qty", "REAL DEFAULT 0"),
        ("catalogo", "stock_min", "REAL DEFAULT 0"),
        ("cotizaciones", "seguimiento_fecha", "TEXT DEFAULT ''"),
        ("cotizaciones", "proxima_accion", "TEXT DEFAULT ''"),
        ("cotizaciones", "motivo_perdida", "TEXT DEFAULT ''"),
        ("cotizaciones", "etapa", "TEXT DEFAULT 'LEAD'"),
        ("cotizaciones", "notas_internas", "TEXT DEFAULT ''"),
        ("cotizaciones", "checklist_json", "TEXT DEFAULT ''"),
        ("cotizaciones", "price_list_code", "TEXT DEFAULT 'PUBLICO'"),
        ("cotizaciones", "price_list_desc_pct", "REAL DEFAULT 0"),
        ("cotizaciones", "public_token", "TEXT DEFAULT ''"),
        ("cotizaciones", "public_expires_at", "TEXT DEFAULT ''"),
        ("cotizaciones", "public_revoked", "INTEGER DEFAULT 0"),
        ("cotizaciones", "accepted", "INTEGER DEFAULT 0"),
        ("cotizaciones", "accepted_at", "TEXT DEFAULT ''"),
        ("cotizaciones", "accepted_name", "TEXT DEFAULT ''"),
        ("cotizaciones", "accepted_ip", "TEXT DEFAULT ''"),
        # anticipo/abonos (cliente)
        ("cotizaciones", "anticipo_val_manual", "REAL DEFAULT 0"),
        ("cotizaciones", "abonado_val", "REAL DEFAULT 0"),
    ]
    for t, c, td in safe_cols:
        try: conn.execute(f"ALTER TABLE {t} ADD COLUMN {c} {td}"); print(f"  âœ“ MigraciÃ³n: {t}.{c}")
        except: pass
    for k, v in PARAMS_DATA:
        conn.execute("INSERT OR IGNORE INTO parametros (clave, valor) VALUES (?,?)", (k, v))

    # Gastos por proyecto (ligados a una cotizaciÃ³n)
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS gastos_proyecto (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cot_id INTEGER NOT NULL,
                fecha TEXT DEFAULT '',
                concepto TEXT NOT NULL,
                valor REAL NOT NULL,
                nota TEXT DEFAULT '',
                created_at TEXT DEFAULT '',
                FOREIGN KEY(cot_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
            )
        """)
    except Exception:
        pass

    # Gastos por proyecto (asociados a una cotizaciÃ³n)
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS gastos_proyecto (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cot_id INTEGER NOT NULL,
            fecha TEXT DEFAULT '',
            concepto TEXT NOT NULL,
            valor REAL NOT NULL DEFAULT 0,
            notas TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            FOREIGN KEY(cot_id) REFERENCES cotizaciones(id)
        )""")
    except Exception:
        pass

    # Bodega por defecto + sincronizar stock_bodega
    try:
        conn.execute("INSERT OR IGNORE INTO bodegas(nombre) VALUES ('PRINCIPAL')")
        # crear filas stock_bodega para todos los productos
        conn.execute("INSERT OR IGNORE INTO stock_bodega(bodega,id_producto,stock_qty) "
                     "SELECT 'PRINCIPAL', id_producto, COALESCE(stock_qty,0) FROM catalogo")
    except Exception:
        pass
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bot_chat_context (
                chat_id TEXT PRIMARY KEY,
                cot_id INTEGER NOT NULL,
                mode TEXT DEFAULT 'editar',
                created_at TEXT DEFAULT '',
                updated_at TEXT DEFAULT '',
                FOREIGN KEY(cot_id) REFERENCES cotizaciones(id) ON DELETE CASCADE
            )
        """)
    except Exception:
        pass
    # Si hay archivos locales en uploads/products, asegura imagen_url en catalogo.
    _sync_catalog_images_from_uploads(conn)
    _sync_amazon_catalog_images(conn)
    # Rescate de la cotización A05-00060 de Sr. Gilberto si la base quedó vacía o la perdió.
    try:
        _seed_gilberto_quote(conn)
    except Exception as e:
        print("[WARN] No se pudo sembrar la cotización de rescate:", e)
    _sync_cotizacion_consecutivo(conn)
    conn.commit()
    conn.close()

def _sync_cotizacion_consecutivo(conn):
    try:
        rows = conn.execute("SELECT no_cotizacion FROM cotizaciones").fetchall()
        max_seq = 0
        for (no_cot,) in rows:
            m = re.search(r'-(\d+)$', str(no_cot or '').strip())
            if m:
                max_seq = max(max_seq, int(m.group(1)))
        if max_seq <= 0:
            return
        current = conn.execute("SELECT valor FROM parametros WHERE clave='consecutivo'").fetchone()
        cur_seq = int(current[0]) if current and str(current[0]).strip().isdigit() else 0
        if cur_seq <= max_seq:
            conn.execute("UPDATE parametros SET valor=? WHERE clave='consecutivo'", (str(max_seq + 1),))
    except Exception as e:
        print("[WARN] No se pudo sincronizar consecutivo de cotizaciones:", e)

def _seed_gilberto_quote(conn):
    exists = conn.execute(
        "SELECT id FROM cotizaciones WHERE no_cotizacion=?",
        (RECOVERY_GILBERTO_QUOTE['no_cotizacion'],)
    ).fetchone()
    if exists:
        return

    fields = [
        'no_cotizacion', 'fecha', 'vendedor', 'cliente', 'empresa', 'nit_cc',
        'telefono', 'email_cliente', 'direccion', 'ciudad', 'proyecto',
        'tipo_cotizacion', 'forma_pago', 'anticipo_pct', 'anticipo_val_manual',
        'abonado_val', 'descuento_pct', 'descuento_val', 'notas', 'notas_internas',
        'checklist_json', 'price_list_code', 'price_list_desc_pct', 'public_token',
        'public_expires_at', 'public_revoked', 'accepted', 'accepted_at',
        'accepted_name', 'accepted_ip', 'estado', 'created_at', 'etapa'
    ]
    values = [RECOVERY_GILBERTO_QUOTE.get(f, '') for f in fields]
    conn.execute(
        f"INSERT INTO cotizaciones ({','.join(fields)}) VALUES ({','.join(['?'] * len(fields))})",
        values,
    )
    cot_id = conn.execute("SELECT id FROM cotizaciones WHERE no_cotizacion=?", (RECOVERY_GILBERTO_QUOTE['no_cotizacion'],)).fetchone()[0]
    for item in RECOVERY_GILBERTO_QUOTE['items']:
        conn.execute(
            """INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item)
               VALUES (?,?,?,?,?,?,?,?)""",
            (cot_id, item['linea'], item['id_producto'], item['cantidad'],
             item['precio_manual'], item['inst_manual'], item['cfg_manual'], item.get('notas_item', ''))
        )
    # El siguiente consecutivo debe ser mayor al número rescatado.
    conn.execute("INSERT OR REPLACE INTO parametros (clave, valor) VALUES ('consecutivo', ?)", ('61',))
    print("  ✓ Cotización de rescate A05-00060 sembrada")

# â”€â”€â”€ CÃLCULOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def calcular_item(item, prod, price_list_desc_pct: float = 0.0):
    """Calcula montos del Ã­tem. Si no hay precio_manual, aplica descuento de lista."""
    pl = float(price_list_desc_pct or 0)
    base_precio = prod['precio']
    if item['precio_manual'] and item['precio_manual'] > 0:
        precio_final = item['precio_manual']
    else:
        precio_final = base_precio * (1 - pl)
    # IVA SOLO sobre productos (nunca inst/cfg)
    iva_monto = round(precio_final * item['cantidad'] * prod['pct_iva'], 0) if prod['aplica_iva'] else 0
    inst_final = item['inst_manual'] if item['inst_manual'] > 0 else prod['inst_default']
    cfg_final  = item['cfg_manual']  if item['cfg_manual']  > 0 else prod['config_default']
    subtotal   = round(precio_final * item['cantidad'], 0)
    total_item = round(subtotal + iva_monto + inst_final * item['cantidad'] + cfg_final, 0)
    return {'precio_final': precio_final, 'iva_monto': iva_monto,
            'inst_final': inst_final, 'cfg_final': cfg_final,
            'subtotal': subtotal, 'total_item': total_item}

def calcular_cotizacion(cot_id):
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot: return {}
    pl = float(cot.get('price_list_desc_pct') or 0)
    items = query("""SELECT i.*, c.precio, c.aplica_iva, c.pct_iva,
                     c.inst_default, c.config_default, c.nombre
                     FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                     WHERE i.cot_id=?""", (cot_id,))
    total_bruto = 0
    for it in items:
        calc = calcular_item(it, it, pl)
        total_bruto += calc['total_item']
    descuento = cot['descuento_val'] if cot['descuento_val'] > 0 else round(total_bruto * cot['descuento_pct'], 0)
    total_final  = total_bruto - descuento
    # Anticipo: por defecto %; si hay valor manual, lo respetamos ("50% pero dio X")
    anticipo_calc = round(total_final * float(cot.get('anticipo_pct') or 0), 0)
    avm = float(cot.get('anticipo_val_manual') or 0)
    anticipo_val = round(avm, 0) if avm and avm > 0 else anticipo_calc

    # Abonado: lo que el cliente ya pagÃ³ (puede ser igual al anticipo o mÃ¡s)
    abonado_val = float(cot.get('abonado_val') or 0)
    if abonado_val < 0:
        abonado_val = 0

    # Saldo: por defecto es total - anticipo; si ya hay abonado, usamos total - abonado
    saldo_val = total_final - (abonado_val if abonado_val > 0 else anticipo_val)
    if saldo_val < 0:
        saldo_val = 0
    return {
        'total_bruto': total_bruto,
        'descuento': descuento,
        'total_final': total_final,
        'anticipo_val': anticipo_val,
        'abonado_val': round(abonado_val, 0),
        'saldo_val': round(saldo_val, 0)
    }

CATEGORY_LABELS = {
    'AUDIOVISUAL': 'SONIDO / AUDIOVISUAL',
    'CCTV': 'CAMARAS / CCTV',
    'CERRADURAS': 'CERRADURAS',
    'DOMOTICA': 'DOMOTICA',
    'OTROS': 'OTROS',
    'REDES': 'REDES',
    'SERVICIOS': 'SERVICIOS',
}

CATEGORY_ORDER = ['DOMOTICA', 'CCTV', 'AUDIOVISUAL', 'REDES', 'CERRADURAS', 'OTROS', 'SERVICIOS']

QUOTE_ACCENT_SOFT = '#B7C4BB'
QUOTE_ACCENT_SOFT_BG = '#EDF2EE'
QUOTE_ACCENT_SOFT_TEXT = '#2F4137'

def agrupar_items_por_categoria(items, total_key='total_item'):
    ordered_groups = []
    groups = {}
    for it in items:
        cat = (it.get('categoria') or 'OTROS').upper().strip() or 'OTROS'
        if cat not in groups:
            group = {
                'categoria': cat,
                'label': CATEGORY_LABELS.get(cat, cat.replace('_', ' ').title()),
                'items': [],
                'subtotal': 0.0,
            }
            groups[cat] = group
            ordered_groups.append(group)
        groups[cat]['items'].append(it)
        groups[cat]['subtotal'] += float(it.get(total_key) or 0)
    ordered_groups.sort(key=lambda g: (CATEGORY_ORDER.index(g['categoria']) if g['categoria'] in CATEGORY_ORDER else 999, g['label']))
    for g in ordered_groups:
        g['subtotal'] = round(g['subtotal'], 0)
    return ordered_groups

def preparar_presentacion_cotizacion(items):
    display_items = []
    productos_total = 0.0
    inst_total = 0.0
    cfg_total = 0.0
    for it in items:
        row = dict(it)
        row['display_total'] = round(float(row.get('subtotal') or 0) + float(row.get('iva_monto') or 0), 0)
        row['inst_line_total'] = round(float(row.get('inst_final') or 0) * float(row.get('cantidad') or 0), 0)
        row['cfg_line_total'] = round(float(row.get('cfg_final') or 0), 0)
        productos_total += row['display_total']
        inst_total += row['inst_line_total']
        cfg_total += row['cfg_line_total']
        display_items.append(row)

    service_items = []
    if round(inst_total, 0) > 0:
        service_items.append({'label': 'Instalacion total', 'total': round(inst_total, 0)})
    if round(cfg_total, 0) > 0:
        service_items.append({'label': 'Configuracion total', 'total': round(cfg_total, 0)})

    return {
        'items': display_items,
        'grouped_items': agrupar_items_por_categoria(display_items, total_key='display_total'),
        'service_items': service_items,
        'productos_total': round(productos_total, 0),
        'servicios_total': round(inst_total + cfg_total, 0),
    }

def next_no_cotizacion():
    p = query("SELECT valor FROM parametros WHERE clave='prefijo_cot'", one=True)
    c = query("SELECT valor FROM parametros WHERE clave='consecutivo'", one=True)
    prefix = p['valor'] if p else 'A05'
    num = int(c['valor']) if c else 60
    no_cot = f"{prefix}-{num:05d}"
    execute("UPDATE parametros SET valor=? WHERE clave='consecutivo'", (num + 1,))
    return no_cot

def ensure_public_token(cot_id:int):
    cot = query("SELECT public_token, public_expires_at, public_revoked FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return None
    if int(cot.get('public_revoked') or 0) == 1:
        return None
    tok = (cot.get('public_token') or '').strip()
    exp = (cot.get('public_expires_at') or '').strip()

    now = datetime.datetime.utcnow()
    # si existe expiraciÃ³n y ya pasÃ³, rotamos token
    expired = False
    if exp:
        try:
            exp_dt = datetime.datetime.strptime(exp, '%Y-%m-%d %H:%M:%S')
            if now > exp_dt:
                expired = True
        except Exception:
            expired = True

    if tok and not expired:
        return tok

    # token fuerte
    tok = secrets.token_urlsafe(32)
    exp_dt = now + datetime.timedelta(days=45)
    execute("UPDATE cotizaciones SET public_token=?, public_expires_at=?, public_revoked=0 WHERE id=?",
            (tok, exp_dt.strftime('%Y-%m-%d %H:%M:%S'), cot_id))
    return tok

    tok = uuid.uuid4().hex[:18]
    execute("UPDATE cotizaciones SET public_token=? WHERE id=?", (tok, cot_id))
    return tok

# â”€â”€â”€ MÃRGENES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def calcular_margenes(cot_id):
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot: return None
    pl = float(cot.get('price_list_desc_pct') or 0)
    items_raw = query("""SELECT i.*, c.precio, c.aplica_iva, c.pct_iva, c.inst_default,
               c.config_default, c.nombre, c.imagen_url,
               COALESCE(c.costo_unitario, 0) as costo_unitario
        FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
        WHERE i.cot_id=? ORDER BY i.linea""", (cot_id,))
    dp = float(cot['descuento_pct']); dv = float(cot['descuento_val'])
    bp = 0; bs = 0; iv = 0; cp = 0; det = []
    costos_ok = False
    for it in items_raw:
        calc = calcular_item(it, it, pl)
        q = float(it['cantidad']); pf = calc['precio_final']
        inst = calc['inst_final']; cfg = calc['cfg_final']
        co = float(it['costo_unitario'])
        if co > 0:
            costos_ok = True
        bp += pf * q; bs += inst * q + cfg
        ci = co * q; cp += ci; iv += calc['iva_monto']
        det.append({'id_producto': it['id_producto'], 'nombre': it['nombre'],
                    'cantidad': q, 'precio_unitario': pf, 'costo_unitario': co,
                    'costo_item': round(ci), 'venta_item': round(calc['total_item'])})
    bt = bp + bs
    if dv > 0: dm = dv; dr = dm / bt if bt > 0 else 0
    else: dm = round(bt * dp); dr = dp
    tots = calcular_cotizacion(cot_id)
    tf = tots['total_final']; un = tf - iv - cp; bsi = tf - iv

    # Gastos adicionales del proyecto (eg: materiales extra, transportes, terceros)
    try:
        g = query("SELECT COALESCE(SUM(valor),0) AS s FROM gastos_proyecto WHERE cot_id=?", (cot_id,), one=True)
        total_gastos = float(g['s']) if g else 0.0
    except Exception:
        total_gastos = 0.0
    un_post = un - total_gastos
    bpd = bp * (1 - dr)
    mni = (un_post / bsi * 100) if bsi > 0 else 0
    mst = (un_post / tf * 100) if tf > 0 else 0
    msp = ((bpd - cp) / bpd * 100) if bpd > 0 else 0
    mk = (un_post / cp * 100) if cp > 0 else 0
    for d in det:
        vn = d['venta_item'] * (1 - dr)
        d['utilidad_item'] = round(vn - d['costo_item'])
        d['margen_item'] = round((d['utilidad_item'] / vn * 100) if vn > 0 else 0, 1)
    return {'base_total_bruta': round(tots['total_bruto']), 'descuento': round(tots['descuento']),
            'total_final': round(tf), 'iva_total': round(iv), 'base_sin_iva': round(bsi),
            'costo_productos_total': round(cp),
            'gastos_proyecto_total': round(total_gastos),
            'utilidad_neta': round(un_post),
            'utilidad_antes_gastos': round(un),
            'margen_neto_sin_iva': round(mni, 2), 'margen_sobre_total': round(mst, 2),
            'margen_solo_producto': round(msp, 2), 'markup': round(mk, 2),
            'costos_ok': costos_ok,
            'items': det}

# â”€â”€â”€ IMÃGENES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def allowed_file(fn):
    return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def _sync_catalog_images_from_uploads(conn):
    """
    Sincroniza imagen_url en catalogo usando archivos locales en uploads/products.
    Regla de nombre esperada: <ID_PRODUCTO>_*.jpg|png|webp...
    Preferencia: amazon > local > catalogo > otros.
    """
    try:
        folders = list(_upload_folders())
        if not folders:
            return 0
        mapping = {}

        def _img_priority(fn: str) -> int:
            low = fn.lower()
            if '_amazon' in low:
                return 3
            if '_local' in low:
                return 2
            if '_catalogo' in low:
                return 1
            return 0

        primary_abs = os.path.abspath(UPLOAD_FOLDER)
        for folder in folders:
            is_primary = os.path.abspath(folder) == primary_abs
            for fn in sorted(os.listdir(folder)):
                fpath = os.path.join(folder, fn)
                if not os.path.isfile(fpath):
                    continue
                low = fn.lower()
                if not low.endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.svg')):
                    continue
                m = re.match(r'^([A-Za-z]{2,5}-\d{3})_', fn)
                if not m:
                    continue
                pid = m.group(1).upper()
                score = (_img_priority(fn), 1 if is_primary else 0)
                current = mapping.get(pid)
                if (not current) or (score > current[0]):
                    mapping[pid] = (score, fn)
        if not mapping:
            return 0

        rows = conn.execute("SELECT id_producto, imagen_url FROM catalogo").fetchall()
        updates = 0
        for row in rows:
            pid = str(row[0] or '').upper()
            cur = str(row[1] or '').strip()
            fallback_data = mapping.get(pid)
            if not fallback_data:
                continue
            fallback_fn = fallback_data[1]
            fallback_url = f"/uploads/products/{fallback_fn}"
            must_update = False
            if (not cur) or (cur.upper() == 'SIN_IMG'):
                must_update = True
            elif cur.startswith('/uploads/products/'):
                cur_fn = cur.split('/uploads/products/', 1)[1]
                cur_path = _resolve_upload_filepath(cur_fn)
                if not cur_path:
                    must_update = True
                else:
                    if _img_priority(fallback_fn) > _img_priority(cur_fn):
                        must_update = True
            if must_update:
                conn.execute("UPDATE catalogo SET imagen_url=? WHERE id_producto=?", (fallback_url, pid))
                updates += 1
        if updates:
            conn.commit()
            print(f"IMG_SYNC: {updates} imagenes actualizadas desde uploads/products")
        return updates
    except Exception as e:
        print("[WARN] No se pudo sincronizar imagenes de catalogo:", e)
        return 0

def _sync_amazon_catalog_images(conn):
    """
    Rellena con fotos Amazon solo los productos que estamos corrigiendo
    en esta tanda. No toca precios, nombres ni otras columnas.
    """
    overrides = {
        'CAM-008': 'https://m.media-amazon.com/images/I/615MkukhHIL._AC_UY218_.jpg',
        'RED-002': 'https://m.media-amazon.com/images/I/61D7vtclcsL._AC_UY218_.jpg',
        'RED-005': 'https://m.media-amazon.com/images/I/313fppQuz6L._AC_SR250,250_QL65_.jpg',
        'RED-008': 'https://m.media-amazon.com/images/I/41qa9-id8tL._AC_SR250,250_QL65_.jpg',
    }
    try:
        updates = 0
        for pid, url in overrides.items():
            current = conn.execute("SELECT imagen_url FROM catalogo WHERE id_producto=?", (pid,)).fetchone()
            if not current:
                continue
            cur_url = str(current[0] or '').strip()
            if cur_url != url:
                conn.execute("UPDATE catalogo SET imagen_url=? WHERE id_producto=?", (url, pid))
                updates += 1
        if updates:
            conn.commit()
            print(f"IMG_SYNC: {updates} imagenes Amazon aplicadas")
        return updates
    except Exception as e:
        print("[WARN] No se pudo aplicar fotos Amazon:", e)
        return 0

@app.route('/uploads/products/<path:filename>')
def serve_upload(filename):
    fp = _resolve_upload_filepath(filename)
    if not fp:
        abort(404)
    return send_from_directory(os.path.dirname(fp), os.path.basename(fp))

@app.post('/api/catalogo/sync_images')
@role_required('admin')

def sync_catalog_images():
    """Admin: fuerza sync de imagenes desde uploads/products."""
    try:
        conn = sqlite3.connect(DB_PATH)
        n = _sync_catalog_images_from_uploads(conn)
        return jsonify({'ok': True, 'updated': n})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/catalogo/<id_producto>/imagen')
@role_required('admin')
def upload_product_image(id_producto):
    if 'file' not in request.files: return jsonify({'error': 'No se enviÃ³ archivo'}), 400
    f = request.files['file']
    if not f or not allowed_file(f.filename): return jsonify({'error': 'Archivo no permitido'}), 400
    ext = f.filename.rsplit('.', 1)[1].lower()
    fname = f"{id_producto}_{uuid.uuid4().hex[:8]}.{ext}"
    path = os.path.join(UPLOAD_FOLDER, fname)
    old = query("SELECT imagen_url FROM catalogo WHERE id_producto=?", (id_producto,), one=True)
    if old and old['imagen_url']:
        _delete_primary_upload_from_url(old['imagen_url'])
    f.save(path)

    # opcional: quitar fondo (fotos con fondo blanco)
    if request.args.get('remove_bg') == '1':
        png_name = f"{id_producto}_{uuid.uuid4().hex[:8]}_cut.png"
        png_path = os.path.join(UPLOAD_FOLDER, png_name)
        try:
            _remove_white_bg_to_png(path, png_path)
            # si se generÃ³ png, usamos ese y borramos el original
            try:
                os.remove(path)
            except Exception:
                pass
            fname = png_name
            path = png_path
        except Exception as e:
            print('[WARN] remove_bg failed:', e)

    url = f"/uploads/products/{fname}"
    execute("UPDATE catalogo SET imagen_url=? WHERE id_producto=?", (url, id_producto))
    return jsonify({'imagen_url': url})


def _is_private_ip(ip: str) -> bool:
    try:
        addr = ipaddress.ip_address(ip)
        return (addr.is_private or addr.is_loopback or addr.is_link_local or addr.is_reserved or addr.is_multicast)
    except Exception:
        return True

def _safe_download_image(url: str, timeout: int = 12, max_bytes: int = 2_500_000) -> bytes:
    """Descarga segura: solo http/https, bloquea localhost/redes privadas, limita tamaÃ±o."""
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ('http', 'https'):
        raise ValueError('Solo URLs http/https')
    host = parsed.hostname
    if not host:
        raise ValueError('Host invÃ¡lido')
    # Resolver DNS y bloquear IPs privadas
    try:
        infos = socket.getaddrinfo(host, None)
        ips = sorted({i[4][0] for i in infos})
        for ip in ips:
            if _is_private_ip(ip):
                raise ValueError('Host no permitido')
    except ValueError:
        raise
    except Exception:
        # Si no podemos resolver, rechazamos por seguridad
        raise ValueError('No se pudo validar el host')

    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = resp.read(max_bytes + 1)
        if len(data) > max_bytes:
            raise ValueError('Archivo demasiado grande')
        return data

@app.post('/api/catalogo/<id_producto>/import_url')
@role_required('admin')
def import_product_image_from_url(id_producto):
    """Descarga una imagen desde URL, la guarda en /uploads/products y opcionalmente quita fondo blanco."""
    data = request.json or {}
    img_url = (data.get('url') or '').strip()
    if not img_url:
        return jsonify({'error': 'URL requerida'}), 400
    remove_bg = str(data.get('remove_bg', '0')) == '1'

    try:
        parsed = urllib.parse.urlparse(img_url)
        ext = os.path.splitext(parsed.path)[1].lower().lstrip('.')
        if ext not in ALLOWED_EXTENSIONS:
            ext = 'jpg'
        tmp_name = f"{id_producto}_{uuid.uuid4().hex[:8]}_dl.{ext}"
        tmp_path = os.path.join(UPLOAD_FOLDER, tmp_name)
        content = _safe_download_image(img_url, timeout=12, max_bytes=2_500_000)
        with open(tmp_path, 'wb') as f:
            f.write(content)
    except Exception as e:
        return jsonify({'error': f'No se pudo descargar la imagen: {e}'}), 400

    final_name = tmp_name
    if remove_bg:
        png_name = f"{id_producto}_{uuid.uuid4().hex[:8]}_cut.png"
        png_path = os.path.join(UPLOAD_FOLDER, png_name)
        try:
            _remove_white_bg_to_png(tmp_path, png_path)
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            final_name = png_name
        except Exception as e:
            print('[WARN] import_url remove_bg failed:', e)

    old = query("SELECT imagen_url FROM catalogo WHERE id_producto=?", (id_producto,), one=True)
    if old and old.get('imagen_url'):
        _delete_primary_upload_from_url(old['imagen_url'])

    url_local = f"/uploads/products/{final_name}"
    execute("UPDATE catalogo SET imagen_url=? WHERE id_producto=?", (url_local, id_producto))
    return jsonify({'imagen_url': url_local})


@app.post('/api/catalogo/<id_producto>/remove_bg')
@role_required('admin')
def remove_bg_existing(id_producto):
    """Convierte la imagen actual del producto a PNG con fondo transparente (si es fondo blanco)."""
    row = query("SELECT imagen_url FROM catalogo WHERE id_producto=?", (id_producto,), one=True)
    if not row or not row.get('imagen_url'):
        return jsonify({'ok': False, 'error': 'Producto sin imagen'}), 400
    src = _resolve_upload_from_url(row.get('imagen_url'))
    if not src or not os.path.isfile(src):
        return jsonify({'ok': False, 'error': 'Imagen no encontrada en servidor'}), 404
    png_name = f"{id_producto}_{uuid.uuid4().hex[:8]}_cut.png"
    png_path = os.path.join(UPLOAD_FOLDER, png_name)
    _remove_white_bg_to_png(src, png_path)
    url = f"/uploads/products/{png_name}"
    execute("UPDATE catalogo SET imagen_url=? WHERE id_producto=?", (url, id_producto))
    return jsonify({'ok': True, 'imagen_url': url})

@app.delete('/api/catalogo/<id_producto>/imagen')
@role_required('admin')
def delete_product_image(id_producto):
    old = query("SELECT imagen_url FROM catalogo WHERE id_producto=?", (id_producto,), one=True)
    if old and old['imagen_url']:
        _delete_primary_upload_from_url(old['imagen_url'])
    execute("UPDATE catalogo SET imagen_url='' WHERE id_producto=?", (id_producto,))
    return jsonify({'ok': True})

# â”€â”€â”€ RUTAS API â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get('/health')
def health():
    return jsonify({'ok': True})

@app.get('/api/me')
def api_me():
    u = current_user()
    if not u:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    # refrescar/asegurar token CSRF
    tok = _csrf_token()
    return jsonify({'ok': True, 'user': u, 'csrf_token': tok})


@app.post('/api/login')
@limit('5 per minute; 20 per hour')
def api_login():
    d = request.json or {}
    username = (d.get('username') or '').strip()
    password = d.get('password') or ''
    if not username or not password:
        _audit_event('login_failed', outcome='rejected', username=username or '', reason='missing_fields')
        return jsonify({'ok': False, 'error': 'Usuario y contraseÃ±a requeridos'}), 400

    u = query("SELECT * FROM users WHERE username=?", (username,), one=True)
    if not u or not verify_password(u.get('password_hash') or '', password):
        # Respuesta genÃ©rica para no filtrar existencia de usuarios
        _audit_event('login_failed', outcome='rejected', username=username, reason='invalid_credentials')
        return jsonify({'ok': False, 'error': 'Credenciales invÃ¡lidas'}), 401

    # RotaciÃ³n automÃ¡tica a Argon2 cuando estÃ© disponible
    try:
        ph = u.get('password_hash') or ''
        if _ARGON2 and (not ph.startswith('$argon2')):
            execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(password), u['id']))
    except Exception:
        pass

    # Renovar la sesiÃ³n para evitar fixation y limpiar residuos previos.
    session.clear()
    session.permanent = True
    must_change = int(u.get('must_change_password') or 0) == 1
    session['user'] = {'id': u['id'], 'username': u['username'], 'role': u.get('role', 'admin'), 'must_change_password': must_change}
    # CSRF token de sesiÃ³n
    tok = _csrf_token()
    actor = {'type': 'user', 'id': u['id'], 'username': u['username'], 'role': u.get('role', 'admin')}
    _audit_event('login_success', actor=actor, must_change_password=must_change)
    return jsonify({'ok': True, 'user': session['user'], 'csrf_token': tok, 'must_change_password': must_change})


@app.post('/api/logout')
@login_required
def api_logout():
    u = current_user()
    if u:
        _audit_event('logout', actor=u)
    session.clear()
    return jsonify({'ok': True})

@app.post('/api/change_password')
@login_required
@limit('10 per hour')
def api_change_password():
    d = request.json or {}
    old = d.get('old_password') or ''
    new = d.get('new_password') or ''
    if not new or len(new) < 10:
        _audit_event('password_change_failed', outcome='rejected', reason='weak_password')
        return jsonify({'ok': False, 'error': 'La nueva contraseÃ±a debe tener mÃ­nimo 10 caracteres.'}), 400
    u = current_user()
    row = query("SELECT * FROM users WHERE id=?", (u['id'],), one=True)
    if not row:
        _audit_event('password_change_failed', outcome='error', reason='user_not_found')
        return jsonify({'ok': False, 'error': 'Usuario no encontrado'}), 404
    if not verify_password(row.get('password_hash') or '', old):
        _audit_event('password_change_failed', outcome='rejected', reason='wrong_current_password')
        return jsonify({'ok': False, 'error': 'ContraseÃ±a actual incorrecta'}), 401
    execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?", (hash_password(new), row['id']))
    # refrescar bandera en sesiÃ³n
    session['user']['must_change_password'] = False
    _audit_event('password_changed', actor=current_user(), user_id=row['id'])
    return jsonify({'ok': True})

@app.post('/api/admin/reauth')
@role_required('admin')
@limit('20 per hour')
def api_admin_reauth():
    d = request.json or {}
    sec = _runtime_security_settings()
    reauth_seconds = int(sec.get('admin_reauth_seconds') or 0)
    pw = str(d.get('password') or d.get('admin_password') or '').strip()
    if not pw:
        _audit_event('admin_reauth_failed', outcome='rejected', actor=current_user(), reason='password_required')
        return jsonify({'ok': False, 'error': 'Contrasena requerida'}), 400
    if not _verify_current_user_password(pw):
        _audit_event('admin_reauth_failed', outcome='rejected', actor=current_user(), reason='invalid_password')
        return jsonify({'ok': False, 'error': 'Contrasena invalida'}), 401
    _mark_admin_reauth_ok()
    _audit_event('admin_reauth', actor=current_user(), source='explicit', valid_for_seconds=reauth_seconds)
    return jsonify({'ok': True, 'valid_for_seconds': reauth_seconds})


@app.get('/api/catalogo')
def get_catalogo():
    cat = request.args.get('categoria')
    solo_activos = request.args.get('activos', '1') == '1'
    sql = "SELECT * FROM catalogo WHERE 1=1"; p = []
    if cat: sql += " AND categoria=?"; p.append(cat.upper())
    if solo_activos: sql += " AND activo=1"
    sql += " ORDER BY categoria, nombre"
    return jsonify(query(sql, p))

@app.get('/api/catalogo/categorias')
def get_categorias():
    return jsonify([r['categoria'] for r in query("SELECT DISTINCT categoria FROM catalogo WHERE activo=1 ORDER BY categoria")])

@app.get('/api/catalogo/<id_producto>')
def get_producto(id_producto):
    p = query("SELECT * FROM catalogo WHERE id_producto=?", (id_producto,), one=True)
    return jsonify(p) if p else ('Not found', 404)

@app.put('/api/catalogo/<id_producto>')
@role_required('admin')
def update_producto(id_producto):
    d = request.json
    execute("""UPDATE catalogo SET nombre=?,descripcion=?,precio=?,aplica_iva=?,
               pct_iva=?,inst_default=?,config_default=?,imagen_url=?,activo=?,
               categoria=?,unidad=?,costo_unitario=? WHERE id_producto=?""",
            (d.get('nombre'), d.get('descripcion'), d.get('precio',0), d.get('aplica_iva',0),
             d.get('pct_iva',0), d.get('inst_default',0), d.get('config_default',0),
             d.get('imagen_url',''), d.get('activo',1), d.get('categoria'), d.get('unidad','Und'),
             d.get('costo_unitario',0), id_producto))
    return jsonify({'ok': True})


# â”€â”€â”€ ID CONSECUTIVO POR CATEGORÃA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PREFIX_MAP = {
    'CCTV': 'CAM',
    'DOMOTICA': 'DOM',
    'AUDIOVISUAL': 'AV',
    'CERRADURAS': 'CEK',
    'REDES': 'RED',
    'SERVICIOS': 'MO',
    'OTROS': 'OTR',
}


def next_id_for_categoria(cat: str) -> str:
    cat = (cat or 'OTROS').upper().strip()
    pref = PREFIX_MAP.get(cat, 'PRD')
    row = query(
        """SELECT MAX(CAST(SUBSTR(id_producto, INSTR(id_producto,'-')+1) AS INTEGER)) as mx
           FROM catalogo
           WHERE id_producto LIKE ?""",
        (f"{pref}-%",),
        one=True,
    )
    mx = int((row or {}).get('mx') or 0)
    return f"{pref}-{mx+1:03d}"


@app.get('/api/catalogo/next_id')
def api_next_id():
    cat = request.args.get('categoria')
    return jsonify({'ok': True, 'categoria': (cat or 'OTROS').upper(), 'id': next_id_for_categoria(cat)})

@app.post('/api/catalogo')
@role_required('admin')
def create_producto():
    d = request.json
    if not d.get('id_producto'):
        d['id_producto'] = next_id_for_categoria(d.get('categoria'))
    execute("""INSERT INTO catalogo (id_producto,categoria,nombre,descripcion,unidad,
               precio,aplica_iva,pct_iva,inst_default,config_default,costo_unitario,stock_qty,stock_min,imagen_url,activo)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d['id_producto'], d.get('categoria','OTROS'), d['nombre'], d.get('descripcion',''),
             d.get('unidad','Und'), d.get('precio',0), d.get('aplica_iva',0), d.get('pct_iva',0),
             d.get('inst_default',0), d.get('config_default',0), d.get('costo_unitario',0),
             d.get('stock_qty',0), d.get('stock_min',0), d.get('imagen_url',''), d.get('activo',1)))
    return jsonify({'ok': True, 'id': d['id_producto']})


# â”€â”€â”€ INVENTARIO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/inventario')
@role_required('admin')
def api_inventario():
    bodega = (request.args.get('bodega') or 'PRINCIPAL').upper()
    # aseguramos bodega
    execute("INSERT OR IGNORE INTO bodegas(nombre) VALUES (?)", (bodega,))
    rows = query("""SELECT c.id_producto, c.categoria, c.nombre, c.imagen_url,
               COALESCE(sb.stock_qty, c.stock_qty, 0) as stock_qty,
               COALESCE(c.stock_min, 0) as stock_min,
               COALESCE(c.costo_unitario, 0) as costo_unitario,
               c.precio, c.activo
        FROM catalogo c
        LEFT JOIN stock_bodega sb ON sb.id_producto=c.id_producto AND sb.bodega=?
        ORDER BY c.categoria, c.nombre""", (bodega,))
    for r in rows:
        r['alerta_reponer'] = 1 if float(r.get('stock_qty') or 0) <= float(r.get('stock_min') or 0) and float(r.get('stock_min') or 0) > 0 else 0
    return jsonify({'bodega': bodega, 'items': rows})

@app.get('/api/inventario/movimientos')
@role_required('admin')
def api_inv_movs():
    bodega = (request.args.get('bodega') or 'PRINCIPAL').upper()
    rows = query("""SELECT m.*, c.nombre, c.categoria
        FROM movimientos_inventario m
        JOIN catalogo c ON c.id_producto=m.id_producto
        WHERE m.bodega=?
        ORDER BY m.id DESC LIMIT 500""", (bodega,))
    return jsonify({'bodega': bodega, 'movimientos': rows})

@app.post('/api/inventario/movimiento')
@role_required('admin')
def api_inv_mov():
    d = request.json or {}
    bodega = (d.get('bodega') or 'PRINCIPAL').upper()
    pid = (d.get('id_producto') or '').strip()
    tipo = (d.get('tipo') or 'AJUSTE').upper()
    cant = float(d.get('cantidad') or 0)
    costo_unit = d.get('costo_unit', None)
    try:
        costo_unit = None if costo_unit in (None, '', 'null') else float(costo_unit)
    except Exception:
        costo_unit = None
    nota = d.get('nota','') or ''
    ref  = d.get('ref','') or ''
    if not pid or cant == 0:
        return jsonify({'ok': False, 'error': 'Faltan datos'}), 400
    execute("INSERT OR IGNORE INTO bodegas(nombre) VALUES (?)", (bodega,))
    execute("INSERT OR IGNORE INTO stock_bodega(bodega,id_producto,stock_qty) VALUES (?,?,0)", (bodega,pid))

    # calcular delta segun tipo
    delta = cant
    if tipo in ('SALIDA','RESERVA'):
        delta = -abs(cant)
    elif tipo in ('ENTRADA','LIBERACION'):
        delta = abs(cant)
    elif tipo == 'AJUSTE':
        # AJUSTE: setear stock a cant (cant = nuevo stock)
        current = query("SELECT stock_qty FROM stock_bodega WHERE bodega=? AND id_producto=?", (bodega,pid), one=True)
        cur = float((current or {}).get('stock_qty') or 0)
        delta = cant - cur

    execute("UPDATE stock_bodega SET stock_qty=COALESCE(stock_qty,0)+? WHERE bodega=? AND id_producto=?", (delta,bodega,pid))
    execute("INSERT INTO movimientos_inventario(id_producto,bodega,tipo,cantidad,nota,ref) VALUES (?,?,?,?,?,?)",
            (pid,bodega,tipo,abs(cant) if tipo!='AJUSTE' else cant,nota,ref))

    # Si llega un costo_unit, actualizar costo_unitario (ENTRADA: promedio ponderado, AJUSTE: set)
    if costo_unit is not None and costo_unit >= 0:
        cur = query("SELECT costo_unitario, stock_qty, precio FROM catalogo WHERE id_producto=?", (pid,), one=True) or {}
        cur_cost = float(cur.get('costo_unitario') or 0)
        cur_stock = float(cur.get('stock_qty') or 0)
        if tipo == 'ENTRADA':
            new_stock = cur_stock + abs(cant)
            new_cost = costo_unit if cur_stock <= 0 else ((cur_cost*cur_stock) + (costo_unit*abs(cant))) / (new_stock if new_stock>0 else 1)
        else:
            new_cost = costo_unit
        execute("UPDATE catalogo SET costo_unitario=? WHERE id_producto=?", (new_cost, pid))
        execute("INSERT INTO historial_precios(id_producto,precio,costo_unit,nota) VALUES (?,?,?,?)",
                (pid, float(cur.get('precio') or 0), new_cost, f'Inv {tipo}'))

    # sincronizar total en catalogo como sumatoria de bodegas
    total = query("SELECT SUM(stock_qty) as t FROM stock_bodega WHERE id_producto=?", (pid,), one=True)
    execute("UPDATE catalogo SET stock_qty=COALESCE(?,0) WHERE id_producto=?", (float(total.get('t') or 0), pid))
    return jsonify({'ok': True})

# â”€â”€â”€ PROVEEDORES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/proveedores')
@role_required('admin')
def api_proveedores_list():
    rows = query("SELECT * FROM proveedores ORDER BY nombre")
    return jsonify(rows)

@app.post('/api/proveedores')
@role_required('admin')
def api_proveedores_create():
    d = request.json or {}
    if not (d.get('nombre') or '').strip():
        return jsonify({'ok': False, 'error': 'Nombre requerido'}), 400
    pid = execute("INSERT INTO proveedores(nombre,whatsapp,email,condiciones) VALUES (?,?,?,?)",
                  (d['nombre'].strip(), d.get('whatsapp',''), d.get('email',''), d.get('condiciones','')))
    return jsonify({'ok': True, 'id': pid})

# â”€â”€â”€ COMPRAS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _gen_no_compra():
    y = datetime.datetime.now().strftime('%y')
    n = query("SELECT COUNT(*) as n FROM compras", one=True)['n']
    return f"CP{y}-{int(n)+1:05d}"

@app.get('/api/compras')
@role_required('admin')
def api_compras_list():
    rows = query("""SELECT co.*, p.nombre as proveedor
        FROM compras co LEFT JOIN proveedores p ON p.id=co.proveedor_id
        ORDER BY co.id DESC LIMIT 200""")
    return jsonify(rows)

@app.post('/api/compras')
@role_required('admin')
def api_compras_create():
    d = request.json or {}
    items = d.get('items') or []
    if not items:
        return jsonify({'ok': False, 'error': 'Items requeridos'}), 400
    no = d.get('no_compra') or _gen_no_compra()
    compra_id = execute("INSERT INTO compras(no_compra,fecha,proveedor_id,notas,total) VALUES (?,?,?,?,0)",
                        (no, d.get('fecha') or datetime.date.today().isoformat(), d.get('proveedor_id'), d.get('notas','')))
    total = 0.0
    for it in items:
        pid = it.get('id_producto'); qty = float(it.get('cantidad') or 0); cost = float(it.get('costo_unit') or 0)
        if not pid or qty<=0: continue
        subtotal = qty*cost; total += subtotal
        execute("INSERT INTO compra_items(compra_id,id_producto,cantidad,costo_unit,subtotal) VALUES (?,?,?,?,?)",
                (compra_id,pid,qty,cost,subtotal))
        # actualizar costo_unitario (promedio ponderado sencillo)
        cur = query("SELECT costo_unitario, stock_qty FROM catalogo WHERE id_producto=?", (pid,), one=True) or {}
        cur_cost = float(cur.get('costo_unitario') or 0)
        cur_stock = float(cur.get('stock_qty') or 0)
        new_stock = cur_stock + qty
        new_cost = cost if cur_stock<=0 else ((cur_cost*cur_stock) + (cost*qty)) / (new_stock if new_stock>0 else 1)
        execute("UPDATE catalogo SET costo_unitario=? WHERE id_producto=?", (new_cost, pid))
        execute("INSERT INTO historial_precios(id_producto,precio,costo_unit,nota) VALUES (?,?,?,?)",
                (pid, float(cur.get('precio') or 0), new_cost, f'Compra {no}'))
        # movimiento inventario ENTRADA
        execute("INSERT OR IGNORE INTO bodegas(nombre) VALUES ('PRINCIPAL')")
        execute("INSERT OR IGNORE INTO stock_bodega(bodega,id_producto,stock_qty) VALUES ('PRINCIPAL',?,0)", (pid,))
        execute("UPDATE stock_bodega SET stock_qty=COALESCE(stock_qty,0)+? WHERE bodega='PRINCIPAL' AND id_producto=?", (qty,pid))
        execute("INSERT INTO movimientos_inventario(id_producto,bodega,tipo,cantidad,nota,ref) VALUES (?,?,?,?,?,?)",
                (pid,'PRINCIPAL','ENTRADA',qty,'Compra',no))
        # sincronizar total
        tot = query("SELECT SUM(stock_qty) as t FROM stock_bodega WHERE id_producto=?", (pid,), one=True)
        execute("UPDATE catalogo SET stock_qty=? WHERE id_producto=?", (float(tot.get('t') or 0), pid))
    execute("UPDATE compras SET total=? WHERE id=?", (total, compra_id))
    return jsonify({'ok': True, 'id': compra_id, 'no_compra': no, 'total': total})


# â”€â”€â”€ PAQUETES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/paquetes')
def api_paquetes_list():
    q = request.args.get('q','').strip()
    sql = """
    SELECT p.*, 
      (SELECT COALESCE(SUM(pi.cantidad),0) FROM paquete_items pi WHERE pi.paquete_id=p.id) as items_count,
      (SELECT COALESCE(SUM(pi.cantidad * c.precio),0) 
         FROM paquete_items pi JOIN catalogo c ON c.id_producto=pi.id_producto 
        WHERE pi.paquete_id=p.id) as total_productos
    FROM paquetes p
    WHERE p.activo=1
    """
    params = []
    if q:
        sql += " AND nombre LIKE ?"
        params.append(f"%{q}%")
    sql += " ORDER BY id DESC"
    return jsonify(query(sql, params))

@app.post('/api/paquetes')
def api_paquetes_create():
    d = request.json or {}
    nombre = (d.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'ok': False, 'error': 'Nombre requerido'}), 400
    categoria = (d.get('categoria') or 'MIXTA').upper().strip()
    notas = d.get('notas','')
    pid = execute("INSERT INTO paquetes(nombre,categoria,notas,activo) VALUES (?,?,?,1)", (nombre,categoria,notas))
    items = d.get('items') or []
    for it in items:
        ip = it.get('id_producto'); qty = float(it.get('cantidad') or 0)
        if ip and qty>0:
            execute("INSERT INTO paquete_items(paquete_id,id_producto,cantidad) VALUES (?,?,?)", (pid,ip,qty))
    return jsonify({'ok': True, 'id': pid})

@app.get('/api/paquetes/<int:pid>')
def api_paquetes_get(pid):
    pack = query("SELECT * FROM paquetes WHERE id=?", (pid,), one=True)
    if not pack:
        return jsonify({'ok': False, 'error': 'No existe'}), 404
    items = query("SELECT pi.id_producto, pi.cantidad, c.nombre, c.categoria, c.precio, c.imagen_url "
                  "FROM paquete_items pi JOIN catalogo c ON c.id_producto=pi.id_producto "
                  "WHERE pi.paquete_id=? ORDER BY c.categoria, c.nombre", (pid,))
    return jsonify({'ok': True, 'pack': pack, 'items': items})

# â”€â”€â”€ CRM â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/crm/pipeline')
def api_crm_pipeline():
    etapa = (request.args.get('etapa') or '').upper().strip()
    q = (request.args.get('q') or '').strip()
    sql = "SELECT * FROM cotizaciones WHERE 1=1"; params=[]
    if etapa:
        sql += " AND etapa=?"; params.append(etapa)
    if q:
        sql += " AND (cliente LIKE ? OR no_cotizacion LIKE ? OR empresa LIKE ?)"; params += [f"%{q}%"]*3
    sql += " ORDER BY created_at DESC LIMIT 300"
    rows = query(sql, params)
    for r in rows:
        # Normalizar etapa segÃºn estado para que el pipeline se vea coherente
        est = (r.get('estado') or 'BORRADOR').upper().strip()
        et = (r.get('etapa') or '').upper().strip()
        if not et or et == 'LEAD':
            if est == 'ENVIADA':
                et = 'ENVIADA'
            elif est == 'APROBADA':
                et = 'APROBADA'
            elif est in ('RECHAZADA','PERDIDA'):
                et = 'PERDIDA'
            else:
                et = 'COTIZADA'
            r['etapa'] = et
        r.update(calcular_cotizacion(r['id']))
    return jsonify(rows)

@app.post('/api/crm/cot/<int:cid>')
def api_crm_update(cid):
    d = request.json or {}
    etapa = (d.get('etapa') or '').upper().strip()
    estado = (d.get('estado') or '').upper().strip()
    seguimiento_fecha = d.get('seguimiento_fecha') or ''
    proxima_accion = d.get('proxima_accion') or ''
    motivo_perdida = d.get('motivo_perdida') or ''
    fields=[]; params=[]
    if etapa: fields.append('etapa=?'); params.append(etapa)
    if estado: fields.append('estado=?'); params.append(estado)
    fields.append('seguimiento_fecha=?'); params.append(seguimiento_fecha)
    fields.append('proxima_accion=?'); params.append(proxima_accion)
    fields.append('motivo_perdida=?'); params.append(motivo_perdida)
    params.append(cid)
    execute(f"UPDATE cotizaciones SET {', '.join(fields)} WHERE id=?", tuple(params))
    return jsonify({'ok': True})

@app.post('/api/crm/actividad')
def api_crm_actividad():
    d = request.json or {}
    cid = int(d.get('cot_id') or 0)
    if cid<=0:
        return jsonify({'ok': False, 'error': 'cot_id requerido'}), 400
    tipo = (d.get('tipo') or 'NOTA').upper().strip()
    nota = d.get('nota') or ''
    prox = d.get('proxima_fecha') or ''
    execute("INSERT INTO crm_actividades(cot_id,tipo,nota,proxima_fecha) VALUES (?,?,?,?)", (cid,tipo,nota,prox))
    return jsonify({'ok': True})

@app.get('/api/crm/actividades/<int:cid>')
def api_crm_actividades(cid):
    return jsonify(query("SELECT * FROM crm_actividades WHERE cot_id=? ORDER BY id DESC LIMIT 200", (cid,)))

# Cotizaciones
@app.get('/api/cotizaciones')
def get_cotizaciones():
    estado = (request.args.get('estado') or '').strip()
    if estado.lower() in ('todos', 'all', 'todas'):
        estado = ''
    busca = request.args.get('q', '')
    sql = "SELECT * FROM cotizaciones WHERE 1=1"; p = []
    if estado: sql += " AND estado=?"; p.append(estado)
    if busca:
        sql += " AND (cliente LIKE ? OR no_cotizacion LIKE ? OR empresa LIKE ?)"
        p += [f'%{busca}%'] * 3
    sql += " ORDER BY created_at DESC"
    cots = query(sql, p)
    for c in cots:
        c.update(calcular_cotizacion(c['id']))
        # AÃ±adir rentabilidad resumida para tarjetas (solo interno)
        try:
            m = calcular_margenes(c['id'])
            c['utilidad_neta'] = m.get('utilidad_neta', 0)
            c['margen_neto_sin_iva'] = m.get('margen_neto_sin_iva', 0)
            c['costos_ok'] = 1 if m.get('costos_ok') else 0
        except Exception:
            c['utilidad_neta'] = 0
            c['margen_neto_sin_iva'] = 0
            c['costos_ok'] = 0
    return jsonify(cots)

@app.post('/api/cotizaciones')
@role_required('admin','vendedor')
def create_cotizacion():
    d = request.json or {}
    no_cot = next_no_cotizacion()
    # Lista de precios (descuento aplicado a precios base si no hay precio_manual)
    price_list_code = (d.get('price_list_code') or 'PUBLICO').upper().strip()
    pl_row = query("SELECT desc_pct FROM price_lists WHERE code=?", (price_list_code,), one=True)
    price_list_desc_pct = float(pl_row['desc_pct']) if pl_row else 0.0
    # Checklist
    chk = d.get('checklist_json')
    if isinstance(chk, (dict, list)):
        checklist_json = json.dumps(chk, ensure_ascii=False)
    elif isinstance(chk, str) and chk.strip():
        checklist_json = chk
    else:
        checklist_json = json.dumps([
            {"label":"Visita tÃ©cnica", "done": False},
            {"label":"Medidas confirmadas", "done": False},
            {"label":"Cliente aprueba equipos", "done": False},
            {"label":"Anticipo recibido", "done": False},
            {"label":"ProgramaciÃ³n agenda", "done": False},
            {"label":"InstalaciÃ³n completa", "done": False},
            {"label":"Entrega y capacitaciÃ³n", "done": False},
        ], ensure_ascii=False)
    # Al crear manualmente, tambiÃ©n queda en etapa COTIZADA por defecto
    cot_id = execute("""INSERT INTO cotizaciones
        (no_cotizacion,cliente,empresa,nit_cc,telefono,email_cliente,direccion,ciudad,
         proyecto,tipo_cotizacion,forma_pago,
         anticipo_pct,anticipo_val_manual,abonado_val,
         descuento_pct,descuento_val,notas,
         notas_internas,checklist_json,price_list_code,price_list_desc_pct,vendedor,etapa)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (no_cot, d.get('cliente',''), d.get('empresa',''), d.get('nit_cc',''), d.get('telefono',''),
         d.get('email_cliente',''), d.get('direccion',''), d.get('ciudad',''),
         d.get('proyecto',''), d.get('tipo_cotizacion','MIXTA'), d.get('forma_pago','70% - 30%'),
         d.get('anticipo_pct',0.70), float(d.get('anticipo_val_manual',0) or 0), float(d.get('abonado_val',0) or 0),
         d.get('descuento_pct',0), d.get('descuento_val',0),
         d.get('notas',''), d.get('notas_internas',''), checklist_json,
         price_list_code, price_list_desc_pct, d.get('vendedor','Admin'), 'COTIZADA'))
    for idx, it in enumerate(d.get('items', []), 1):
        execute("""INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (cot_id, idx, it['id_producto'], it.get('cantidad',1),
                 it.get('precio_manual',0), it.get('inst_manual',0),
                 it.get('cfg_manual',0), it.get('notas_item','')))
    return jsonify({'ok': True, 'id': cot_id, 'no_cotizacion': no_cot})

@app.get('/api/cotizaciones/<int:cot_id>')
def get_cotizacion(cot_id):
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot: return ('Not found', 404)
    pl = float(cot.get('price_list_desc_pct') or 0)
    items = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                     c.aplica_iva, c.pct_iva, c.inst_default, c.config_default,
                     c.categoria, c.imagen_url, COALESCE(c.costo_unitario,0) as costo_unitario
                     FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                     WHERE i.cot_id=? ORDER BY i.linea""", (cot_id,))
    ic = []
    for it in items:
        calc = calcular_item(it, it, pl)
        ic.append({**it, **calc})
    tots = calcular_cotizacion(cot_id); cot.update(tots); cot['items'] = ic
    return jsonify(cot)


# â”€â”€â”€ GASTOS POR PROYECTO (ligados a cotizaciÃ³n) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/cotizaciones/<int:cot_id>/gastos')
@role_required('admin','vendedor')
def api_list_gastos(cot_id):
    rows = query("SELECT * FROM gastos_proyecto WHERE cot_id=? ORDER BY fecha DESC, id DESC", (cot_id,))
    total = 0.0
    items = []
    for r in rows:
        v = float(r.get('valor') or 0)
        total += v
        items.append({**r, 'valor': v})
    return jsonify({'items': items, 'total': round(total, 0)})


@app.post('/api/cotizaciones/<int:cot_id>/gastos')
@role_required('admin','vendedor')
def api_add_gasto(cot_id):
    d = request.json or {}
    concepto = str(d.get('concepto','')).strip()
    if not concepto:
        return jsonify({'error':'Concepto requerido'}), 400
    try:
        valor = float(d.get('valor',0) or 0)
    except Exception:
        valor = 0.0
    if valor <= 0:
        return jsonify({'error':'Valor debe ser mayor a 0'}), 400
    fecha = str(d.get('fecha','')).strip()
    if not fecha:
        fecha = datetime.now().strftime('%Y-%m-%d')
    nota = str(d.get('nota','')).strip()
    gid = execute("INSERT INTO gastos_proyecto (cot_id,fecha,concepto,valor,nota) VALUES (?,?,?,?,?)",
                  (cot_id, fecha, concepto, valor, nota))
    return jsonify({'ok': True, 'id': gid})


@app.delete('/api/gastos/<int:gasto_id>')
@role_required('admin','vendedor')
def api_delete_gasto(gasto_id):
    execute("DELETE FROM gastos_proyecto WHERE id=?", (gasto_id,))
    return jsonify({'ok': True})

@app.put('/api/cotizaciones/<int:cot_id>')
@role_required('admin','vendedor')
def update_cotizacion(cot_id):
    d = request.json or {}

    # Traer lo que ya existe para NO pisar con None/"" (NULL)
    curr = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not curr:
        return ("Not found", 404)

    # Atajo: si SOLO viene estado, actualizar solo estado (evita NOT NULL cliente)
    if set(d.keys()) <= {"estado"}:
        execute("UPDATE cotizaciones SET estado=? WHERE id=?",
                (d.get("estado", curr.get("estado", "BORRADOR")), cot_id))
        return jsonify({"ok": True})

    # Campos obligatorios que nunca pueden quedar vacÃ­os
    REQUIRED_FIELDS = {"cliente", "tipo_cotizacion", "forma_pago"}

    def keep(key, default=""):
        v = d.get(key, None)
        if v is None:
            return curr.get(key, default)
        # Solo bloquear vaciado en campos requeridos
        if key in REQUIRED_FIELDS and isinstance(v, str) and v.strip() == "":
            return curr.get(key, default)
        return v

    def norm_pct(val, fallback):
        # En BD se guarda 0.70 (70%). Si llega 70, lo convertimos.
        if val is None:
            return fallback
        try:
            x = float(val)
        except Exception:
            return fallback
        if x > 1:
            x = x / 100.0
        if x < 0:
            x = 0
        if x > 1:
            x = 1
        return x

    cliente = keep("cliente", "")
    if not str(cliente).strip():
        cliente = curr.get("cliente", "")

    anticipo_pct  = norm_pct(d.get("anticipo_pct", None), float(curr.get("anticipo_pct", 0.70) or 0.70))
    descuento_pct = norm_pct(d.get("descuento_pct", None), float(curr.get("descuento_pct", 0.0) or 0.0))

    # Montos manuales (opcionales)
    def norm_money(val, fallback=0.0):
        if val is None:
            return float(fallback or 0)
        try:
            x = float(val)
        except Exception:
            return float(fallback or 0)
        if x < 0:
            x = 0
        return x

    anticipo_val_manual = norm_money(d.get("anticipo_val_manual", None), curr.get("anticipo_val_manual", 0))
    abonado_val = norm_money(d.get("abonado_val", None), curr.get("abonado_val", 0))

    # Lista de precios
    pl_code_raw = d.get('price_list_code', None)
    if pl_code_raw is None:
        price_list_code = (curr.get('price_list_code') or 'PUBLICO').upper()
        price_list_desc_pct = float(curr.get('price_list_desc_pct') or 0)
    else:
        price_list_code = (str(pl_code_raw) or 'PUBLICO').upper().strip()
        pl_row = query("SELECT desc_pct FROM price_lists WHERE code=?", (price_list_code,), one=True)
        price_list_desc_pct = float(pl_row['desc_pct']) if pl_row else 0.0

    # Checklist
    chk = d.get('checklist_json', None)
    if chk is None:
        checklist_json = curr.get('checklist_json') or ''
    elif isinstance(chk, (dict, list)):
        checklist_json = json.dumps(chk, ensure_ascii=False)
    else:
        checklist_json = str(chk)

    execute("""UPDATE cotizaciones SET
        cliente=?, empresa=?, nit_cc=?, telefono=?, email_cliente=?,
        direccion=?, ciudad=?, proyecto=?, tipo_cotizacion=?, forma_pago=?,
        anticipo_pct=?, anticipo_val_manual=?, abonado_val=?, descuento_pct=?, descuento_val=?, notas=?, notas_internas=?,
        checklist_json=?, price_list_code=?, price_list_desc_pct=?, estado=?
        WHERE id=?""",
        (
            cliente,
            keep("empresa", ""),
            keep("nit_cc", ""),
            keep("telefono", ""),
            keep("email_cliente", ""),
            keep("direccion", ""),
            keep("ciudad", ""),
            keep("proyecto", ""),
            keep("tipo_cotizacion", "MIXTA"),
            keep("forma_pago", "70% - 30%"),
            anticipo_pct,
            anticipo_val_manual,
            abonado_val,
            descuento_pct,
            float(keep("descuento_val", 0) or 0),
            keep("notas", ""),
            keep("notas_internas", ""),
            checklist_json,
            price_list_code,
            price_list_desc_pct,
            keep("estado", "BORRADOR"),
            cot_id
        )
    )

    # Actualizar items SOLO si llegan en el request
    if "items" in d:
        execute("DELETE FROM items WHERE cot_id=?", (cot_id,))
        for idx, it in enumerate(d.get("items", []), 1):
            execute("""INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (cot_id, idx, it["id_producto"], it.get("cantidad", 1),
                     it.get("precio_manual", 0), it.get("inst_manual", 0),
                     it.get("cfg_manual", 0), it.get("notas_item", "")))

    return jsonify({"ok": True})


@app.delete('/api/cotizaciones/<int:cot_id>')
@role_required('admin','vendedor')
def delete_cotizacion(cot_id):
    if not query("SELECT id FROM cotizaciones WHERE id=?", (cot_id,), one=True):
        return jsonify({'ok': False, 'error': 'No encontrada'}), 404
    try:
        snapshot_db(f'before_delete_cot_{cot_id}')
    except Exception:
        pass
    execute("DELETE FROM cotizaciones WHERE id=?", (cot_id,))
    return jsonify({'ok': True})


@app.post('/api/cotizaciones/<int:cot_id>/clone')
@role_required('admin','vendedor')
def clone_cotizacion(cot_id):
    src = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not src:
        return jsonify({'ok': False, 'error': 'No existe'}), 404
    no_cot = next_no_cotizacion()
    # Copia sin aceptaciÃ³n ni token pÃºblico
    new_id = execute("""INSERT INTO cotizaciones (
        no_cotizacion,cliente,empresa,nit_cc,telefono,email_cliente,direccion,ciudad,
        proyecto,tipo_cotizacion,forma_pago,anticipo_pct,descuento_pct,descuento_val,notas,
        notas_internas,checklist_json,price_list_code,price_list_desc_pct,vendedor,estado,etapa
    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
    (
        no_cot,
        src.get('cliente',''), src.get('empresa',''), src.get('nit_cc',''), src.get('telefono',''),
        src.get('email_cliente',''), src.get('direccion',''), src.get('ciudad',''),
        src.get('proyecto',''), src.get('tipo_cotizacion','MIXTA'), src.get('forma_pago','70% - 30%'),
        src.get('anticipo_pct',0.70), src.get('descuento_pct',0), src.get('descuento_val',0), src.get('notas',''),
        src.get('notas_internas',''), src.get('checklist_json',''), src.get('price_list_code','PUBLICO'),
        src.get('price_list_desc_pct',0), src.get('vendedor','Admin'), 'BORRADOR', 'COTIZADA'
    ))
    items = query("SELECT * FROM items WHERE cot_id=? ORDER BY linea", (cot_id,))
    for idx, it in enumerate(items, 1):
        execute("""INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (new_id, idx, it['id_producto'], it.get('cantidad',1), it.get('precio_manual',0),
                 it.get('inst_manual',0), it.get('cfg_manual',0), it.get('notas_item','')))
    return jsonify({'ok': True, 'id': new_id, 'no_cotizacion': no_cot})


@app.put('/api/items/<int:item_id>')
@role_required('admin','vendedor')
def update_item(item_id):
    d = request.json
    execute("UPDATE items SET id_producto=?,cantidad=?,precio_manual=?,inst_manual=?,cfg_manual=?,notas_item=? WHERE id=?",
            (d['id_producto'], d.get('cantidad',1), d.get('precio_manual',0),
             d.get('inst_manual',0), d.get('cfg_manual',0), d.get('notas_item',''), item_id))
    return jsonify({'ok': True})

@app.delete('/api/items/<int:item_id>')
@role_required('admin','vendedor')
def delete_item(item_id):
    execute("DELETE FROM items WHERE id=?", (item_id,)); return jsonify({'ok': True})

@app.get('/api/parametros')
@role_required('admin')
def get_parametros():
    return jsonify({r['clave']: r['valor'] for r in query("SELECT * FROM parametros")})

@app.get('/api/admin/db_status')
@role_required('admin')
def api_db_status():
    try:
        sec = _runtime_security_settings(force=True)
        cot_count = query("SELECT COUNT(*) as n FROM cotizaciones", one=True)['n']
        item_count = query("SELECT COUNT(*) as n FROM items", one=True)['n']
        backups = _list_backups(limit=20)
        latest = backups[0] if backups else None
        latest_verify = _verify_sqlite_file(latest['path']) if latest else {'ok': False, 'integrity': 'missing'}
        return jsonify({
            'ok': True,
            'base_dir': BASE_DIR,
            'data_dir': DATA_DIR,
            'db_path': DB_PATH,
            'uploads_dir': UPLOAD_FOLDER,
            'backup_dir': BACKUP_DIR,
            'db_exists': os.path.isfile(DB_PATH),
            'db_size': os.path.getsize(DB_PATH) if os.path.isfile(DB_PATH) else 0,
            'cotizaciones': cot_count,
            'items': item_count,
            'backups': backups,
            'latest_backup': latest,
            'latest_backup_verify': latest_verify,
            'auto_backup': {
                'enabled': bool(sec.get('auto_backup_enabled')),
                'every_min': int(sec.get('auto_backup_every_min') or 0),
                'keep': int(sec.get('auto_backup_keep') or 0),
                'max_age_hours': int(sec.get('backup_max_age_hours') or 0),
            },
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.get('/api/admin/security_check')
@role_required('admin')
def api_security_check():
    """Checklist de hardening y persistencia para validar la nube con evidencia."""
    try:
        sec = _runtime_security_settings(force=True)
        now_utc = datetime.datetime.utcnow().replace(microsecond=0).isoformat() + 'Z'
        secret = app.secret_key or ''
        data_dir_abs = os.path.abspath(DATA_DIR)
        db_path_abs = os.path.abspath(DB_PATH)
        backup_dir_abs = os.path.abspath(BACKUP_DIR)
        audit_path_abs = os.path.abspath(AUDIT_LOG_PATH)

        latest_backup = _latest_backup_info()
        latest_verify = _verify_sqlite_file(latest_backup['path']) if latest_backup else {'ok': False, 'integrity': 'missing'}
        latest_age_s = max(0, int(time.time() - int(latest_backup.get('mtime') or 0))) if latest_backup else None
        max_age_s = int((sec.get('backup_max_age_hours') or 0) * 3600)
        has_security_webhook = bool(str(sec.get('security_alert_webhook') or '').strip())
        has_backup_webhook = bool(str(sec.get('backup_webhook') or '').strip())
        auto_thread_ok = bool(_AUTO_BACKUP_THREAD and _AUTO_BACKUP_THREAD.is_alive()) if bool(sec.get('auto_backup_enabled')) else True

        events = _read_audit_tail(300)
        now_dt = datetime.datetime.utcnow()
        rejected_24h = 0
        error_24h = 0
        for ev in events:
            try:
                ts_raw = str(ev.get('ts') or '')
                if not ts_raw:
                    continue
                ts = datetime.datetime.fromisoformat(ts_raw.replace('Z', '+00:00')).replace(tzinfo=None)
                if (now_dt - ts).total_seconds() > 86400:
                    continue
                out = str(ev.get('outcome') or '').lower()
                if out == 'rejected':
                    rejected_24h += 1
                elif out == 'error':
                    error_24h += 1
            except Exception:
                continue

        is_persist_path = lambda p: p.startswith('/var/data')
        checks = []
        def add(cid, label, ok, value=None, recommendation=''):
            checks.append({
                'id': cid,
                'label': label,
                'ok': bool(ok),
                'value': value,
                'recommendation': recommendation,
            })

        add(
            'secret_not_default',
            'SECRET_KEY no usa valor por defecto',
            secret != _DEFAULT_SECRET_KEY,
            value='configured' if secret and secret != _DEFAULT_SECRET_KEY else 'default_or_empty',
            recommendation='Configurar SECRET_KEY larga y aleatoria en Render.'
        )
        add(
            'secret_strength',
            'SECRET_KEY con longitud recomendada (>=32)',
            len(secret) >= 32,
            value=len(secret),
            recommendation='Usar una clave de al menos 32 caracteres.'
        )
        add(
            'bot_key_configured',
            'BOT_KEY configurada',
            len(BOT_KEYS) > 0,
            value=len(BOT_KEYS),
            recommendation='Definir BOT_KEY fuerte (o varias, separadas por coma).'
        )
        add(
            'bot_key_not_default',
            'BOT_KEY sin valor de desarrollo',
            (_DEFAULT_BOT_KEY not in BOT_KEYS),
            value='default_present' if (_DEFAULT_BOT_KEY in BOT_KEYS) else 'ok',
            recommendation='Quitar el BOT_KEY de desarrollo en entornos reales.'
        )
        add(
            'session_cookie_secure',
            'Cookie de sesion segura (HTTPS)',
            bool(app.config.get('SESSION_COOKIE_SECURE')),
            value=bool(app.config.get('SESSION_COOKIE_SECURE')),
            recommendation='Mantener SESSION_COOKIE_SECURE=True en produccion.'
        )
        add(
            'session_cookie_httponly',
            'Cookie de sesion HttpOnly',
            bool(app.config.get('SESSION_COOKIE_HTTPONLY')),
            value=bool(app.config.get('SESSION_COOKIE_HTTPONLY')),
            recommendation='Mantener SESSION_COOKIE_HTTPONLY=True.'
        )
        add(
            'session_cookie_samesite',
            'Cookie SameSite en Lax/Strict',
            str(app.config.get('SESSION_COOKIE_SAMESITE', '')).lower() in ('lax', 'strict'),
            value=app.config.get('SESSION_COOKIE_SAMESITE'),
            recommendation='Usar SameSite Lax o Strict.'
        )
        add(
            'db_persistent_path',
            'Base de datos en ruta persistente',
            is_persist_path(db_path_abs),
            value=db_path_abs,
            recommendation='Usar DB_PATH en /var/data/rc_domotic.db.'
        )
        add(
            'data_dir_persistent',
            'DATA_DIR en almacenamiento persistente',
            is_persist_path(data_dir_abs),
            value=data_dir_abs,
            recommendation='Configurar DATA_DIR=/var/data en Render.'
        )
        add(
            'backup_dir_persistent',
            'BACKUP_DIR en almacenamiento persistente',
            is_persist_path(backup_dir_abs),
            value=backup_dir_abs,
            recommendation='Configurar BACKUP_DIR dentro de /var/data.'
        )
        add(
            'audit_log_persistent',
            'Audit log en almacenamiento persistente',
            is_persist_path(audit_path_abs),
            value=audit_path_abs,
            recommendation='Configurar AUDIT_LOG_PATH dentro de /var/data.'
        )
        add(
            'db_exists',
            'Archivo de base existe',
            os.path.isfile(DB_PATH),
            value=os.path.getsize(DB_PATH) if os.path.isfile(DB_PATH) else 0,
            recommendation='Verificar montaje de disco y ruta de DB_PATH.'
        )
        add(
            'backup_dir_writable',
            'Directorio de backups escribible',
            os.access(BACKUP_DIR, os.W_OK),
            value=backup_dir_abs,
            recommendation='Asegurar permisos de escritura en BACKUP_DIR.'
        )
        add(
            'audit_log_writable',
            'Archivo/directorio de auditoria escribible',
            os.access(os.path.dirname(audit_path_abs) or DATA_DIR, os.W_OK),
            value=audit_path_abs,
            recommendation='Asegurar permisos de escritura de AUDIT_LOG_PATH.'
        )
        add(
            'limiter_enabled',
            'Rate limit habilitado',
            limiter is not None,
            value=bool(limiter is not None),
            recommendation='Instalar flask-limiter en entorno productivo.'
        )
        add(
            'admin_ip_allowlist',
            'Allowlist IP para rutas admin',
            (int(sec.get('admin_ip_allowlist_count') or 0) > 0) if IS_PROD else True,
            value=int(sec.get('admin_ip_allowlist_count') or 0),
            recommendation='Configurar ADMIN_IP_ALLOWLIST con IP(s)/CIDR de acceso admin.'
        )
        add(
            'bot_ip_allowlist',
            'Allowlist IP para rutas bot',
            (int(sec.get('bot_ip_allowlist_count') or 0) > 0) if IS_PROD else True,
            value=int(sec.get('bot_ip_allowlist_count') or 0),
            recommendation='Configurar BOT_IP_ALLOWLIST para restringir peticiones del bot.'
        )
        add(
            'admin_reauth_enabled',
            'Reautenticacion admin para restaurar base',
            int(sec.get('admin_reauth_seconds') or 0) > 0,
            value=int(sec.get('admin_reauth_seconds') or 0),
            recommendation='Definir ADMIN_REAUTH_SECONDS (recomendado 600-1800).'
        )
        add(
            'auto_backup_enabled',
            'Backup automatico habilitado',
            bool(sec.get('auto_backup_enabled')),
            value={
                'enabled': bool(sec.get('auto_backup_enabled')),
                'every_min': int(sec.get('auto_backup_every_min') or 0),
                'keep': int(sec.get('auto_backup_keep') or 0)
            },
            recommendation='Activar AUTO_BACKUP_ENABLED y revisar frecuencia.'
        )
        add(
            'auto_backup_worker_running',
            'Worker de backup automatico activo',
            bool(auto_thread_ok),
            value='running' if auto_thread_ok else 'stopped',
            recommendation='Reiniciar servicio si el worker no aparece activo.'
        )
        add(
            'latest_backup_exists',
            'Existe al menos un backup',
            bool(latest_backup),
            value=(latest_backup or {}).get('name') if latest_backup else 'none',
            recommendation='Ejecutar backup inicial y validar ruta persistente.'
        )
        add(
            'latest_backup_fresh',
            f"Backup reciente (<={int(sec.get('backup_max_age_hours') or 0)}h)",
            bool(latest_backup and latest_age_s is not None and latest_age_s <= max_age_s),
            value={'age_seconds': latest_age_s, 'max_age_seconds': max_age_s},
            recommendation='Reducir AUTO_BACKUP_EVERY_MIN o generar backup manual ahora.'
        )
        add(
            'latest_backup_integrity',
            'Ultimo backup pasa integridad sqlite',
            bool(latest_backup and latest_verify.get('ok')),
            value=latest_verify.get('integrity'),
            recommendation='No restaurar backups con integridad fallida.'
        )
        add(
            'security_webhook',
            'Webhook de alertas de seguridad configurado',
            has_security_webhook if IS_PROD else True,
            value='configured' if has_security_webhook else 'not_set',
            recommendation='Configurar SECURITY_ALERT_WEBHOOK para alertas en tiempo real.'
        )
        add(
            'backup_webhook',
            'Webhook de backup/restore configurado',
            has_backup_webhook if IS_PROD else True,
            value='configured' if has_backup_webhook else 'not_set',
            recommendation='Configurar BACKUP_WEBHOOK_URL para observabilidad externa.'
        )

        total = len(checks)
        ok_n = len([c for c in checks if c['ok']])
        fail_n = total - ok_n
        level = 'ok' if fail_n == 0 else ('warn' if fail_n <= 4 else 'error')
        return jsonify({
            'ok': True,
            'generated_at': now_utc,
            'level': level,
            'summary': {
                'ok': ok_n,
                'fail': fail_n,
                'total': total,
            },
            'env': {
                'is_prod': bool(IS_PROD),
                'render_env': bool(os.environ.get('RENDER') or os.environ.get('RENDER_SERVICE_ID')),
                'data_dir': data_dir_abs,
                'db_path': db_path_abs,
                'backup_dir': backup_dir_abs,
                'audit_log_path': audit_path_abs,
                'max_upload_bytes': int(app.config.get('MAX_CONTENT_LENGTH') or 0),
                'admin_ip_allowlist': int(sec.get('admin_ip_allowlist_count') or 0),
                'bot_ip_allowlist': int(sec.get('bot_ip_allowlist_count') or 0),
                'admin_reauth_seconds': int(sec.get('admin_reauth_seconds') or 0),
                'auto_backup_enabled': bool(sec.get('auto_backup_enabled')),
                'auto_backup_every_min': int(sec.get('auto_backup_every_min') or 0),
                'auto_backup_keep': int(sec.get('auto_backup_keep') or 0),
                'security_backup_max_age_hours': int(sec.get('backup_max_age_hours') or 0),
                'security_alert_webhook': has_security_webhook,
                'backup_webhook': has_backup_webhook,
                'latest_backup': latest_backup,
                'latest_backup_verify': latest_verify,
                'latest_backup_age_seconds': latest_age_s,
                'audit_events_last_24h': {
                    'rejected': int(rejected_24h),
                    'error': int(error_24h),
                },
            },
            'checks': checks,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/admin/security_quick_harden')
@role_required('admin')
def api_security_quick_harden():
    """Aplica ajustes de seguridad recomendados en un solo paso (sin tocar datos)."""
    try:
        current = _runtime_security_settings(force=True)
        updates = {}

        ip = _client_ip().strip()
        try:
            ip_obj = ipaddress.ip_address(ip)
            ip_cidr = f"{ip}/{32 if ip_obj.version == 4 else 128}"
        except Exception:
            ip_cidr = ''

        if not str(current.get('security_alert_webhook') or '').strip():
            updates['security_alert_webhook'] = 'internal://audit'
        if not str(current.get('backup_webhook') or '').strip():
            updates['security_backup_webhook'] = 'internal://audit'

        if int(current.get('admin_ip_allowlist_count') or 0) <= 0 and ip_cidr:
            updates['security_admin_ip_allowlist'] = ip_cidr

        if int(current.get('bot_ip_allowlist_count') or 0) <= 0:
            bot_ips = []
            for ev in _read_audit_tail(800):
                try:
                    p = str(ev.get('path') or '')
                    out = str(ev.get('outcome') or '').lower()
                    ev_ip = str(ev.get('ip') or '').strip()
                    if not p.startswith('/api/bot/') or out != 'ok' or not ev_ip:
                        continue
                    ip_o = ipaddress.ip_address(ev_ip)
                    cidr = f"{ev_ip}/{32 if ip_o.version == 4 else 128}"
                    if cidr not in bot_ips:
                        bot_ips.append(cidr)
                except Exception:
                    continue
            if bot_ips:
                updates['security_bot_ip_allowlist'] = ','.join(bot_ips[:5])
            elif ip_cidr:
                # Fallback seguro: limita bot a la misma IP actual hasta definir IP final de n8n.
                updates['security_bot_ip_allowlist'] = ip_cidr

        if updates:
            for k, v in updates.items():
                execute("INSERT OR REPLACE INTO parametros (clave, valor) VALUES (?,?)", (k, str(v)))
            try:
                _SEC_SETTINGS_CACHE['ts'] = 0.0
                _SEC_SETTINGS_CACHE['data'] = None
            except Exception:
                pass
            # Si auto-backup quedó habilitado por parámetros, inicia el worker.
            _ensure_auto_backup_thread()

        updated = _runtime_security_settings(force=True)
        _audit_event('security_quick_harden', actor=current_user(), changed=list(updates.keys()), current_ip=ip, applied=bool(updates))
        return jsonify({
            'ok': True,
            'updated_keys': list(updates.keys()),
            'settings': {
                'admin_ip_allowlist_count': int(updated.get('admin_ip_allowlist_count') or 0),
                'bot_ip_allowlist_count': int(updated.get('bot_ip_allowlist_count') or 0),
                'security_alert_webhook': bool(str(updated.get('security_alert_webhook') or '').strip()),
                'backup_webhook': bool(str(updated.get('backup_webhook') or '').strip()),
            }
        })
    except Exception as e:
        _audit_event('security_quick_harden', outcome='error', actor=current_user(), error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/admin/db_backup')
@role_required('admin')
def api_db_backup():
    try:
        sec = _runtime_security_settings(force=True)
        out = snapshot_db('admin')
        if not out:
            _audit_event('db_backup', outcome='rejected', reason='no_database_available')
            return jsonify({'ok': False, 'error': 'No existe base para respaldar'}), 404
        verify = _verify_sqlite_file(out)
        removed = _prune_backup_files(int(sec.get('auto_backup_keep') or AUTO_BACKUP_KEEP_DEFAULT))
        size = os.path.getsize(out) if os.path.isfile(out) else 0
        _audit_event(
            'db_backup',
            outcome='ok' if verify.get('ok') else 'error',
            actor=current_user(),
            backup=os.path.basename(out),
            backup_path=out,
            size=size,
            integrity=verify.get('integrity'),
            removed_old=removed,
        )
        if str(sec.get('backup_webhook') or '').strip():
            _post_json_webhook_async(str(sec.get('backup_webhook') or '').strip(), {
                'event': 'manual_backup',
                'ok': bool(verify.get('ok')),
                'backup': {'name': os.path.basename(out), 'path': out, 'size': size},
                'verify': verify,
                'removed_old': removed,
                'actor': (current_user() or {}).get('username') or '',
            })
        return jsonify({
            'ok': True,
            'backup': os.path.basename(out),
            'path': out,
            'size': size,
            'verify': verify,
            'removed_old': removed,
        })
    except Exception as e:
        _audit_event('db_backup', outcome='error', error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.get('/api/admin/db_backup_download')
@role_required('admin')
def api_db_backup_download():
    try:
        sec = _runtime_security_settings(force=True)
        out = snapshot_db('download')
        if not out:
            _audit_event('db_backup_download', outcome='rejected', actor=current_user(), reason='no_database_available')
            return jsonify({'ok': False, 'error': 'No existe base para respaldar'}), 404
        verify = _verify_sqlite_file(out)
        removed = _prune_backup_files(int(sec.get('auto_backup_keep') or AUTO_BACKUP_KEEP_DEFAULT))
        size = os.path.getsize(out) if os.path.isfile(out) else 0
        _audit_event(
            'db_backup_download',
            outcome='ok' if verify.get('ok') else 'error',
            actor=current_user(),
            backup=os.path.basename(out),
            size=size,
            integrity=verify.get('integrity'),
            removed_old=removed,
        )
        return send_file(
            out,
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=os.path.basename(out)
        )
    except Exception as e:
        _audit_event('db_backup_download', outcome='error', actor=current_user(), error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/admin/db_restore_upload')
@role_required('admin')
def api_db_restore_upload():
    """Sube un backup sqlite para poder restaurarlo desde la nube de forma segura."""
    upload_path = ''
    try:
        f = (request.files or {}).get('file')
        if not f or not getattr(f, 'filename', ''):
            _audit_event('db_restore_upload', outcome='rejected', actor=current_user(), reason='missing_file')
            return jsonify({'ok': False, 'error': 'Archivo requerido'}), 400

        original_name = secure_filename(f.filename or '')
        ext = (os.path.splitext(original_name)[1] or '').lower()
        if ext not in ('.db', '.sqlite', '.sqlite3'):
            _audit_event('db_restore_upload', outcome='rejected', actor=current_user(), reason='invalid_extension', filename=original_name)
            return jsonify({'ok': False, 'error': 'Formato no permitido. Usa .db/.sqlite/.sqlite3'}), 400

        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        stored_name = f"restore_upload_{ts}_{uuid.uuid4().hex[:8]}.db"
        upload_path = os.path.join(BACKUP_DIR, stored_name)
        f.save(upload_path)
        if not os.path.isfile(upload_path) or os.path.getsize(upload_path) <= 0:
            _audit_event('db_restore_upload', outcome='rejected', actor=current_user(), reason='empty_upload', filename=original_name)
            return jsonify({'ok': False, 'error': 'El archivo subido está vacío'}), 400

        verify = _verify_sqlite_file(upload_path)
        if not verify.get('ok'):
            try:
                os.remove(upload_path)
            except Exception:
                pass
            _audit_event('db_restore_upload', outcome='rejected', actor=current_user(), reason='integrity_failed', filename=original_name, integrity=verify.get('integrity'))
            return jsonify({'ok': False, 'error': 'Backup inválido (integridad SQLite falló)', 'verify': verify}), 400

        size = os.path.getsize(upload_path)
        _audit_event(
            'db_restore_upload',
            actor=current_user(),
            filename=original_name,
            stored_as=stored_name,
            size=size,
            integrity=verify.get('integrity'),
        )
        return jsonify({
            'ok': True,
            'name': stored_name,
            'path': os.path.abspath(upload_path),
            'size': size,
            'verify': verify,
        })
    except Exception as e:
        if upload_path and os.path.isfile(upload_path):
            try:
                os.remove(upload_path)
            except Exception:
                pass
        _audit_event('db_restore_upload', outcome='error', actor=current_user(), error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/admin/auto_backup_now')
@role_required('admin')
def api_auto_backup_now():
    try:
        result = _run_auto_backup_once(force=True)
        if not result.get('ok'):
            return jsonify({'ok': False, 'error': result.get('error') or result.get('skipped') or 'backup_failed', 'result': result}), 500
        _audit_event('auto_backup_manual_trigger', actor=current_user(), result=result)
        return jsonify({'ok': True, 'result': result})
    except Exception as e:
        _audit_event('auto_backup_manual_trigger', outcome='error', actor=current_user(), error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.get('/api/admin/db_scan')
@role_required('admin')
def api_db_scan():
    try:
        roots = []
        for d in [DATA_DIR, BASE_DIR, '/var/data', '/tmp']:
            if d and os.path.isdir(d):
                roots.append(os.path.abspath(d))
        seen = set()
        found = []
        for root in roots:
            for dirpath, _, filenames in os.walk(root):
                for fn in filenames:
                    if not fn.lower().endswith('.db'):
                        continue
                    fp = os.path.abspath(os.path.join(dirpath, fn))
                    if fp in seen:
                        continue
                    seen.add(fp)
                    try:
                        cot_count, item_count = _db_counts_from_file(fp)
                    except Exception:
                        continue
                    found.append({
                        'path': fp,
                        'name': fn,
                        'size': os.path.getsize(fp),
                        'mtime': int(os.path.getmtime(fp)),
                        'cotizaciones': cot_count,
                        'items': item_count,
                        'is_current': os.path.abspath(fp) == os.path.abspath(DB_PATH),
                    })
        found.sort(key=lambda x: (x['cotizaciones'], x['items'], x['mtime']), reverse=True)
        return jsonify({'ok': True, 'current_db': DB_PATH, 'candidates': found[:100]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.get('/api/admin/backup_verify')
@role_required('admin')
def api_backup_verify():
    try:
        sec = _runtime_security_settings(force=True)
        latest = _latest_backup_info()
        if not latest:
            _audit_event('backup_verify', outcome='rejected', actor=current_user(), reason='no_backups')
            return jsonify({'ok': False, 'error': 'No hay backups disponibles'}), 404
        verify = _verify_sqlite_file(latest['path'])
        age_seconds = max(0, int(time.time() - int(latest.get('mtime') or 0)))
        max_age_seconds = int((sec.get('backup_max_age_hours') or SECURITY_BACKUP_MAX_AGE_HOURS_DEFAULT) * 3600)
        _audit_event(
            'backup_verify',
            outcome='ok' if verify.get('ok') else 'error',
            actor=current_user(),
            backup=latest.get('name'),
            integrity=verify.get('integrity'),
            age_seconds=age_seconds,
        )
        return jsonify({
            'ok': True,
            'backup': latest,
            'verify': verify,
            'age_seconds': age_seconds,
            'max_age_seconds': max_age_seconds,
            'fresh': age_seconds <= max_age_seconds,
        })
    except Exception as e:
        _audit_event('backup_verify', outcome='error', actor=current_user(), error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/admin/db_restore')
@role_required('admin')
def api_db_restore():
    try:
        sec = _runtime_security_settings(force=True)
        data = request.json or {}
        reauth_err = _require_admin_reauth_for_action(data, target_action='db_restore')
        if reauth_err:
            return reauth_err
        src = os.path.abspath(str(data.get('path') or '').strip())
        if not src or not os.path.isfile(src):
            _audit_event('db_restore', outcome='rejected', reason='source_missing', requested_path=src)
            return jsonify({'ok': False, 'error': 'Archivo origen no existe'}), 404
        # Permitir restaurar solo desde rutas locales esperadas
        allowed_roots = [os.path.abspath(x) for x in [DATA_DIR, BASE_DIR, '/var/data', '/tmp'] if x and os.path.isdir(x)]
        if not any(src.startswith(root + os.sep) or src == root for root in allowed_roots):
            _audit_event('db_restore', outcome='rejected', reason='forbidden_source', requested_path=src)
            return jsonify({'ok': False, 'error': 'Ruta origen no permitida'}), 403
        source_verify = _verify_sqlite_file(src)
        if not source_verify.get('ok'):
            _audit_event('db_restore', outcome='rejected', reason='source_integrity_failed', requested_path=src, integrity=source_verify.get('integrity'))
            return jsonify({'ok': False, 'error': 'Backup origen no paso integridad sqlite', 'verify': source_verify}), 400
        cot_count, item_count = _db_counts_from_file(src)
        backup_prev = snapshot_db('pre_restore')
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        src_conn = sqlite3.connect(src)
        try:
            dst_conn = sqlite3.connect(DB_PATH)
            try:
                src_conn.backup(dst_conn)
                dst_conn.commit()
            finally:
                dst_conn.close()
        finally:
            src_conn.close()
        restored_verify = _verify_sqlite_file(DB_PATH)
        _audit_event(
            'db_restore',
            outcome='ok' if restored_verify.get('ok') else 'error',
            restored_from=src,
            backup_previous=backup_prev,
            cotizaciones=cot_count,
            items=item_count,
            source_integrity=source_verify.get('integrity'),
            restored_integrity=restored_verify.get('integrity'),
        )
        if str(sec.get('backup_webhook') or '').strip():
            _post_json_webhook_async(str(sec.get('backup_webhook') or '').strip(), {
                'event': 'db_restore',
                'ok': bool(restored_verify.get('ok')),
                'source': src,
                'backup_previous': backup_prev,
                'verify': restored_verify,
                'actor': (current_user() or {}).get('username') or '',
            })
        return jsonify({
            'ok': True,
            'restored_from': src,
            'backup_previous': backup_prev,
            'cotizaciones': cot_count,
            'items': item_count,
            'db_path': DB_PATH,
            'verify': restored_verify,
        })
    except Exception as e:
        _audit_event('db_restore', outcome='error', error=str(e))
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.get('/api/admin/security_events')
@role_required('admin')
def api_security_events():
    try:
        try:
            limit_n = int(request.args.get('limit', 50))
        except Exception:
            limit_n = 50
        limit_n = max(1, min(limit_n, 200))
        events = _read_audit_tail(limit_n)
        _audit_event('security_events_view', actor=current_user(), limit=limit_n, returned=len(events))
        return jsonify({'ok': True, 'path': AUDIT_LOG_PATH, 'count': len(events), 'events': events})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.put('/api/parametros')
@role_required('admin')
def update_parametros():
    for k, v in request.json.items():
        execute("INSERT OR REPLACE INTO parametros (clave, valor) VALUES (?,?)", (k, str(v)))
    # Refresca cache de seguridad cuando cambian parametros.
    try:
        _SEC_SETTINGS_CACHE['ts'] = 0.0
        _SEC_SETTINGS_CACHE['data'] = None
    except Exception:
        pass
    _ensure_auto_backup_thread()
    return jsonify({'ok': True})


# â”€â”€â”€ Listas de precios (perfil) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/price_lists')
@role_required('admin')
def api_price_lists():
    rows = query("SELECT * FROM price_lists WHERE active=1 ORDER BY desc_pct, code")
    return jsonify(rows)

@app.post('/api/price_lists')
@role_required('admin')
def api_price_lists_create():
    d = request.json or {}
    code = (d.get('code') or '').strip().upper()
    name = (d.get('name') or '').strip()
    desc_pct = float(d.get('desc_pct') or 0)
    if not code or not name:
        return jsonify({'ok': False, 'error': 'code y name requeridos'}), 400
    execute("INSERT OR REPLACE INTO price_lists(code,name,desc_pct,active) VALUES (?,?,?,1)", (code, name, desc_pct))
    return jsonify({'ok': True})

@app.put('/api/price_lists/<code>')
@role_required('admin')
def api_price_lists_update(code):
    d = request.json or {}
    name = (d.get('name') or '').strip()
    desc_pct = float(d.get('desc_pct') or 0)
    active = 1 if str(d.get('active','1')) == '1' else 0
    execute("UPDATE price_lists SET name=?, desc_pct=?, active=? WHERE code=?", (name, desc_pct, active, code.upper()))
    return jsonify({'ok': True})
@app.get('/api/cotizaciones/<int:cot_id>/margenes')
def get_margenes(cot_id):
    m = calcular_margenes(cot_id)
    return jsonify(m) if m else ('Not found', 404)

@app.get('/api/cotizaciones/<int:cot_id>/whatsapp')
def whatsapp_link(cot_id):
    """Genera enlace de WhatsApp al nÃºmero del cliente con un mensaje SIN valores (solo link pÃºblico)."""
    import urllib.parse
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return ('Not found', 404)

    # Normalizar telÃ©fono (Colombia): si viene 10 dÃ­gitos, prefijar 57
    tel = re.sub(r'\D', '', (cot.get('telefono') or ''))
    if tel:
        if len(tel) == 10:
            tel = '57' + tel
        elif len(tel) == 11 and tel.startswith('3'):
            tel = '57' + tel
        elif not tel.startswith('57') and len(tel) < 12:
            tel = '57' + tel

    token = ensure_public_token(cot_id)
    base = request.host_url.rstrip('/')
    public_url = f"{base}/q/{token}" if token else base
    pdf_url = f"{public_url}/pdf"

    proyecto = cot.get('proyecto') or 'N/A'
    msg = (
        f"Hola {cot.get('cliente','')}!\n\n"
        f"Te comparto tu cotizaciÃ³n de RC DOMOTIC para {proyecto}.\n"
        f"NÂ° {cot.get('no_cotizacion','')}\n\n"
        f"PDF: {pdf_url}\n"
        f"Ver y aceptar en web: {public_url}\n\n"
        f"Quedo atento.\n"
        f"RaÃºl Cuello â€” RC DOMOTIC\n"
        f"Tel: 3123042156"
    )

    q = urllib.parse.quote(msg)
    wa_url = f"https://wa.me/{tel}?text={q}" if tel else f"https://wa.me/?text={q}"
    return jsonify({'url': wa_url, 'mensaje': msg})



@app.get('/api/cotizaciones/<int:cot_id>/share')
def cot_share(cot_id):
    """Devuelve enlaces listos para copiar/WhatsApp/Email.

    Regla: el mensaje NO incluye valores (ni total, ni anticipo). Los valores se ven en el PDF o en el link.
    """
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return ('Not found', 404)

    token = ensure_public_token(cot_id)
    base = request.host_url.rstrip('/')
    import urllib.parse

    no = (cot.get('no_cotizacion') or '').strip()
    no_q = urllib.parse.quote(no, safe='')
    if no:
        public_url = f"{base}/cotizacion/{no_q}"
    else:
        public_url = f"{base}/q/{token}"

    pdf_url = f"{public_url}/pdf"
    accept_url = f"{public_url}/accept"

    # Mensaje sin valores (amigable para WhatsApp)
    cliente = (cot.get('cliente') or '').strip()
    proyecto = (cot.get('proyecto') or '').strip() or 'N/A'
    resumen = (
        f"Hola {cliente}!\n\n"
        f"Te comparto tu cotizaciÃ³n de RC DOMOTIC para {proyecto}.\n"
        f"NÂ° {no}\n\n"
        f"PDF: {pdf_url}\n"
        f"Ver y aceptar en web: {public_url}\n\n"
        f"Quedo atento.\n"
        f"RaÃºl Cuello â€” RC DOMOTIC\n"
        f"Tel: 3123042156"
    )

    tel = re.sub(r'\D', '', cot.get('telefono', '') or '')
    if tel and not tel.startswith('57'):
        tel = '57' + tel
    wa_url = f"https://wa.me/{tel}?text={urllib.parse.quote(resumen)}" if tel else f"https://wa.me/?text={urllib.parse.quote(resumen)}"

    subj = f"CotizaciÃ³n {no or ''} - RC DOMOTIC".strip()
    body = resumen
    mailto = f"mailto:?subject={urllib.parse.quote(subj)}&body={urllib.parse.quote(body)}"

    return jsonify({
        'ok': True,
        'public_url': public_url,
        'pdf_url': pdf_url,
        'accept_url': accept_url,
        'whatsapp_url': wa_url,
        'mailto': mailto,
        'resumen': resumen
    })

# â”€â”€â”€ DASHBOARD STATS (solo APROBADA para ventas) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/stats')
@role_required('admin')
def get_stats():
    total_cots = query("SELECT COUNT(*) as n FROM cotizaciones", one=True)['n']
    por_estado = query("SELECT estado, COUNT(*) as n FROM cotizaciones GROUP BY estado")
    top_prods = query("""SELECT c.nombre, SUM(i.cantidad) as total_cant
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         GROUP BY i.id_producto ORDER BY total_cant DESC LIMIT 5""")
    return jsonify({'total_cotizaciones': total_cots,
                    'por_estado': {r['estado']: r['n'] for r in por_estado},
                    'top_productos': [dict(r) for r in top_prods]})

@app.get('/api/stats/dashboard')
@role_required('admin')
def get_dashboard():
    """Dashboard avanzado: ventas SOLO de cotizaciones APROBADA."""
    now = datetime.date.today()
    mes_actual = now.strftime('%Y-%m')
    mes_ant = (now.replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m')
    # Ventas mes actual
    r = query("""SELECT COALESCE(SUM(total),0) as venta, COUNT(*) as n FROM (
        SELECT c.id, (SELECT SUM(calc) FROM (
            SELECT CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END * ci.cantidad
                   + CASE WHEN cat.aplica_iva THEN (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad*cat.pct_iva ELSE 0 END
                   + (CASE WHEN ci.inst_manual>0 THEN ci.inst_manual ELSE cat.inst_default END)*ci.cantidad
                   + CASE WHEN ci.cfg_manual>0 THEN ci.cfg_manual ELSE cat.config_default END as calc
            FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto WHERE ci.cot_id=c.id
        )) as total
        FROM cotizaciones c WHERE c.estado='APROBADA' AND c.fecha LIKE ?||'%'
    ) WHERE total IS NOT NULL""", (mes_actual,), one=True)
    venta_mes = r['venta'] if r else 0; n_mes = r['n'] if r else 0

    r2 = query("""SELECT COALESCE(SUM(total),0) as venta, COUNT(*) as n FROM (
        SELECT c.id, (SELECT SUM(calc) FROM (
            SELECT CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END * ci.cantidad
                   + CASE WHEN cat.aplica_iva THEN (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad*cat.pct_iva ELSE 0 END
                   + (CASE WHEN ci.inst_manual>0 THEN ci.inst_manual ELSE cat.inst_default END)*ci.cantidad
                   + CASE WHEN ci.cfg_manual>0 THEN ci.cfg_manual ELSE cat.config_default END as calc
            FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto WHERE ci.cot_id=c.id
        )) as total
        FROM cotizaciones c WHERE c.estado='APROBADA' AND c.fecha LIKE ?||'%'
    ) WHERE total IS NOT NULL""", (mes_ant,), one=True)
    venta_ant = r2['venta'] if r2 else 0

    # Top productos por valor (mes actual, APROBADA)
    top_val = query("""SELECT cat.nombre, cat.id_producto,
        SUM((CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad) as valor
        FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto
        JOIN cotizaciones c ON ci.cot_id=c.id
        WHERE c.estado='APROBADA' AND c.fecha LIKE ?||'%'
        GROUP BY ci.id_producto ORDER BY valor DESC LIMIT 10""", (mes_actual,))

    # Top productos por cantidad
    top_cant = query("""SELECT cat.nombre, cat.id_producto, SUM(ci.cantidad) as cant
        FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto
        JOIN cotizaciones c ON ci.cot_id=c.id
        WHERE c.estado='APROBADA' AND c.fecha LIKE ?||'%'
        GROUP BY ci.id_producto ORDER BY cant DESC LIMIT 10""", (mes_actual,))

    # Top clientes
    top_cli = query("""SELECT c.cliente, COUNT(*) as n_cots,
        SUM(total) as valor FROM (
            SELECT c2.id, c2.cliente, (SELECT SUM(
                (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad
                + CASE WHEN cat.aplica_iva THEN (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad*cat.pct_iva ELSE 0 END
                + (CASE WHEN ci.inst_manual>0 THEN ci.inst_manual ELSE cat.inst_default END)*ci.cantidad
                + CASE WHEN ci.cfg_manual>0 THEN ci.cfg_manual ELSE cat.config_default END
            ) FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto WHERE ci.cot_id=c2.id) as total
            FROM cotizaciones c2 WHERE c2.estado='APROBADA' AND c2.fecha LIKE ?||'%'
        ) c WHERE total IS NOT NULL
        GROUP BY c.cliente ORDER BY valor DESC LIMIT 10""", (mes_actual,))

    # ProyecciÃ³n
    dia_hoy = now.day
    prom_diario = venta_mes / dia_hoy if dia_hoy > 0 else 0
    import calendar
    dias_mes = calendar.monthrange(now.year, now.month)[1]
    proj_conservador = round(prom_diario * dias_mes * 0.85)
    proj_probable = round(prom_diario * dias_mes)
    proj_agresivo = round(prom_diario * dias_mes * 1.20)

    return jsonify({
        'venta_mes': round(venta_mes), 'venta_ant': round(venta_ant),
        'n_aprobadas_mes': n_mes, 'mes_actual': mes_actual, 'mes_anterior': mes_ant,
        'top_valor': [dict(r) for r in top_val],
        'top_cantidad': [dict(r) for r in top_cant],
        'top_clientes': [dict(r) for r in top_cli],
        'proyeccion': {'conservador': proj_conservador, 'probable': proj_probable, 'agresivo': proj_agresivo}
    })


@app.get('/api/stats/series')
@role_required('admin')
def get_stats_series():
    """Series para el grÃ¡fico de Ventas.

    Params:
      - period: week | month | quarter | year   (default: month)
      - scope: APROBADA | ALL                   (default: APROBADA)

    Nota: calculamos el total igual que en el dashboard (precio + IVA + inst + cfg),
    sin depender de columnas materializadas.
    """
    try:
        period = (request.args.get('period') or 'month').strip().lower()
        scope = (request.args.get('scope') or 'APROBADA').strip().upper()

        today = datetime.date.today()

        def total_for(extra_where_sql:str, params:tuple):
            where_parts = []
            if scope != 'ALL':
                where_parts.append("c.estado='APROBADA'")
            if extra_where_sql:
                where_parts.append(extra_where_sql)
            where_sql = " AND ".join(where_parts) if where_parts else "1=1"

            r = query(f"""SELECT COALESCE(SUM(total),0) as t FROM (
                SELECT c.id,
                (SELECT SUM(
                    (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad
                    + CASE WHEN cat.aplica_iva THEN (CASE WHEN ci.precio_manual>0 THEN ci.precio_manual ELSE cat.precio END)*ci.cantidad*cat.pct_iva ELSE 0 END
                    + (CASE WHEN ci.inst_manual>0 THEN ci.inst_manual ELSE cat.inst_default END)*ci.cantidad
                    + CASE WHEN ci.cfg_manual>0 THEN ci.cfg_manual ELSE cat.config_default END
                )
                FROM items ci JOIN catalogo cat ON ci.id_producto=cat.id_producto
                WHERE ci.cot_id=c.id) as total
                FROM cotizaciones c
                WHERE {where_sql}
            ) WHERE total IS NOT NULL""", params, one=True)
            return float((r or {}).get('t', 0) or 0)

        dias = []

        if period == 'week':
            for i in range(6, -1, -1):
                d = today - datetime.timedelta(days=i)
                ds = d.strftime('%Y-%m-%d')
                dias.append({'dia': ds, 'total': total_for('c.fecha=?', (ds,))})

        elif period == 'quarter':
            # 13 semanas (incluye semana actual). Etiqueta: lunes de cada semana.
            mon = today - datetime.timedelta(days=today.weekday())
            for i in range(12, -1, -1):
                ws = mon - datetime.timedelta(weeks=i)
                we = ws + datetime.timedelta(days=6)
                dias.append({
                    'dia': ws.strftime('%Y-%m-%d'),
                    'total': total_for('c.fecha BETWEEN ? AND ?', (ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d')))
                })

        elif period == 'year':
            # 12 meses (incluye el mes actual). Etiqueta: YYYY-MM
            first = today.replace(day=1)
            for i in range(11, -1, -1):
                yy = first.year
                mm = first.month - i
                while mm <= 0:
                    mm += 12
                    yy -= 1
                ms = f"{yy:04d}-{mm:02d}"
                dias.append({'dia': ms, 'total': total_for('c.fecha LIKE ?', (ms + '%',))})

        else:
            # month (default): Ãºltimos 30 dÃ­as
            for i in range(29, -1, -1):
                d = today - datetime.timedelta(days=i)
                ds = d.strftime('%Y-%m-%d')
                dias.append({'dia': ds, 'total': total_for('c.fecha=?', (ds,))})

        return jsonify({'dias': dias, 'period': period, 'scope': scope})
    except Exception as e:
        return jsonify({'dias': [], 'error': str(e), 'ok': False})


@app.get('/api/stats/bi')
@role_required('admin')
def get_stats_bi():
    """BI para Inicio (estilo dashboard).

    Params:
      - scope: APROBADA | ALL     (default: APROBADA)
      - period: week|month|quarter|year|all (default: year)

    Nota: inventario/SKUs/categorÃ­as se calculan sobre catÃ¡logo activo (no dependen del periodo).
    """
    scope = (request.args.get('scope') or 'APROBADA').strip().upper()
    period = (request.args.get('period') or 'year').strip().lower()

    # Rango de fechas
    today = datetime.date.today()
    start = None
    if period == 'week':
        start = today - datetime.timedelta(days=6)
    elif period == 'month':
        start = today - datetime.timedelta(days=29)
    elif period == 'quarter':
        start = today - datetime.timedelta(days=89)
    elif period == 'year':
        start = today - datetime.timedelta(days=364)
    elif period == 'all':
        start = None

    where_parts = []
    params = []
    if scope != 'ALL':
        where_parts.append("c.estado='APROBADA'")
    if start:
        where_parts.append("c.fecha BETWEEN ? AND ?")
        params.extend([start.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d')])

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    # ventas, unidades y costo (solo costo de productos)
    rows = query(f"""
      SELECT c.id as cot_id,
             c.descuento_pct as descuento_pct,
             SUM((CASE WHEN i.precio_manual>0 THEN i.precio_manual ELSE cat.precio END)*i.cantidad) as venta_prod,
             SUM(i.cantidad) as unidades,
             SUM(COALESCE(cat.costo_unitario,0)*i.cantidad) as costo_prod,
             SUM((CASE WHEN i.inst_manual>0 THEN i.inst_manual ELSE cat.inst_default END)*i.cantidad) as inst_total,
             SUM((CASE WHEN i.cfg_manual>0 THEN i.cfg_manual ELSE cat.config_default END)) as cfg_total
      FROM cotizaciones c
      JOIN items i ON i.cot_id=c.id
      JOIN catalogo cat ON cat.id_producto=i.id_producto
      {where_sql}
      GROUP BY c.id
    """, tuple(params))

    venta_total = 0.0
    costo_total = 0.0
    unidades_total = 0.0

    for r in rows:
        venta_cot = float((r.get('venta_prod') or 0) + (r.get('inst_total') or 0) + (r.get('cfg_total') or 0))
        desc = float(r.get('descuento_pct') or 0) / 100.0
        venta_total += venta_cot * (1 - desc)
        costo_total += float(r.get('costo_prod') or 0)
        unidades_total += float(r.get('unidades') or 0)

    beneficio = venta_total - costo_total
    margen_pct = (beneficio / venta_total * 100) if venta_total > 0 else 0

    # inventario
    inv = query("SELECT COALESCE(SUM(stock_qty),0) as u FROM catalogo WHERE activo=1", one=True) or {'u': 0}
    unidades_inv = float(inv.get('u') or 0)
    skus = (query("SELECT COUNT(*) as n FROM catalogo WHERE activo=1", one=True) or {'n': 0})['n']
    cats = (query("SELECT COUNT(DISTINCT categoria) as n FROM catalogo WHERE activo=1", one=True) or {'n': 0})['n']

    # ventas por "departamento" (categorÃ­a) â€” del periodo
    dept = query(f"""
      SELECT cat.categoria as dept,
             SUM((CASE WHEN i.precio_manual>0 THEN i.precio_manual ELSE cat.precio END)*i.cantidad
                 + (CASE WHEN i.inst_manual>0 THEN i.inst_manual ELSE cat.inst_default END)*i.cantidad
                 + (CASE WHEN i.cfg_manual>0 THEN i.cfg_manual ELSE cat.config_default END)
             ) as venta,
             SUM(COALESCE(cat.costo_unitario,0)*i.cantidad) as costo,
             SUM(i.cantidad) as unidades
      FROM cotizaciones c
      JOIN items i ON i.cot_id=c.id
      JOIN catalogo cat ON cat.id_producto=i.id_producto
      {where_sql}
      GROUP BY cat.categoria
      ORDER BY venta DESC
    """, tuple(params))

    dept_rows = []
    for d in dept:
        v = float(d.get('venta') or 0)
        cst = float(d.get('costo') or 0)
        util = v - cst
        m = (util / v * 100) if v > 0 else 0
        dept_rows.append({
            'dept': d.get('dept') or 'N/A',
            'venta': round(v),
            'costo': round(cst),
            'beneficio': round(util),
            'margen_pct': round(m, 1),
            'unidades': float(d.get('unidades') or 0)
        })

    return jsonify({
        'ventas_netas': round(venta_total),
        'beneficio': round(beneficio),
        'margen_pct': round(margen_pct, 1),
        'unidades_vendidas': round(unidades_total),
        'unidades_inventario': round(unidades_inv),
        'categorias': cats,
        'skus': skus,
        'dept': dept_rows,
        'period': period,
        'date_from': start.strftime('%Y-%m-%d') if start else None,
        'date_to': today.strftime('%Y-%m-%d')
    })

    return jsonify({
        'ventas_netas': round(venta_total),
        'beneficio': round(beneficio),
        'margen_pct': round(margen_pct, 1),
        'unidades_vendidas': round(unidades_total),
        'unidades_inventario': round(unidades_inv),
        'categorias': cats,
        'skus': skus,
        'dept': dept_rows
    })

@app.get('/api/stats/ops')
@role_required('admin')
def get_stats_ops():
    """Dashboard operacional: pipeline, alertas, meta, actividad reciente."""
    try:
        import datetime as dt_mod
        today = dt_mod.date.today()
        mes_ini = today.replace(day=1).strftime('%Y-%m-%d')

        # Pipeline por estado
        pipeline_rows = query(
            "SELECT estado, COUNT(*) as cnt, COALESCE(SUM(total_final),0) as total "
            "FROM cotizaciones GROUP BY estado ORDER BY estado"
        )

        # Por cobrar: aprobadas con saldo pendiente
        por_cobrar_row = query(
            "SELECT COALESCE(SUM(total_final - COALESCE(abonado_val,0)),0) as v "
            "FROM cotizaciones WHERE estado='APROBADA'", one=True)
        por_cobrar = float((por_cobrar_row or {}).get('v') or 0)

        # Vendido este mes (aprobadas)
        venta_mes_row = query(
            "SELECT COALESCE(SUM(total_final),0) as v FROM cotizaciones "
            "WHERE estado='APROBADA' AND fecha >= ?", (mes_ini,), one=True)
        venta_mes = float((venta_mes_row or {}).get('v') or 0)

        # Vigencia configurada
        vig_row = query("SELECT valor FROM parametros WHERE clave='vigencia_dias'", one=True)
        vigencia = int(vig_row['valor']) if vig_row and vig_row.get('valor') else 8

        # Meta mensual
        meta_row = query("SELECT valor FROM parametros WHERE clave='meta_ventas_mes'", one=True)
        meta_mes = float(meta_row['valor']) if meta_row and meta_row.get('valor') else 0

        # Alertas: enviadas sin respuesta hace >3 dÃ­as
        sin_resp = query(
            "SELECT id, no_cotizacion, cliente, fecha, total_final FROM cotizaciones "
            "WHERE estado='ENVIADA' AND date(fecha) <= date('now','-3 days') "
            "ORDER BY fecha ASC LIMIT 10")

        # Alertas: por vencer (borrador/enviada, creadas hace > vigencia-2 dÃ­as)
        umbral = max(vigencia - 2, 1)
        por_vencer = query(
            f"SELECT id, no_cotizacion, cliente, fecha, total_final FROM cotizaciones "
            f"WHERE estado IN ('BORRADOR','ENVIADA') "
            f"AND date(fecha) <= date('now','-{umbral} days') "
            f"ORDER BY fecha ASC LIMIT 10")

        # Recientemente aprobadas (Ãºltimas 5, este mes)
        recien_aprobadas = query(
            "SELECT id, no_cotizacion, cliente, fecha, total_final FROM cotizaciones "
            "WHERE estado='APROBADA' ORDER BY id DESC LIMIT 5")

        # Actividad reciente (Ãºltimas 8 cotizaciones por id)
        actividad = query(
            "SELECT id, no_cotizacion, cliente, fecha, total_final, estado "
            "FROM cotizaciones ORDER BY id DESC LIMIT 8")

        total_alertas = len(sin_resp) + len(por_vencer)

        return jsonify({
            'ok': True,
            'pipeline': pipeline_rows,
            'por_cobrar': por_cobrar,
            'venta_mes': venta_mes,
            'meta_mes': meta_mes,
            'vigencia_dias': vigencia,
            'alertas_sin_respuesta': sin_resp,
            'alertas_por_vencer': por_vencer,
            'recien_aprobadas': recien_aprobadas,
            'actividad': actividad,
            'total_alertas': total_alertas,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# â”€â”€â”€ COMANDOS (SEGURO â€” no SQL libre) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.post('/api/commands/create_quote')
def cmd_create_quote():
    """Crear cotizaciÃ³n desde JSON validado. SIN SQL libre."""
    payload = request.json or {}
    payload_str = json.dumps(payload, ensure_ascii=False)
    try:
        # Validaciones
        errors = []
        cliente = (payload.get('cliente') or '').strip()
        if not cliente: errors.append("'cliente' es requerido")
        desc_pct = float(payload.get('descuento_pct', 0))
        ant_pct = float(payload.get('anticipo_pct', 70))
        if not (0 <= desc_pct <= 100): errors.append("descuento_pct debe estar entre 0 y 100")
        if not (0 <= ant_pct <= 100): errors.append("anticipo_pct debe estar entre 0 y 100")
        raw_items = payload.get('items', [])
        if not raw_items: errors.append("'items' no puede estar vacÃ­o")

        # Validar Ã­tems y merge duplicados
        merged = {}
        missing = []
        for it in raw_items:
            cod = (it.get('codigo') or '').strip().upper()
            cant = float(it.get('cant', 1))
            if cant < 1: errors.append(f"cant de '{cod}' debe ser >= 1")
            inst = float(it.get('inst', 0))
            cfg = float(it.get('cfg', 0))
            if inst < 0 or cfg < 0: errors.append(f"inst/cfg de '{cod}' no pueden ser negativos")
            # Verificar que existe en catÃ¡logo
            prod = query("SELECT * FROM catalogo WHERE id_producto=? AND activo=1", (cod,), one=True)
            if not prod:
                missing.append(cod)
                continue
            if cod in merged:
                merged[cod]['cantidad'] += cant  # MERGE: sumar cantidad
            else:
                merged[cod] = {'id_producto': cod, 'cantidad': cant,
                               'precio_manual': 0, 'inst_manual': inst, 'cfg_manual': cfg, 'notas_item': ''}

        if missing: errors.append(f"CÃ³digos no encontrados en catÃ¡logo: {', '.join(missing)}")
        if errors:
            execute("INSERT INTO commands_log (payload, status, error_msg) VALUES (?,?,?)",
                    (payload_str, 'ERROR', '; '.join(errors)))
            return jsonify({'ok': False, 'errors': errors}), 400

        # Crear cotizaciÃ³n
        items_list = list(merged.values())
        data = {
            'cliente': cliente,
            'empresa': payload.get('empresa', ''),
            'nit_cc': payload.get('nit_cc', ''),
            'telefono': payload.get('telefono', ''),
            'email_cliente': payload.get('email', ''),
            'direccion': payload.get('direccion', ''),
            'ciudad': payload.get('ciudad', ''),
            'proyecto': payload.get('proyecto', ''),
            'tipo_cotizacion': payload.get('tipo', 'MIXTA'),
            'forma_pago': payload.get('forma_pago', '70% - 30%'),
            'anticipo_pct': ant_pct / 100,
            'descuento_pct': desc_pct / 100,
            'descuento_val': 0,
            'notas': payload.get('notas', ''),
            'vendedor': payload.get('vendedor', 'Admin'),
            'items': items_list,
        }
        no_cot = next_no_cotizacion()
        # Por defecto: al crear desde comando queda en etapa COTIZADA (pipeline) y estado BORRADOR
        cot_id = execute("""INSERT INTO cotizaciones
            (no_cotizacion,cliente,empresa,nit_cc,telefono,email_cliente,direccion,ciudad,
             proyecto,tipo_cotizacion,forma_pago,anticipo_pct,descuento_pct,descuento_val,notas,vendedor,etapa,estado)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (no_cot, data['cliente'], data['empresa'], data['nit_cc'], data['telefono'],
             data['email_cliente'], data['direccion'], data['ciudad'], data['proyecto'],
             data['tipo_cotizacion'], data['forma_pago'], data['anticipo_pct'],
             data['descuento_pct'], 0, data['notas'], data['vendedor'], 'COTIZADA', 'BORRADOR'))
        for idx, it in enumerate(items_list, 1):
            execute("""INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (cot_id, idx, it['id_producto'], it['cantidad'], it['precio_manual'],
                     it['inst_manual'], it['cfg_manual'], it['notas_item']))

        execute("INSERT INTO commands_log (payload, status, cot_id) VALUES (?,?,?)",
                (payload_str, 'OK', cot_id))
        return jsonify({'ok': True, 'id': cot_id, 'no_cotizacion': no_cot,
                        'items_merged': len(items_list), 'message': f'CotizaciÃ³n {no_cot} creada'})

    except Exception as e:
        tb = traceback.format_exc()
        print(f"[CMD ERROR] {tb}")
        execute("INSERT INTO commands_log (payload, status, error_msg) VALUES (?,?,?)",
                (payload_str, 'ERROR', str(e)))
        return jsonify({'ok': False, 'errors': [str(e)]}), 500

# â”€â”€â”€ EXCEL EXPORT (CON TRY/EXCEPT ROBUSTO) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/export/excel/<int:cot_id>')
@login_required
def export_excel(cot_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
        if not cot: return ('Not found', 404)
        pl = float(cot.get('price_list_desc_pct') or 0)
        items = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                         c.aplica_iva, c.pct_iva, c.inst_default, c.config_default,
                         COALESCE(c.costo_unitario,0) as costo_unitario
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         WHERE i.cot_id=? ORDER BY i.linea""", (cot_id,))
        tots = calcular_cotizacion(cot_id)
        marg = calcular_margenes(cot_id)
        params = {r['clave']: r['valor'] for r in query("SELECT * FROM parametros")}

        wb = Workbook()
        hf = Font(bold=True, color="FFFFFF", size=10)
        hfl = PatternFill("solid", fgColor="0F0F0F")
        mf = '#,##0'
        tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        def sh(ws, row, n):
            for c in range(1, n+1):
                cl = ws.cell(row=row, column=c)
                cl.font = hf; cl.fill = hfl
                cl.alignment = Alignment(horizontal='center', wrap_text=True); cl.border = tb

        def aw(ws):
            for col in ws.columns:
                mx = 0; lt = col[0].column_letter
                for cl in col:
                    try: mx = max(mx, len(str(cl.value or '')))
                    except: pass
                ws.column_dimensions[lt].width = min(mx + 3, 40)

        # HOJA 1
        ws1 = wb.active; ws1.title = "Cotizacion"
        info = [("Empresa:", params.get('empresa','')), ("NIT:", params.get('nit','')), ("",""),
                ("CotizaciÃ³n:", cot.get('no_cotizacion','')), ("Fecha:", cot.get('fecha','')),
                ("Cliente:", cot.get('cliente','')), ("Empresa:", cot.get('empresa','')),
                ("Proyecto:", cot.get('proyecto','')), ("TelÃ©fono:", cot.get('telefono','')),
                ("Ciudad:", cot.get('ciudad',''))]
        for i, (l, v) in enumerate(info, 1):
            ws1.cell(row=i, column=1, value=l).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=str(v or ''))

        r = len(info) + 2
        hs = ["#","CÃ³digo","Producto","DescripciÃ³n","Und","Cant.","P.Unit.","IVA","Inst.","Config.","Total"]
        for c, h in enumerate(hs, 1): ws1.cell(row=r, column=c, value=h)
        sh(ws1, r, len(hs)); r += 1

        for idx, it in enumerate(items, 1):
            calc = calcular_item(it, it, pl)
            vs = [idx, str(it.get('id_producto','')), str(it.get('nombre','')),
                  str(it.get('descripcion','')), str(it.get('unidad','')),
                  float(it.get('cantidad',0)), float(calc['precio_final']),
                  float(calc['iva_monto']), float(calc['inst_final']),
                  float(calc['cfg_final']), float(calc['total_item'])]
            for c, v in enumerate(vs, 1):
                cl = ws1.cell(row=r, column=c, value=v); cl.border = tb
                if c >= 7: cl.number_format = mf
            r += 1

        r += 1
        for l, v in [("Subtotal:", tots.get('total_bruto',0)), ("Descuento:", -tots.get('descuento',0)),
                      ("TOTAL COP:", tots.get('total_final',0)),
                      (f"Anticipo ({int(cot.get('anticipo_pct',0)*100)}%):", tots.get('anticipo_val',0)),
                      ("Saldo:", tots.get('saldo_val',0))]:
            ws1.cell(row=r, column=9, value=l).font = Font(bold=True)
            cl = ws1.cell(row=r, column=11, value=float(v or 0))
            cl.number_format = mf; cl.font = Font(bold=True); r += 1
        aw(ws1)

        # HOJA 2
        if marg:
            ws2 = wb.create_sheet("Interno"); ws2.sheet_properties.tabColor = "25D366"
            h2 = ["#","CÃ³digo","Producto","Cant.","P.Venta","Costo Unit.","Costo Item","Venta Item","Utilidad","Margen %"]
            for c, h in enumerate(h2, 1): ws2.cell(row=1, column=c, value=h)
            sh(ws2, 1, len(h2)); r2 = 2
            for idx, mi in enumerate(marg.get('items',[]), 1):
                vs = [idx, str(mi.get('id_producto','')), str(mi.get('nombre','')),
                      float(mi.get('cantidad',0)), float(mi.get('precio_unitario',0)),
                      float(mi.get('costo_unitario',0)), float(mi.get('costo_item',0)),
                      float(mi.get('venta_item',0)), float(mi.get('utilidad_item',0)),
                      float(mi.get('margen_item',0))]
                for c, v in enumerate(vs, 1):
                    cl = ws2.cell(row=r2, column=c, value=v); cl.border = tb
                    if c in (5,6,7,8,9): cl.number_format = mf
                r2 += 1
            r2 += 2
            for l, v in [("Base Bruta:", marg.get('base_total_bruta',0)),
                          ("Descuento:", marg.get('descuento',0)),
                          ("Total Final:", marg.get('total_final',0)),
                          ("IVA:", marg.get('iva_total',0)),
                          ("Base sin IVA:", marg.get('base_sin_iva',0)),
                          ("Costo:", marg.get('costo_productos_total',0)),
                          ("UTILIDAD:", marg.get('utilidad_neta',0)),
                          ("",""),
                          ("Margen Neto sin IVA:", f"{marg.get('margen_neto_sin_iva',0)}%"),
                          ("Margen s/Total:", f"{marg.get('margen_sobre_total',0)}%"),
                          ("Margen Producto:", f"{marg.get('margen_solo_producto',0)}%"),
                          ("Markup:", f"{marg.get('markup',0)}%")]:
                ws2.cell(row=r2, column=1, value=l).font = Font(bold=True)
                cl = ws2.cell(row=r2, column=3, value=v if not isinstance(v, str) else v)
                if isinstance(v, (int, float)): cl.number_format = mf
                cl.font = Font(bold=True); r2 += 1
            aw(ws2)

        no = cot.get('no_cotizacion', str(cot_id))
        fname = f"Cotizacion_{no}.xlsx"
        import tempfile
        tmp = os.path.join(tempfile.gettempdir(), fname)
        wb.save(tmp)
        return send_file(tmp, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        tb_str = traceback.format_exc()
        print(f"[EXCEL ERROR] {tb_str}")
        return jsonify({'error': str(e), 'traceback': tb_str}), 500

# â”€â”€â”€ PDF / Print â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PRINT_TEMPLATE = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Cotizacion {{ cot.no_cotizacion }}</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#111;padding:20px;position:relative}
  .watermark{position:fixed;top:50%;left:50%;transform:translate(-50%,-50%) rotate(-10deg);opacity:.06;z-index:0;pointer-events:none}
  .watermark img{width:520px;max-width:78vw;height:auto}
  .content{position:relative;z-index:1}
  .cover{background:linear-gradient(135deg, {{ primary }} 0%, #000 100%);color:#fff;border-radius:12px;padding:16px 18px;margin-bottom:14px;display:flex;align-items:center;justify-content:space-between;gap:14px}
  .cover .left{display:flex;align-items:center;gap:12px}
  .cover img.logo{height:54px;object-fit:contain;filter:drop-shadow(0 2px 4px rgba(0,0,0,.25))}
  .cover .title{font-size:22px;font-weight:800;letter-spacing:.5px}
  .cover .sub{font-size:11px;opacity:.85;margin-top:2px;line-height:1.35}
  .cover .box{background:{{ accent_soft }};color:{{ accent_soft_text }};padding:10px 14px;border-radius:10px;text-align:right;min-width:210px}
  .cover .box .num{font-size:18px;font-weight:900}
  .cover .box .meta{font-size:11px;font-weight:700}
  .header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px;padding-bottom:12px;border-bottom:3px solid {{ accent_soft }}}
  .header-left{display:flex;align-items:center;gap:12px}
  .header-left img.logo{height:48px;object-fit:contain}
  .brand{font-size:22px;font-weight:700;color:{{ primary }}}
  .brand-sub{font-size:11px;color:#666;margin-top:3px}
  .cot-box{background:{{ primary }};color:#fff;padding:8px 16px;border-radius:6px;text-align:right}
  .cot-box .num{font-size:16px;font-weight:700}
  .section-title{background:{{ primary }};color:#fff;padding:5px 10px;font-weight:700;margin:12px 0 6px;font-size:11px}
  .info-grid{display:grid;grid-template-columns:1fr 1fr;gap:3px 12px;margin-bottom:8px}
  .info-row{display:flex;gap:6px;font-size:11px}
  .info-label{font-weight:700;color:{{ primary }};min-width:90px}
  table{width:100%;border-collapse:collapse;font-size:10px;margin-bottom:10px}
  th{background:{{ primary }};color:#fff;padding:5px 4px;text-align:left;font-size:9px}
  td{padding:4px;border-bottom:1px solid #e0e0e0;vertical-align:middle}
  tr:nth-child(even) td{background:#f5f8ff}
  .right{text-align:right}.center{text-align:center}
  .cat-band{background:{{ accent_soft }};color:{{ accent_soft_text }};font-weight:800;padding:6px 10px;border-radius:8px 8px 0 0;margin-top:10px;font-size:11px;letter-spacing:.3px}
  .cat-subtotal td{background:{{ accent_soft_bg }} !important;font-weight:800}
  .img-cell{width:50px;text-align:center}
  .img-cell img{width:45px;height:45px;object-fit:contain;border-radius:3px;border:1px solid #e0e0e0}
  .img-cell .no-img{width:45px;height:45px;display:inline-flex;align-items:center;justify-content:center;background:#f0f0f0;border-radius:3px;color:#bbb;font-size:8px;text-align:center}
  .totals{float:right;width:320px;margin-top:8px}
  .totals table td{border:none;padding:3px 7px}
  .totals .grand{background:{{ accent_soft }};color:{{ accent_soft_text }};font-weight:700;font-size:13px}
  .service-table td{background:#fffdfa}
  .footer{margin-top:30px;padding-top:10px;border-top:1px solid #ccc;font-size:10px;color:#666;display:flex;justify-content:space-between}
  .sign{display:flex;gap:40px;margin-top:20px}
  .sign-line{border-top:1px solid #555;width:200px;padding-top:4px;font-size:10px;text-align:center}
  @media print{body{padding:10px} .no-print{display:none} .watermark{position:fixed;top:50%;left:50%;transform:translate(-50%,-50%) rotate(-10deg);opacity:.06;z-index:0;pointer-events:none}}
</style></head><body>
{% if logo_exists %}<div class="watermark"><img src="{{ watermark_url }}" alt=""></div>{% endif %}
<div class="content">
<div class="no-print" style="margin-bottom:16px">
  <button onclick="window.print()" style="background:{{ accent_soft }};color:{{ accent_soft_text }};border:none;padding:8px 20px;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">Imprimir / Guardar PDF</button>
  <button onclick="window.close()" style="background:#666;color:#fff;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;margin-left:8px">&times; Cerrar</button>
</div>
<div class="cover">
  <div class="left">
    {% if logo_exists %}<img class="logo" src="{{ logo_url }}" alt="Logo">{% endif %}
    <div>
      <div class="title">{{ params.empresa or 'RC DOMOTIC' }}</div>
      <div class="sub">{{ params.ciudad }} &middot; {{ params.direccion }}<br>Tel: {{ params.telefono }} &middot; {{ params.email }}</div>
    </div>
  </div>
  <div class="box">
    <div class="meta">COTIZACION</div>
    <div class="num">{{ cot.no_cotizacion }}</div>
    <div class="meta">Fecha: {{ cot.fecha }}</div>
  </div>
</div>
<!-- header reemplazado por cover -->
<div class="section-title">DATOS DEL CLIENTE</div>
<div class="info-grid">
  <div class="info-row"><span class="info-label">Cliente:</span> <strong>{{ cot.cliente }}</strong></div>
  <div class="info-row"><span class="info-label">Proyecto:</span> {{ cot.proyecto or 'N/A' }}</div>
  <div class="info-row"><span class="info-label">Empresa:</span> {{ cot.empresa or 'N/A' }}</div>
  <div class="info-row"><span class="info-label">Ciudad:</span> {{ cot.ciudad or 'N/A' }}</div>
  <div class="info-row"><span class="info-label">NIT/CC:</span> {{ cot.nit_cc or 'N/A' }}</div>
  <div class="info-row"><span class="info-label">Tipo:</span> {{ cot.tipo_cotizacion }}</div>
  <div class="info-row"><span class="info-label">Telefono:</span> {{ cot.telefono or 'N/A' }}</div>
  <div class="info-row"><span class="info-label">Forma pago:</span> {{ cot.forma_pago }}</div>
</div>
<div class="section-title">DETALLE DE PRODUCTOS</div>
{% for grp in grouped_items %}
<div class="cat-band">{{ grp['label'] }}</div>
<table><thead>
<tr><th>#</th><th class="img-cell">Img</th><th>ID</th><th>Producto</th><th>Descripcion</th><th class="center">Und</th><th class="center">Cant.</th><th class="right">P.Unit.</th><th class="right">IVA</th><th class="right">TOTAL</th></tr>
</thead><tbody>
{% for it in grp['items'] %}
<tr><td class="center">{{ loop.index }}</td>
<td class="img-cell">{% if it.imagen_url %}<img src="{{ it.imagen_url }}" alt="">{% else %}<div class="no-img">Sin<br>img</div>{% endif %}</td>
<td style="font-weight:700;color:{{ primary }}">{{ it.id_producto }}</td>
<td><strong>{{ it.nombre }}</strong></td>
<td style="color:#555;font-size:9px">{{ it.descripcion or '' }}</td>
<td class="center">{{ it.unidad }}</td>
<td class="center">{{ it.cantidad|int }}</td>
<td class="right">$ {{ "{:,.0f}".format(it.precio_final).replace(",", ".") }}</td>
<td class="right">{% if it.iva_monto %}$ {{ "{:,.0f}".format(it.iva_monto).replace(",", ".") }}{% else %}&mdash;{% endif %}</td>
<td class="right"><strong>$ {{ "{:,.0f}".format(it.display_total).replace(",", ".") }}</strong></td></tr>
{% endfor %}
<tr class="cat-subtotal"><td colspan="9" class="right">Subtotal {{ grp['label'] }}</td><td class="right">$ {{ "{:,.0f}".format(grp.subtotal).replace(",", ".") }}</td></tr>
</tbody></table>
{% endfor %}
{% if service_items %}
<div class="cat-band">SERVICIOS COMPLEMENTARIOS</div>
<table class="service-table"><thead>
<tr><th>Concepto</th><th class="right">Total</th></tr>
</thead><tbody>
{% for svc in service_items %}
<tr><td><strong>{{ svc.label }}</strong></td><td class="right"><strong>$ {{ "{:,.0f}".format(svc.total).replace(",", ".") }}</strong></td></tr>
{% endfor %}
<tr class="cat-subtotal"><td class="right">Subtotal servicios complementarios</td><td class="right">$ {{ "{:,.0f}".format(servicios_total).replace(",", ".") }}</td></tr>
</tbody></table>
{% endif %}
<div class="totals"><table>
<tr><td>Subtotal productos:</td><td class="right">$ {{ "{:,.0f}".format(productos_total).replace(",",".") }}</td></tr>
{% if servicios_total > 0 %}<tr><td>Subtotal servicios:</td><td class="right">$ {{ "{:,.0f}".format(servicios_total).replace(",",".") }}</td></tr>{% endif %}
{% if tots.descuento > 0 %}<tr><td>(-) Descuento:</td><td class="right" style="color:#C00000">- $ {{ "{:,.0f}".format(tots.descuento).replace(",",".") }}</td></tr>{% endif %}
<tr class="grand"><td>TOTAL COP:</td><td class="right">$ {{ "{:,.0f}".format(tots.total_final).replace(",",".") }}</td></tr>
<tr><td>
  {% if cot.anticipo_val_manual and cot.anticipo_val_manual|float > 0 %}Anticipo acordado{% else %}Anticipo ({{ (cot.anticipo_pct * 100)|int }}%){% endif %}:
  </td><td class="right">$ {{ "{:,.0f}".format(tots.anticipo_val).replace(",",".") }}</td></tr>
<tr><td>Abonado:</td><td class="right">$ {{ "{:,.0f}".format(tots.abonado_val).replace(",",".") }}</td></tr>
<tr><td>Saldo:</td><td class="right">$ {{ "{:,.0f}".format(tots.saldo_val).replace(",",".") }}</td></tr>
</table></div><div style="clear:both"></div>
{% if cot.notas %}<div style="margin-top:10px;padding:7px 10px;background:#f5f5f5;border-left:3px solid {{ accent_soft }};font-size:11px"><strong>Notas:</strong> {{ cot.notas }}</div>{% endif %}
<div class="sign"><div class="sign-line">Firma Cliente<br>{{ cot.cliente }}</div><div class="sign-line">RC DOMOTIC<br>{{ params.contacto }}</div></div>
<div class="footer"><div>&#10003; Garantia: {{ params.garantia }}<br>&#10003; Vigencia: {{ params.vigencia_dias }} dias &middot; &#10003; Plazo: {{ params.plazo_entrega }}</div>
<div style="text-align:right">Consignaciones: {{ params.banco }}<br>Cta. {{ params.cuenta }} &middot; {{ params.titular }}<br>CC {{ params.nit }}</div></div>
</div></body></html>"""

@app.get('/print/<int:cot_id>')
@login_required
def print_cotizacion(cot_id):
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot: return "Not found", 404
    cot = _repair_mojibake_obj(cot)
    pl = float(cot.get('price_list_desc_pct') or 0)
    items_raw = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                         c.aplica_iva, c.pct_iva, c.inst_default, c.config_default, c.imagen_url, c.categoria
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         WHERE i.cot_id=? ORDER BY i.linea""", (cot_id,))
    items = []
    for it in items_raw:
        items.append(_repair_mojibake_obj({**it, **calcular_item(it, it, pl)}))
    presentation = _repair_mojibake_obj(preparar_presentacion_cotizacion(items))
    grouped_items = presentation['grouped_items']
    service_items = presentation['service_items']
    tots = _repair_mojibake_obj(calcular_cotizacion(cot_id))
    params = _repair_mojibake_obj({r['clave']: r['valor'] for r in query("SELECT * FROM parametros")})
    lp = params.get('logo_path','/static/brand_logo.png')
    wp = params.get('watermark_path','/static/watermark.png')
    le = os.path.isfile(os.path.join(BASE_DIR, lp.lstrip('/')))
    we = os.path.isfile(os.path.join(BASE_DIR, wp.lstrip('/')))
    import time
    ts = int(time.time())
    logo_url = f"{lp}?v={ts}" if le else lp
    watermark_url = f"{wp}?v={ts}" if we else wp
    return render_template_string(PRINT_TEMPLATE, cot=cot, items=presentation['items'], grouped_items=grouped_items,
        service_items=service_items, productos_total=presentation['productos_total'], servicios_total=presentation['servicios_total'], tots=tots, params=params,
        primary=params.get('brand_primary','#0F0F0F'), accent=params.get('brand_accent','#25D366'),
        accent_soft=QUOTE_ACCENT_SOFT, accent_soft_bg=QUOTE_ACCENT_SOFT_BG, accent_soft_text=QUOTE_ACCENT_SOFT_TEXT,
        logo_url=logo_url, watermark_url=watermark_url, logo_exists=le, watermark_exists=we)


# â”€â”€â”€ Vista pÃºblica sin login + aceptaciÃ³n â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PUBLIC_VIEW_TEMPLATE = """<!doctype html><html lang='es'><head>
<meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Cotizacion {{ cot.no_cotizacion }} - RC DOMOTIC</title>
<style>
  *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
  body{font-family:system-ui, -apple-system, 'Segoe UI', Roboto, sans-serif; margin:0; background:#f0f2f5; color:#1a1d21; line-height:1.6; -webkit-font-smoothing:antialiased}
  .wrap{max-width:960px; margin:0 auto; padding:20px 16px;}
  .card{background:#fff; border-radius:18px; box-shadow:0 1px 3px rgba(0,0,0,.06), 0 8px 24px rgba(0,0,0,.08); overflow:hidden; border:1px solid rgba(0,0,0,.04)}
  .top{padding:20px 24px; display:flex; gap:16px; align-items:center; border-bottom:1px solid #f0f0f0; background:#fafafa}
  .logo{width:54px; height:54px; border-radius:14px; background:#0f0f0f; display:flex; align-items:center; justify-content:center; color:#9FB3A6; font-weight:900; font-size:18px; letter-spacing:-1px; flex-shrink:0}
  .grow{flex:1; min-width:0}
  .h1{font-size:20px; font-weight:800; margin:0; letter-spacing:-.3px}
  .sub{font-size:12px; color:#6b7280; margin-top:2px}
  .btn{display:inline-flex; align-items:center; gap:6px; padding:10px 16px; border-radius:10px; background:#0f0f0f; color:#fff; text-decoration:none; font-weight:700; font-size:12px; transition:all .2s; border:none; cursor:pointer}
  .btn:hover{background:#222}
  .btn2{background:#B7C4BB; color:#2F4137}
  .btn2:hover{background:#AAB8AE}
  .content{padding:24px}
  .grid{display:grid; grid-template-columns:1fr 1fr; gap:10px; font-size:13px; color:#1a1d21}
  .row{padding:12px 14px; background:#f8f9fa; border-radius:10px; border:1px solid #f0f0f0}
  .row strong{color:#6b7280; font-size:11px; text-transform:uppercase; letter-spacing:.4px; display:block; margin-bottom:2px}
  table{width:100%; border-collapse:collapse; margin-top:16px; font-size:13px}
  th,td{border-bottom:1px solid #f0f2f5; padding:12px 10px; vertical-align:top}
  th{background:#0f0f0f; color:rgba(255,255,255,.92); font-size:11px; text-align:left; text-transform:uppercase; letter-spacing:.3px; font-weight:700}
  th:first-child{border-radius:10px 0 0 0} th:last-child{border-radius:0 10px 0 0}
  tbody tr:hover{background:rgba(183,196,187,.18)}
  .r{text-align:right}
  .cat-band{margin-top:16px;padding:10px 14px;background:#B7C4BB;color:#2F4137;border-radius:12px 12px 0 0;font-weight:900;font-size:12px;letter-spacing:.3px}
  .cat-subtotal td{background:#EDF2EE;font-weight:800}
  .tot{margin-top:18px; display:flex; justify-content:flex-end}
  .tot table{width:360px}
  .tot td{border-bottom:1px solid #f0f2f5; padding:8px 10px}
  .stamp{display:inline-flex; align-items:center; gap:6px; padding:8px 14px; border-radius:10px; font-weight:800; font-size:13px}
  .ok{background:#EEF4F0; color:#2F4137; border:1.5px solid #B7C4BB}
  .warn{background:#fffbeb; color:#92400e; border:1.5px solid #fcd34d}
  .foot{font-size:12px; color:#6b7280; margin-top:20px; display:flex; justify-content:space-between; gap:14px; flex-wrap:wrap; padding-top:16px; border-top:1px solid #f0f2f5}
  .accept{margin-top:16px; padding:16px; border-radius:14px; background:#f8f9fa; border:1px solid #e5e7eb}
  input,button{font:inherit}
  input[type=text]{width:100%; padding:11px 14px; border-radius:10px; border:1.5px solid #e5e7eb; margin-top:8px; font-size:14px; transition:border-color .2s, box-shadow .2s}
  input[type=text]:focus{border-color:#B7C4BB; box-shadow:0 0 0 3px rgba(183,196,187,.22); outline:none}
  @media(max-width:680px){.grid{grid-template-columns:1fr}.tot table{width:100%}.top{flex-wrap:wrap}}
</style></head><body>
  <div class='wrap'>
    <div class='card'>
      <div class='top'>
        <div class='logo'>RC</div>
        <div class='grow'>
          <p class='h1'>Cotizacion {{ cot.no_cotizacion }}</p>
          <div class='sub'>{{ cot.fecha }} &middot; {{ cot.cliente }} &middot; {{ cot.proyecto or 'N/A' }}</div>
        </div>
        <a class='btn' href='{{ pdf_url }}' target='_blank'>Descargar PDF</a>
      </div>
      <div class='content'>
        <div style='display:flex; gap:10px; align-items:center; flex-wrap:wrap;'>
          {% if cot.accepted %}
            <span class='stamp ok'>ACEPTADA &#10003;</span>
            <span class='sub'>{{ cot.accepted_name or cot.cliente }} &middot; {{ cot.accepted_at }}</span>
          {% else %}
            <span class='stamp warn'>PENDIENTE DE ACEPTACION</span>
          {% endif %}
        </div>
        <div class='grid' style='margin-top:12px'>
          <div class='row'><strong>Cliente</strong><br>{{ cot.cliente }}</div>
          <div class='row'><strong>Telefono</strong><br>{{ cot.telefono or 'N/A' }}</div>
          <div class='row'><strong>Empresa</strong><br>{{ cot.empresa or 'N/A' }}</div>
          <div class='row'><strong>Forma de pago</strong><br>{{ cot.forma_pago }}</div>
        </div>

        {% for grp in grouped_items %}
        <div class='cat-band'>{{ grp['label'] }}</div>
        <table>
          <thead><tr>
            <th style='width:38px'>#</th><th>Producto</th><th class='r' style='width:70px'>Cant</th><th class='r' style='width:110px'>P.Unit</th><th class='r' style='width:110px'>IVA</th><th class='r' style='width:130px'>Total</th>
          </tr></thead>
          <tbody>
            {% for it in grp['items'] %}
              <tr>
                <td>{{ loop.index }}</td>
                <td><strong>{{ it.nombre }}</strong><br><span style='color:#555;font-size:11px'>{{ it.descripcion or '' }}</span></td>
                <td class='r'>{{ it.cantidad|int }}</td>
                <td class='r'>$ {{ "{:,.0f}".format(it.precio_final).replace(",", ".") }}</td>
                <td class='r'>{% if it.iva_monto %}$ {{ "{:,.0f}".format(it.iva_monto).replace(",", ".") }}{% else %}&mdash;{% endif %}</td>
                <td class='r'><strong>$ {{ "{:,.0f}".format(it.display_total).replace(",", ".") }}</strong></td>
              </tr>
            {% endfor %}
            <tr class='cat-subtotal'>
              <td colspan='5' class='r'>Subtotal {{ grp['label'] }}</td>
              <td class='r'>$ {{ "{:,.0f}".format(grp.subtotal).replace(",", ".") }}</td>
            </tr>
          </tbody>
        </table>
        {% endfor %}
        {% if service_items %}
        <div class='cat-band'>SERVICIOS COMPLEMENTARIOS</div>
        <table>
          <thead><tr><th>Concepto</th><th class='r' style='width:160px'>Total</th></tr></thead>
          <tbody>
            {% for svc in service_items %}
              <tr>
                <td><strong>{{ svc.label }}</strong></td>
                <td class='r'><strong>$ {{ "{:,.0f}".format(svc.total).replace(",", ".") }}</strong></td>
              </tr>
            {% endfor %}
            <tr class='cat-subtotal'>
              <td class='r'>Subtotal servicios complementarios</td>
              <td class='r'>$ {{ "{:,.0f}".format(servicios_total).replace(",", ".") }}</td>
            </tr>
          </tbody>
        </table>
        {% endif %}
        <div class='tot'>
          <table>
            <tr><td>Subtotal productos</td><td class='r'>$ {{ "{:,.0f}".format(productos_total).replace(",",".") }}</td></tr>
            {% if servicios_total > 0 %}<tr><td>Subtotal servicios</td><td class='r'>$ {{ "{:,.0f}".format(servicios_total).replace(",",".") }}</td></tr>{% endif %}
            {% if tots.descuento > 0 %}<tr><td>(-) Descuento</td><td class='r' style='color:#C00000'>- $ {{ "{:,.0f}".format(tots.descuento).replace(",",".") }}</td></tr>{% endif %}
            <tr><td style='font-weight:900'>TOTAL</td><td class='r' style='font-weight:900'>$ {{ "{:,.0f}".format(tots.total_final).replace(",",".") }}</td></tr>
            <tr><td>
              {% if cot.anticipo_val_manual and cot.anticipo_val_manual|float > 0 %}Anticipo acordado{% else %}Anticipo ({{ (cot.anticipo_pct*100)|int }}%){% endif %}
            </td><td class='r'>$ {{ "{:,.0f}".format(tots.anticipo_val).replace(",",".") }}</td></tr>
            <tr><td>Abonado</td><td class='r'>$ {{ "{:,.0f}".format(tots.abonado_val).replace(",",".") }}</td></tr>
            <tr><td>Saldo</td><td class='r'>$ {{ "{:,.0f}".format(tots.saldo_val).replace(",",".") }}</td></tr>
          </table>
        </div>

        {% if cot.notas %}<div class='accept' style='background:#fff; border:1px solid #eee'><strong>Notas:</strong> {{ cot.notas }}</div>{% endif %}

        {% if not cot.accepted %}
        <div class='accept'>
          <form method='post' action='{{ accept_url }}'>
            <label style='font-weight:800'>Nombre de quien acepta (opcional)</label>
            <input type='text' name='nombre' placeholder='Ej: Juan Perez'>
            <div style='display:flex; gap:10px; align-items:center; margin-top:10px; flex-wrap:wrap'>
              <label style='display:flex; gap:8px; align-items:center'>
                <input type='checkbox' name='ok' value='1' required>
                Acepto esta cotizacion
              </label>
              <button type='submit'>Aceptar</button>
            </div>
            <div class='sub' style='margin-top:8px'>Al aceptar, se registra fecha y direccion IP.</div>
          </form>
        </div>
        {% endif %}

        <div class='foot'>
          <div>Garantia: {{ params.garantia }} &middot; Vigencia: {{ params.vigencia_dias }} dias &middot; Plazo: {{ params.plazo_entrega }}</div>
          <div><strong>RC DOMOTIC</strong> &middot; {{ params.contacto }} &middot; {{ params.email }}</div>
        </div>
      </div>
    </div>
  </div>
</body></html>"""

def _get_cot_by_token(token:str):
    cot = query("SELECT * FROM cotizaciones WHERE public_token=?", (token,), one=True)
    if not cot:
        return None
    if int(cot.get('public_revoked') or 0) == 1:
        return None
    exp = (cot.get('public_expires_at') or '').strip()
    if exp:
        try:
            exp_dt = datetime.datetime.strptime(exp, '%Y-%m-%d %H:%M:%S')
            if datetime.datetime.utcnow() > exp_dt:
                return None
        except Exception:
            return None
    return cot

@app.get('/q/<token>')
@limit('60 per minute')
def public_view(token):
    cot = _get_cot_by_token(token)
    if not cot: return "Not found", 404
    cot = _repair_mojibake_obj(cot)
    pl = float(cot.get('price_list_desc_pct') or 0)
    items_raw = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                         c.aplica_iva, c.pct_iva, c.inst_default, c.config_default, c.categoria
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         WHERE i.cot_id=? ORDER BY i.linea""", (cot['id'],))
    items = [_repair_mojibake_obj({**it, **calcular_item(it, it, pl)}) for it in items_raw]
    presentation = _repair_mojibake_obj(preparar_presentacion_cotizacion(items))
    grouped_items = presentation['grouped_items']
    service_items = presentation['service_items']
    tots = _repair_mojibake_obj(calcular_cotizacion(cot['id']))
    params = _repair_mojibake_obj({r['clave']: r['valor'] for r in query("SELECT * FROM parametros")})
    base = request.host_url.rstrip('/')
    return render_template_string(PUBLIC_VIEW_TEMPLATE, cot=cot, items=presentation['items'], grouped_items=grouped_items,
        service_items=service_items, productos_total=presentation['productos_total'], servicios_total=presentation['servicios_total'], tots=tots, params=params,
        pdf_url=f"{base}/q/{token}/pdf", accept_url=f"{base}/q/{token}/accept")

@app.post('/q/<token>/accept')
@limit('20 per minute')
def public_accept(token):
    cot = _get_cot_by_token(token)
    if not cot: return "Not found", 404
    if cot.get('accepted'):
        return ('', 204)
    if (request.form.get('ok') or '') != '1':
        return "Bad request", 400
    nombre = (request.form.get('nombre') or '').strip()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    ip = request.headers.get('X-Forwarded-For', request.remote_addr) or ''
    execute("UPDATE cotizaciones SET accepted=1, accepted_at=?, accepted_name=?, accepted_ip=? WHERE id=?",
            (now, nombre, ip, cot['id']))
    return ("<script>location.href='/q/%s';</script>" % token)

@app.get('/q/<token>/pdf')
@limit('30 per minute')
def public_pdf(token):
    cot = _get_cot_by_token(token)
    if not cot: return "Not found", 404
    cot = _repair_mojibake_obj(cot)
    pl = float(cot.get('price_list_desc_pct') or 0)
    items_raw = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                         c.aplica_iva, c.pct_iva, c.inst_default, c.config_default, c.categoria
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         WHERE i.cot_id=? ORDER BY i.linea""", (cot['id'],))
    items = [_repair_mojibake_obj({**it, **calcular_item(it, it, pl)}) for it in items_raw]
    presentation = _repair_mojibake_obj(preparar_presentacion_cotizacion(items))
    grouped_items = presentation['grouped_items']
    service_items = presentation['service_items']
    tots = _repair_mojibake_obj(calcular_cotizacion(cot['id']))
    params = _repair_mojibake_obj({r['clave']: r['valor'] for r in query("SELECT * FROM parametros")})
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    w, h = letter
    y = h - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, f"Cotizacion {cot.get('no_cotizacion','')}")
    y -= 18
    c.setFont("Helvetica", 10)
    c.drawString(40, y, f"Cliente: {cot.get('cliente','')}  |  Proyecto: {cot.get('proyecto','') or 'N/A'}")
    y -= 14
    c.drawString(40, y, f"Fecha: {cot.get('fecha','')}  |  Contacto: {params.get('contacto','')}")
    y -= 22
    c.setFont("Helvetica-Bold", 9)
    row_no = 1
    for grp in grouped_items:
        if y < 120:
            c.showPage(); y = h - 40
        c.setFillColorRGB(0.72, 0.77, 0.73)
        c.rect(40, y - 10, 515, 16, fill=1, stroke=0)
        c.setFillColorRGB(0.18, 0.25, 0.22)
        c.drawString(46, y - 2, grp['label'])
        y -= 18
        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y, "#")
        c.drawString(55, y, "Producto")
        c.drawRightString(400, y, "Cant")
        c.drawRightString(470, y, "P.Unit")
        c.drawRightString(545, y, "Total")
        y -= 8
        c.line(40, y, 555, y)
        y -= 14
        c.setFont("Helvetica", 9)
        for it in grp['items']:
            if y < 80:
                c.showPage(); y = h - 40
            c.drawString(40, y, str(row_no))
            nombre = (it.get('nombre') or '')[:55]
            c.drawString(55, y, nombre)
            c.drawRightString(400, y, str(int(it.get('cantidad') or 0)))
            c.drawRightString(470, y, f"{int(it.get('precio_final') or 0):,}".replace(',', '.'))
            c.drawRightString(545, y, f"{int(it.get('display_total') or 0):,}".replace(',', '.'))
            y -= 13
            row_no += 1
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(470, y, f"Subtotal {grp['label']}")
        c.drawRightString(545, y, f"{int(grp.get('subtotal') or 0):,}".replace(',', '.'))
        y -= 18
        c.setFont("Helvetica", 9)
    if service_items:
        if y < 120:
            c.showPage(); y = h - 40
        c.setFillColorRGB(0.72, 0.77, 0.73)
        c.rect(40, y - 10, 515, 16, fill=1, stroke=0)
        c.setFillColorRGB(0.18, 0.25, 0.22)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(46, y - 2, "SERVICIOS COMPLEMENTARIOS")
        y -= 20
        c.setFont("Helvetica", 9)
        for svc in service_items:
            if y < 80:
                c.showPage(); y = h - 40
            c.drawString(55, y, svc['label'])
            c.drawRightString(545, y, f"{int(svc.get('total') or 0):,}".replace(',', '.'))
            y -= 13
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(470, y, "Subtotal servicios")
        c.drawRightString(545, y, f"{int(presentation.get('servicios_total') or 0):,}".replace(',', '.'))
        y -= 18
        c.setFont("Helvetica", 9)
    y -= 6
    c.line(330, y, 555, y)
    y -= 16
    c.drawRightString(470, y, "Subtotal productos")
    c.drawRightString(545, y, f"{int(presentation.get('productos_total') or 0):,}".replace(',', '.'))
    y -= 14
    if presentation.get('servicios_total'):
        c.drawRightString(470, y, "Subtotal servicios")
        c.drawRightString(545, y, f"{int(presentation.get('servicios_total') or 0):,}".replace(',', '.'))
        y -= 14
    if tots.get('descuento', 0):
        c.drawRightString(470, y, "Descuento")
        c.drawRightString(545, y, f"- {int(tots.get('descuento',0)):,}".replace(',', '.'))
        y -= 14
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(470, y, "TOTAL")
    c.drawRightString(545, y, f"{int(tots.get('total_final',0)):,}".replace(',', '.'))
    y -= 18
    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Garantia: {params.get('garantia','')} - Vigencia: {params.get('vigencia_dias','')} dias - Plazo: {params.get('plazo_entrega','')}")
    c.showPage(); c.save()
    buf.seek(0)
    filename = f"{cot.get('no_cotizacion','cotizacion')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)

# Serve SPA
def _get_cot_by_no(no:str):
    return query("SELECT * FROM cotizaciones WHERE no_cotizacion=?", (no,), one=True)

@app.get('/cotizacion/<path:no>')
def public_view_no(no):
    cot = _get_cot_by_no(no)
    if not cot:
        return "Not found", 404
    cot = _repair_mojibake_obj(cot)
    token = ensure_public_token(cot['id'])
    pl = float(cot.get('price_list_desc_pct') or 0)
    items_raw = query("""SELECT i.*, c.nombre, c.descripcion, c.unidad, c.precio,
                         c.aplica_iva, c.pct_iva, c.inst_default, c.config_default, c.categoria
                         FROM items i JOIN catalogo c ON i.id_producto=c.id_producto
                         WHERE i.cot_id=? ORDER BY i.linea""", (cot['id'],))
    items = [_repair_mojibake_obj({**it, **calcular_item(it, it, pl)}) for it in items_raw]
    presentation = _repair_mojibake_obj(preparar_presentacion_cotizacion(items))
    grouped_items = presentation['grouped_items']
    service_items = presentation['service_items']
    tots = _repair_mojibake_obj(calcular_cotizacion(cot['id']))
    params = _repair_mojibake_obj({r['clave']: r['valor'] for r in query("SELECT * FROM parametros")})
    base = request.host_url.rstrip('/')
    import urllib.parse
    no_q = urllib.parse.quote(no, safe='')
    return render_template_string(PUBLIC_VIEW_TEMPLATE, cot=cot, items=presentation['items'], grouped_items=grouped_items,
        service_items=service_items, productos_total=presentation['productos_total'], servicios_total=presentation['servicios_total'], tots=tots, params=params,
        pdf_url=f"{base}/cotizacion/{no_q}/pdf", accept_url=f"{base}/cotizacion/{no_q}/accept")

@app.post('/cotizacion/<path:no>/accept')
def public_accept_no(no):
    cot = _get_cot_by_no(no)
    if not cot:
        return "Not found", 404
    if cot.get('accepted'):
        return ('', 204)
    if (request.form.get('ok') or '') != '1':
        return "Bad request", 400
    nombre = (request.form.get('nombre') or '').strip()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    ip = request.headers.get('X-Forwarded-For', request.remote_addr) or ''
    execute("UPDATE cotizaciones SET accepted=1, accepted_at=?, accepted_name=?, accepted_ip=? WHERE id=?",
            (now, nombre, ip, cot['id']))
    import urllib.parse
    no_q = urllib.parse.quote(no, safe='')
    return ("<script>location.href='/cotizacion/%s';</script>" % no_q)

@app.get('/cotizacion/<path:no>/pdf')
def public_pdf_no(no):
    cot = _get_cot_by_no(no)
    if not cot:
        return "Not found", 404
    token = ensure_public_token(cot['id'])
    return public_pdf(token)


@app.get('/')
def serve_index():
    """Home - Single Page App (cotizador interno)."""
    return send_file(os.path.join(BASE_DIR, 'static', 'index.html'))


@app.get('/catalog')
@app.get('/catalogo')
def serve_catalog():
    """Catálogo público standalone — para catalogo.rcdomotic.com."""
    return send_file(os.path.join(BASE_DIR, 'catalog.html'))


@app.get('/print/inventario')
@role_required('admin')
def print_inventario():
    bodega = (request.args.get('bodega') or 'PRINCIPAL').upper()
    data = api_inventario().json
    items = data.get('items', [])
    # Simple imprimible con miniaturas
    def fm(n): 
        try: return f"{int(float(n)):,}".replace(',', '.')
        except: return str(n)
    html = """<!doctype html><html><head><meta charset='utf-8'>
    <meta name='viewport' content='width=device-width,initial-scale=1'>
    <title>Inventario</title>
    <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'Segoe UI',-apple-system,system-ui,sans-serif;font-size:12px;margin:20px;color:#1a1d21;-webkit-font-smoothing:antialiased}
    h1{margin:0 0 4px 0;font-size:20px;font-weight:800;letter-spacing:-.3px}
    .sub{color:#6b7280;margin-bottom:16px;font-size:12px}
    table{width:100%;border-collapse:collapse}
    th,td{border:1px solid #e5e7eb;padding:8px 6px;vertical-align:middle}
    th{background:#0f0f0f;color:#fff;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.4px;font-weight:700}
    .img{width:44px;height:44px;border-radius:8px;overflow:hidden;background:#f3f4f6;display:flex;align-items:center;justify-content:center}
    .img img{width:100%;height:100%;object-fit:cover}
    .muted{color:#6b7280}
    .low{background:#fffbeb}
    tbody tr:hover{background:rgba(37,211,102,.04)}
    @media print{body{margin:10px}th{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
    </style></head><body>
    <h1>Inventario â€” """+bodega+"""</h1>
    <div class='sub'>Generado: """+datetime.datetime.now().strftime('%Y-%m-%d %H:%M')+"""</div>
    <table><thead><tr>
      <th>Foto</th><th>ID</th><th>Producto</th><th>CategorÃ­a</th>
      <th>Stock</th><th>Min</th><th>Costo</th><th>Precio</th>
    </tr></thead><tbody>
    """
    for it in items:
        low = 'low' if it.get('alerta_reponer') else ''
        img = it.get('imagen_url') or ''
        if img and img.startswith('/uploads/'):
            img = img  # served by app
        row = f"<tr class='{low}'>"
        if img:
            row += f"<td><div class='img'><img src='{img}'></div></td>"
        else:
            row += "<td><div class='img'><span class='muted'>â€”</span></div></td>"
        row += f"<td>{it['id_producto']}</td><td>{it['nombre']}</td><td>{it['categoria']}</td>"
        row += f"<td>{fm(it.get('stock_qty',0))}</td><td>{fm(it.get('stock_min',0))}</td>"
        row += f"<td>$ {fm(it.get('costo_unitario',0))}</td><td>$ {fm(it.get('precio',0))}</td></tr>"
        html += row
    html += """</tbody></table></body></html>"""
    return html


# â”€â”€â”€ BOT API (Telegram Cotizador) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _require_bot_key():
    if not BOT_KEYS:
        _audit_event('bot_key_failed', outcome='rejected', actor={'type': 'bot'}, reason='bot_key_not_configured')
        return jsonify({'ok': False, 'error': 'BOT_KEY no configurado en el servidor.'}), 503
    if not _request_has_valid_bot_key():
        _audit_event('bot_key_failed', outcome='rejected', actor={'type': 'bot'}, reason='invalid_bot_key')
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401
    return None

def _normalize_str(s):
    s = str(s or '').lower().strip()
    s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
    for a, b in [('Ã¡','a'),('Ã©','e'),('Ã­','i'),('Ã³','o'),('Ãº','u'),('Ã±','n'),('Ã¼','u')]:
        s = s.replace(a, b)
    s = re.sub(r'[^a-z0-9\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    replacements = [
        ('orvivo', 'orvibo'),
        ('orbivo', 'orvibo'),
        ('orvbio', 'orvibo'),
        ('decopro', 'deco pro'),
        ('decos', 'deco'),
        ('interruptores', 'interruptor'),
        ('grises', 'gris'),
        ('pantallas', 'pantalla'),
        ('grandes', 'grande'),
        ('midpad', 'mixpad'),
        ('mipad', 'mixpad'),
        ('exteriores', 'exterior'),
        ('interiores', 'interior'),
        ('parlantes', 'parlante'),
        ('bocinas', 'bocina'),
        ('techos', 'techo'),
        ('sensores', 'sensor'),
        ('camaras', 'camara'),
        ('megapixeles', 'mp'),
        ('megapixeleses', 'mp'),
        ('megapixeles', 'mp'),
        ('megapixel', 'mp'),
        ('tomacorrientes', 'tomacorriente'),
        ('tomas', 'toma'),
        ('gfcs', 'gfci'),
        ('camarabala', 'camara bala'),
        ('lumion', 'lumios'),
        ('queen', 'wiim'),
        ('quinn', 'wiim'),
        ('defi', 'defy'),
        ('4k1', '4k'),
        ('digitales', 'digital'),
        ('cerraduras', 'cerradura'),
    ]
    for old, new in replacements:
        s = re.sub(rf'\b{old}\b', new, s)
    return re.sub(r'\s+', ' ', s).strip()

def _query_terms(s):
    return set(_normalize_str(s).split())

def _catalog_rows():
    return query("SELECT id_producto, categoria, nombre FROM catalogo WHERE activo=1")

def _find_product(all_p, id_producto):
    for p in all_p:
        if p['id_producto'] == id_producto:
            return p
    return None

def _preferred_categories(qn):
    words = set(qn.split())
    cats = set()
    if words & {'camara', 'cctv', 'nvr', 'solar', 'floodlight', 'lumios', 'lumi', 'bala'}:
        cats.add('CCTV')
    if words & {'deco', 'mesh', 'wifi', 'red', 'router'}:
        cats.add('REDES')
    if words & {'orvibo', 'mixpad', 'relay', 'interruptor', 'sensor', 'hub', 'toma', 'gfci', 'usb', 'tomacorriente'}:
        cats.add('DOMOTICA')
    if words & {'proyector', 'telon', 'parlante', 'bocina', 'techo', 'amplificador', 'wiim'}:
        cats.add('AUDIOVISUAL')
    if words & {'cerradura', 'lock'}:
        cats.add('CERRADURAS')
    return cats

def _alias_match(qn, all_p):
    words = set(qn.split())
    ordered_rules = [
        ({'deco', 'pro'}, 'RED-001'),
        ({'b3000'}, 'RED-001'),
        ({'be3000'}, 'RED-001'),
        ({'deco', 'm5'}, 'RED-002'),
        ({'deco'}, 'RED-002'),
        ({'cerradura', 'v5'}, 'CEK-007'),
        ({'v5'}, 'CEK-007'),
        ({'cerradura', 'digital', 'huella'}, 'CEK-001'),
        ({'interruptor', 'orvibo', 'defy', 'gris'}, 'DOM-024'),
        ({'interruptor', 'orvibo', 'defy'}, 'DOM-024'),
        ({'interruptor', 'orvibo', 'gris'}, 'DOM-001'),
        ({'interruptor', 'orvibo', 'negro'}, 'DOM-002'),
        ({'interruptor', 'orvibo', 'blanco'}, 'DOM-023'),
        ({'mixpad', 'mini'}, 'DOM-003'),
        ({'pantalla', 'mini', 'orvibo'}, 'DOM-003'),
        ({'mixpad', 'grande'}, 'DOM-004'),
        ({'pantalla', 'grande', 'orvibo'}, 'DOM-004'),
        ({'relay', 'micro'}, 'DOM-009'),
        ({'parlante', 'klipsch'}, 'AV-003'),
        ({'parlante', 'empotrado'}, 'AV-003'),
        ({'parlante', 'sala', 'junta'}, 'AV-004'),
        ({'parlante', 'sala', 'central'}, 'AV-005'),
        ({'parlante', 'exterior'}, 'AV-006'),
        ({'bocina', 'exterior'}, 'AV-006'),
        ({'bocina', 'techo'}, 'AV-007'),
        ({'parlante', 'techo'}, 'AV-007'),
        ({'parlante', 'bocina'}, 'AV-006'),
        ({'parlante'}, 'AV-006'),
        ({'bocina'}, 'AV-006'),
        ({'amplificador', 'wiim'}, 'AV-001'),
        ({'amplificador', 'ultra'}, 'AV-002'),
        ({'wiim', 'ultra'}, 'AV-002'),
        ({'sensor', 'presencia'}, 'DOM-015'),
        ({'camara', '360'}, 'CAM-001'),
        ({'camara', 'bala', '4k'}, 'CAM-004'),
        ({'lumios'}, 'CAM-005'),
        ({'camara', 'exterior', '16', 'mp'}, 'CAM-003'),
        ({'camara', 'exterior', '16mp'}, 'CAM-003'),
        ({'camara', 'exterior', '180', 'solar'}, 'CAM-015'),
        ({'solar', 'panoramica'}, 'CAM-015'),
        ({'toma', 'usb'}, 'DOM-005'),
        ({'tomacorriente', 'usb'}, 'DOM-005'),
    ]
    for required, pid in ordered_rules:
        if required.issubset(words):
            prod = _find_product(all_p, pid)
            if prod:
                return prod
    return None

def _match_catalog(codigo_query):
    raw_query = str(codigo_query or '').strip()
    qn = _normalize_str(raw_query)
    if not qn and not raw_query:
        return {'found': False, 'ambiguous': []}
    # 1. exact code match
    prod = query("SELECT id_producto, categoria, nombre FROM catalogo WHERE activo=1 AND UPPER(id_producto)=?",
                 (raw_query.upper(),), one=True)
    if prod:
        return {'found': True, 'id_producto': prod['id_producto'], 'nombre': prod['nombre']}
    all_p = _catalog_rows()
    query_code_norm = re.sub(r'[^a-z0-9]+', '', (raw_query or qn).lower())
    if query_code_norm:
        for p in all_p:
            prod_code_norm = re.sub(r'[^a-z0-9]+', '', str(p['id_producto']).lower())
            if prod_code_norm == query_code_norm:
                return {'found': True, 'id_producto': p['id_producto'], 'nombre': p['nombre']}
    aliased = _alias_match(qn, all_p)
    if aliased:
        return {'found': True, 'id_producto': aliased['id_producto'], 'nombre': aliased['nombre']}
    exact, ranked = [], []
    q_words = _query_terms(qn)
    pref_cats = _preferred_categories(qn)
    for p in all_p:
        pn = _normalize_str(p['nombre'])
        if pn == qn:
            return {'found': True, 'id_producto': p['id_producto'], 'nombre': p['nombre']}
        p_words = _query_terms(p['nombre'])
        if q_words and q_words.issubset(p_words):
            exact.append(p)
            continue
        overlap = len(q_words & p_words)
        if not overlap:
            continue
        score = overlap * 15
        if qn in pn or pn in qn:
            score += 35
        if pref_cats and p['categoria'] in pref_cats:
            score += 20
        if overlap >= max(1, len(q_words) - 1):
            score += 10
        ranked.append((score, p))
    candidates = exact
    if not candidates and ranked:
        ranked.sort(key=lambda x: (-x[0], x[1]['id_producto']))
        best_score = ranked[0][0]
        if best_score >= 20:
            candidates = [p for score, p in ranked if score >= max(20, best_score - 10)][:5]
    if len(candidates) == 1:
        return {'found': True, 'id_producto': candidates[0]['id_producto'], 'nombre': candidates[0]['nombre']}
    elif candidates:
        return {'found': False, 'ambiguous': [{'id_producto': p['id_producto'], 'nombre': p['nombre']} for p in candidates[:5]]}
    return {'found': False, 'ambiguous': []}

def _bot_now_iso():
    return datetime.datetime.now().isoformat(timespec='seconds')

def _bot_ctx_get(chat_id):
    chat_id = str(chat_id or '').strip()
    if not chat_id:
        return None
    return query("SELECT * FROM bot_chat_context WHERE chat_id=?", (chat_id,), one=True)

def _bot_ctx_set(chat_id, cot_id, mode='editar'):
    chat_id = str(chat_id or '').strip()
    if not chat_id or not cot_id:
        return
    ts = _bot_now_iso()
    execute(
        """INSERT INTO bot_chat_context (chat_id,cot_id,mode,created_at,updated_at)
           VALUES (?,?,?,?,?)
           ON CONFLICT(chat_id) DO UPDATE SET
             cot_id=excluded.cot_id,
             mode=excluded.mode,
             updated_at=excluded.updated_at""",
        (chat_id, cot_id, mode, ts, ts)
    )

def _bot_ctx_clear(chat_id):
    chat_id = str(chat_id or '').strip()
    if chat_id:
        execute("DELETE FROM bot_chat_context WHERE chat_id=?", (chat_id,))

def _bot_ctx_clear_by_cot(cot_id):
    if cot_id:
        execute("DELETE FROM bot_chat_context WHERE cot_id=?", (cot_id,))

def _bot_qty_label(qty):
    try:
        qty_num = float(qty)
        if qty_num.is_integer():
            return str(int(qty_num))
        return str(qty_num)
    except Exception:
        return str(qty)

def _bot_add_or_merge_item(cot_id, codigo, cant, notas_item=''):
    codigo = str(codigo or '').strip().upper()
    cant = max(1, int(cant or 1))
    notas_item = str(notas_item or '').strip()
    existing = query("SELECT * FROM items WHERE cot_id=? AND id_producto=?", (cot_id, codigo), one=True)
    if existing:
        if notas_item and not str(existing.get('notas_item') or '').strip():
            execute("UPDATE items SET cantidad=cantidad+?, notas_item=? WHERE id=?", (cant, notas_item, existing['id']))
        else:
            execute("UPDATE items SET cantidad=cantidad+? WHERE id=?", (cant, existing['id']))
        return
    ml = query("SELECT MAX(linea) as ml FROM items WHERE cot_id=?", (cot_id,), one=True)
    execute(
        "INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item) VALUES (?,?,?,?,0,0,0,?)",
        (cot_id, (ml['ml'] or 0) + 1, codigo, cant, notas_item)
    )

def _bot_build_quote_summary(cot_id, items_ambiguos=None, items_no_encontrados=None, intro=''):
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return ''
    rows = query(
        """SELECT i.id_producto, i.cantidad, c.nombre
           FROM items i
           LEFT JOIN catalogo c ON c.id_producto = i.id_producto
           WHERE i.cot_id=?
           ORDER BY i.linea, i.id""",
        (cot_id,)
    )
    lines = [f"Borrador {cot['no_cotizacion']} - {cot.get('cliente') or 'Sin especificar'}"]
    if intro:
        lines.append(intro)
    if cot.get('ciudad'):
        lines.append(f"Ciudad: {cot['ciudad']}")
    if cot.get('proyecto'):
        lines.append(f"Proyecto: {cot['proyecto']}")
    lines.append("")
    for row in rows:
        nombre = row.get('nombre') or row['id_producto']
        lines.append(f"- {_bot_qty_label(row.get('cantidad'))}x {nombre} ({row['id_producto']})")
    items_ambiguos = items_ambiguos or []
    items_no_encontrados = items_no_encontrados or []
    if items_ambiguos:
        lines.append("")
        lines.append("Ambiguos (elige abajo):")
        for item in items_ambiguos:
            lines.append(f"\"{item['query']}\" - varias opciones")
    if items_no_encontrados:
        lines.append("")
        lines.append("No encontre: " + ', '.join(x['query'] for x in items_no_encontrados))
    return '\n'.join(lines)

BOT_QTY_WORDS = {
    'un': 1, 'una': 1, 'uno': 1,
    'dos': 2, 'tres': 3, 'cuatro': 4, 'cinco': 5,
    'seis': 6, 'siete': 7, 'ocho': 8, 'nueve': 9, 'diez': 10,
    'once': 11, 'doce': 12, 'trece': 13, 'catorce': 14, 'quince': 15,
    'dieciseis': 16, 'diecisiete': 17, 'dieciocho': 18, 'diecinueve': 19, 'veinte': 20,
}

def _bot_extract_qty_and_phrase(raw_text):
    text = _normalize_str(raw_text)
    text = re.sub(r'\b(?:de|del|la|el|los|las|uds|unidad|unidades)\b', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    if not text:
        return 1, ''
    parts = text.split(' ', 1)
    head = parts[0]
    rest = parts[1] if len(parts) > 1 else ''
    if head.isdigit():
        return max(1, int(head)), rest.strip()
    if head in BOT_QTY_WORDS:
        return BOT_QTY_WORDS[head], rest.strip()
    return 1, text

def _bot_quote_items(cot_id):
    return query(
        """SELECT i.id, i.id_producto, i.cantidad, c.nombre, c.categoria
           FROM items i
           JOIN catalogo c ON c.id_producto = i.id_producto
           WHERE i.cot_id=?
           ORDER BY i.linea, i.id""",
        (cot_id,)
    )

def _bot_match_existing_item(cot_id, query_text):
    qn = _normalize_str(query_text)
    if not qn:
        return {'found': False, 'ambiguous': []}
    rows = _bot_quote_items(cot_id)
    exact = []
    ranked = []
    q_words = _query_terms(qn)
    for row in rows:
        name_norm = _normalize_str(row['nombre'])
        code_norm = _normalize_str(row['id_producto'])
        combined = f"{name_norm} {code_norm}".strip()
        if qn == name_norm or qn == code_norm or qn == combined:
            return {'found': True, 'id_producto': row['id_producto'], 'nombre': row['nombre']}
        words = _query_terms(combined)
        if q_words and q_words.issubset(words):
            exact.append(row)
            continue
        overlap = len(q_words & words)
        if not overlap:
            continue
        score = overlap * 20
        if qn in combined or combined in qn:
            score += 30
        if overlap >= max(1, len(q_words) - 1):
            score += 10
        ranked.append((score, row))
    candidates = exact
    if not candidates and ranked:
        ranked.sort(key=lambda x: (-x[0], x[1]['id_producto']))
        best = ranked[0][0]
        if best >= 20:
            candidates = [row for score, row in ranked if score >= max(20, best - 10)][:5]
    if len(candidates) == 1:
        row = candidates[0]
        return {'found': True, 'id_producto': row['id_producto'], 'nombre': row['nombre']}
    if candidates:
        return {'found': False, 'ambiguous': [{'id_producto': row['id_producto'], 'nombre': row['nombre']} for row in candidates]}
    return {'found': False, 'ambiguous': []}

def _bot_remove_or_reduce_item(cot_id, codigo, cant):
    codigo = str(codigo or '').strip().upper()
    remaining = max(1, int(cant or 1))
    rows = query("SELECT id, cantidad FROM items WHERE cot_id=? AND id_producto=? ORDER BY linea, id", (cot_id, codigo))
    changed = 0
    for row in rows:
        if remaining <= 0:
            break
        qty = int(row.get('cantidad') or 0)
        if qty <= 0:
            continue
        if remaining >= qty:
            execute("DELETE FROM items WHERE id=?", (row['id'],))
            remaining -= qty
            changed += qty
        else:
            execute("UPDATE items SET cantidad=? WHERE id=?", (qty - remaining, row['id']))
            changed += remaining
            remaining = 0
    return changed

def _bot_parse_edit_ops(cot_id, raw_text):
    text = _normalize_str(raw_text)
    if not text:
        return {'handled': False}
    ops = {
        'handled': False,
        'remove_ok': [],
        'remove_ambiguos': [],
        'remove_no_encontrados': [],
        'add_ok': [],
        'add_ambiguos': [],
        'add_no_encontrados': [],
    }

    def resolve_remove(chunk):
        qty, phrase = _bot_extract_qty_and_phrase(chunk)
        match = _bot_match_existing_item(cot_id, phrase)
        if match['found']:
            ops['remove_ok'].append({'codigo': match['id_producto'], 'nombre': match['nombre'], 'cantidad': qty, 'query': phrase})
        elif match['ambiguous']:
            ops['remove_ambiguos'].append({'query': phrase, 'opciones': match['ambiguous']})
        else:
            ops['remove_no_encontrados'].append({'query': phrase})

    def resolve_add(chunk):
        qty, phrase = _bot_extract_qty_and_phrase(chunk)
        match = _match_catalog(phrase)
        if match['found']:
            ops['add_ok'].append({'codigo': match['id_producto'], 'nombre': match['nombre'], 'cantidad': qty})
        elif match['ambiguous']:
            ops['add_ambiguos'].append({'query': phrase, 'opciones': match['ambiguous']})
        else:
            ops['add_no_encontrados'].append({'query': phrase})

    replacement = re.search(r'(?:cambia|cambiar|reemplaza|reemplazar|sustituye|sustituir)\s+(.+?)\s+por\s+(.+)', text)
    if replacement:
        ops['handled'] = True
        resolve_remove(replacement.group(1))
        resolve_add(replacement.group(2))
        return ops

    remove_add = re.search(r'(?:quita|quitar|quite|borra|borrar|elimina|eliminar|saca|sacar)\s+(.+?)\s+(?:y\s+)?(?:pon|poner|agrega|agregar|anade|anadir|suma|sumar|mete|meter|deja|dejar|coloca|colocar)\s+(.+)', text)
    if remove_add:
        ops['handled'] = True
        resolve_remove(remove_add.group(1))
        resolve_add(remove_add.group(2))
        return ops

    remove_only = re.search(r'(?:quita|quitar|quite|borra|borrar|elimina|eliminar|saca|sacar)\s+(.+)', text)
    if remove_only:
        ops['handled'] = True
        resolve_remove(remove_only.group(1))
        return ops

    add_only = re.search(r'(?:pon|poner|agrega|agregar|anade|anadir|suma|sumar|mete|meter|deja|dejar|coloca|colocar)\s+(.+)', text)
    if add_only:
        ops['handled'] = True
        resolve_add(add_only.group(1))
        return ops

    return ops

def _bot_has_remove_intent(raw_text):
    text = _normalize_str(raw_text)
    if not text:
        return False
    return bool(re.search(
        r'\b(quita|quitar|quite|borra|borrar|elimina|eliminar|saca|sacar|'
        r'cambia|cambiar|reemplaza|reemplazar|sustituye|sustituir)\b',
        text
    ))

@app.post('/api/bot/cotizacion')
def bot_crear_cotizacion():
    err = _require_bot_key()
    if err:
        return err
    d = request.json or {}
    chat_id = str(d.get('chat_id') or '').strip()
    raw_text = str(d.get('texto_original') or '').strip()
    has_remove_intent = _bot_has_remove_intent(raw_text)
    cliente = (d.get('cliente') or 'Sin especificar').strip()
    edit_ctx = _bot_ctx_get(chat_id)
    edit_cot = None
    if edit_ctx:
        edit_cot = query("SELECT * FROM cotizaciones WHERE id=?", (edit_ctx['cot_id'],), one=True)
        if not edit_cot:
            _bot_ctx_clear(chat_id)
    items_ok, items_ambiguos, items_no_encontrados = [], [], []
    for it in (d.get('items') or []):
        codigo = str(it.get('codigo') or '').strip()
        cant = max(1, int(it.get('cant') or 1))
        notas_item = str(it.get('notas_item') or '').strip()
        m = _match_catalog(codigo)
        if m['found']:
            items_ok.append({
                'id_producto': m['id_producto'],
                'nombre': m['nombre'],
                'cantidad': cant,
                'notas_item': notas_item,
            })
        elif m['ambiguous']:
            items_ambiguos.append({'query': codigo, 'opciones': m['ambiguous']})
        else:
            items_no_encontrados.append({'query': codigo})
    if edit_cot:
        edit_ops = _bot_parse_edit_ops(edit_cot['id'], raw_text)
        if edit_ops.get('handled'):
            edit_ambiguos = edit_ops['remove_ambiguos'] + edit_ops['add_ambiguos']
            edit_no_encontrados = edit_ops['remove_no_encontrados'] + edit_ops['add_no_encontrados']

            # Regla anti-duplicados: si el usuario pidio quitar/reemplazar, no agregamos nada
            # hasta poder identificar claramente el producto a retirar del mismo borrador.
            if has_remove_intent and not edit_ops['remove_ok']:
                return jsonify({
                    'ok': False,
                    'id': edit_cot['id'],
                    'error': 'No identifique con precision que producto quitar.',
                    'items_ambiguos': edit_ambiguos,
                    'items_no_encontrados': edit_no_encontrados,
                    'mensaje_telegram': _bot_build_quote_summary(
                        edit_cot['id'],
                        items_ambiguos=edit_ambiguos,
                        items_no_encontrados=edit_no_encontrados,
                        intro='No aplique cambios para evitar duplicados. Dime exactamente que producto debo quitar.'
                    ),
                })

            changed = False
            for it in edit_ops['remove_ok']:
                changed = _bot_remove_or_reduce_item(edit_cot['id'], it['codigo'], it['cantidad']) > 0 or changed
            for it in edit_ops['add_ok']:
                _bot_add_or_merge_item(edit_cot['id'], it['codigo'], it['cantidad'])
                changed = True
            if changed:
                _bot_ctx_clear(chat_id)
                return jsonify({
                    'ok': True,
                    'id': edit_cot['id'],
                    'no_cotizacion': edit_cot['no_cotizacion'],
                    'resumen_texto': _bot_build_quote_summary(
                        edit_cot['id'],
                        items_ambiguos=edit_ambiguos,
                        items_no_encontrados=edit_no_encontrados,
                        intro='Actualice este borrador con tu cambio.'
                    ),
                    'items_ambiguos': edit_ambiguos,
                    'items_no_encontrados': edit_no_encontrados,
                })
            return jsonify({
                'ok': False,
                'id': edit_cot['id'],
                'error': 'No pude aplicar ese cambio completo.',
                'items_ambiguos': edit_ambiguos,
                'items_no_encontrados': edit_no_encontrados,
                'mensaje_telegram': _bot_build_quote_summary(
                    edit_cot['id'],
                    items_ambiguos=edit_ambiguos,
                    items_no_encontrados=edit_no_encontrados,
                    intro='No pude aplicar ese cambio completo. Pruebame con mas detalle.'
                ),
            })
    if not items_ok and not items_ambiguos:
        faltan = ', '.join(x['query'] for x in items_no_encontrados)
        if edit_cot:
            return jsonify({
                'ok': False,
                'id': edit_cot['id'],
                'error': 'No encontre ningun producto en el catalogo.',
                'items_no_encontrados': items_no_encontrados,
                'mensaje_telegram': (
                    f"No pude actualizar el borrador {edit_cot['no_cotizacion']}. "
                    f"No reconoci estos productos: {faltan}"
                ),
            })
        return jsonify({
            'ok': False,
            'error': 'No encontre ningun producto en el catalogo.',
            'items_no_encontrados': items_no_encontrados,
            'mensaje_telegram': f'No encontre ningun producto. Productos no reconocidos: {faltan}',
        })
    if edit_cot:
        if has_remove_intent:
            return jsonify({
                'ok': False,
                'id': edit_cot['id'],
                'error': 'No pude aplicar la parte de quitar/reemplazar con ese texto.',
                'items_ambiguos': items_ambiguos,
                'items_no_encontrados': items_no_encontrados,
                'mensaje_telegram': _bot_build_quote_summary(
                    edit_cot['id'],
                    items_ambiguos=items_ambiguos,
                    items_no_encontrados=items_no_encontrados,
                    intro='No aplique cambios para evitar duplicados. Dime exactamente que producto debo quitar y cual agregar.'
                ),
            })
        for it in items_ok:
            _bot_add_or_merge_item(edit_cot['id'], it['id_producto'], it['cantidad'], it['notas_item'])
        _bot_ctx_clear(chat_id)
        return jsonify({
            'ok': True,
            'id': edit_cot['id'],
            'no_cotizacion': edit_cot['no_cotizacion'],
            'resumen_texto': _bot_build_quote_summary(
                edit_cot['id'],
                items_ambiguos=items_ambiguos,
                items_no_encontrados=items_no_encontrados,
                intro='Actualice este borrador con tu cambio.',
            ),
            'items_ambiguos': items_ambiguos,
            'items_no_encontrados': items_no_encontrados,
        })
    no_cot = next_no_cotizacion()
    desc_pct = float(d.get('descuento_pct') or 0) / 100
    ant_pct = float(d.get('anticipo_pct') or 70) / 100
    cot_id = execute(
        """INSERT INTO cotizaciones
        (no_cotizacion,cliente,telefono,ciudad,proyecto,tipo_cotizacion,forma_pago,
         anticipo_pct,descuento_pct,descuento_val,notas,vendedor,etapa,estado)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            no_cot,
            cliente,
            d.get('telefono', ''),
            d.get('ciudad', ''),
            d.get('proyecto', ''),
            'MIXTA',
            '70% - 30%',
            ant_pct,
            desc_pct,
            0,
            d.get('notas', ''),
            'Bot',
            'COTIZADA',
            'BORRADOR',
        ),
    )
    for idx, it in enumerate(items_ok, 1):
        execute(
            "INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item) VALUES (?,?,?,?,0,0,0,?)",
            (cot_id, idx, it['id_producto'], it['cantidad'], it['notas_item']),
        )
    lines = [f"*Borrador {no_cot}* - {cliente}"]
    if d.get('ciudad'):
        lines.append(f"Ciudad: {d['ciudad']}")
    if d.get('proyecto'):
        lines.append(f"Proyecto: {d['proyecto']}")
    lines.append('')
    for it in items_ok:
        lines.append(f"- {it['cantidad']}x {it['nombre']} ({it['id_producto']})")
    if items_ambiguos:
        lines.append('')
        lines.append('Ambiguos (elige abajo):')
        for a in items_ambiguos:
            lines.append(f'  "{a["query"]}" - varias opciones')
    if items_no_encontrados:
        lines.append('')
        lines.append('No encontre: ' + ', '.join(x['query'] for x in items_no_encontrados))
    return jsonify({
        'ok': True,
        'id': cot_id,
        'no_cotizacion': no_cot,
        'resumen_texto': '\n'.join(lines),
        'items_ambiguos': items_ambiguos,
        'items_no_encontrados': items_no_encontrados,
    })

@app.post('/api/bot/cotizacion/accion')
def bot_accion_cotizacion():
    err = _require_bot_key()
    if err:
        return err
    d = request.json or {}
    cot_id = d.get('id') or d.get('cot_id')
    accion = str(d.get('accion') or '').lower().strip()
    chat_id = str(d.get('chat_id') or '').strip()
    if accion == 'cambiar':
        if not cot_id:
            return jsonify({'ok': False, 'error': 'id requerido para cambiar algo.'})
        cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
        if not cot:
            return jsonify({'ok': False, 'error': f'Cotizacion {cot_id} no encontrada.'})
        if not chat_id:
            return jsonify({'ok': False, 'error': 'chat_id requerido para cambiar algo.'})
        _bot_ctx_set(chat_id, cot_id, 'editar')
        return jsonify({
            'ok': True,
            'estado': cot.get('estado') or 'BORRADOR',
            'modo': 'editar',
            'mensaje_telegram': f"Perfecto. Enviame solo lo que quieres agregar o corregir y lo sumo al borrador {cot['no_cotizacion']}.",
        })
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return jsonify({'ok': False, 'error': f'Cotizacion {cot_id} no encontrada.'})
    if accion == 'confirmar':
        _bot_ctx_clear(chat_id)
        _bot_ctx_clear_by_cot(cot_id)
        execute("UPDATE cotizaciones SET estado='ENVIADA' WHERE id=?", (cot_id,))
        return jsonify({
            'ok': True,
            'estado': 'ENVIADA',
            'mensaje_telegram': f"Cotizacion {cot['no_cotizacion']} confirmada y enviada a revision.",
        })
    elif accion == 'cancelar':
        _bot_ctx_clear(chat_id)
        _bot_ctx_clear_by_cot(cot_id)
        execute("UPDATE cotizaciones SET estado='RECHAZADA' WHERE id=?", (cot_id,))
        return jsonify({
            'ok': True,
            'estado': 'RECHAZADA',
            'mensaje_telegram': f"Cotizacion {cot['no_cotizacion']} cancelada.",
        })
    return jsonify({'ok': False, 'error': f'Accion desconocida: {accion}'})

@app.post('/api/bot/cotizacion/<int:cot_id>/agregar_item')
def bot_agregar_item(cot_id):
    err = _require_bot_key()
    if err: return err
    d = request.json or {}
    codigo = str(d.get('codigo') or '').strip().upper()
    cant = max(1, int(d.get('cant') or 1))
    cot = query("SELECT * FROM cotizaciones WHERE id=?", (cot_id,), one=True)
    if not cot:
        return jsonify({'ok': False, 'error': f'CotizaciÃ³n {cot_id} no encontrada.'})
    prod = query("SELECT * FROM catalogo WHERE id_producto=? AND activo=1", (codigo,), one=True)
    if not prod:
        return jsonify({'ok': False, 'error': f'Producto {codigo} no encontrado en catÃ¡logo.'})
    existing = query("SELECT * FROM items WHERE cot_id=? AND id_producto=?", (cot_id, codigo), one=True)
    if existing:
        execute("UPDATE items SET cantidad=cantidad+? WHERE id=?", (cant, existing['id']))
    else:
        ml = query("SELECT MAX(linea) as ml FROM items WHERE cot_id=?", (cot_id,), one=True)
        execute("INSERT INTO items (cot_id,linea,id_producto,cantidad,precio_manual,inst_manual,cfg_manual,notas_item) VALUES (?,?,?,?,0,0,0,'')",
                (cot_id, (ml['ml'] or 0) + 1, codigo, cant))
    return jsonify({'ok': True,
                    'mensaje_telegram': f"âœ… Agregado: {cant}x {prod['nombre']} a la cotizaciÃ³n {cot['no_cotizacion']}."})

@app.post('/api/bot/transcribir-audio')
def bot_transcribir_audio():
    """Recibe audio en base64 y transcribe vÃ­a OpenAI Whisper."""
    err = _require_bot_key()
    if err: return err
    openai_key = os.environ.get('OPENAI_API_KEY', '')
    if not openai_key:
        return jsonify({'ok': False, 'error': 'OPENAI_API_KEY no configurada.'}), 500
    d = request.json or {}
    audio_b64 = d.get('audio_b64') or ''
    filename = d.get('filename') or 'audio.ogg'
    mimetype = d.get('mimetype') or 'audio/ogg'
    prompt = d.get('prompt') or ''
    model = d.get('model') or 'whisper-1'
    if not audio_b64:
        return jsonify({'ok': False, 'error': 'audio_b64 vacÃ­o.'}), 400
    try:
        import base64 as _b64
        audio_bytes = _b64.b64decode(audio_b64)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error decodificando audio: {e}'}), 400
    boundary = secrets.token_hex(16)
    parts = []
    for name, val in [('model', model), ('language', 'es'), ('response_format', 'json')]:
        parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="{name}"\r\n\r\n{val}'.encode())
    if prompt:
        parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="prompt"\r\n\r\n{prompt}'.encode())
    parts.append(
        f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{filename}"\r\nContent-Type: {mimetype}\r\n\r\n'.encode()
        + audio_bytes)
    parts.append(f'--{boundary}--'.encode())
    body = b'\r\n'.join(parts)
    req = urllib.request.Request(
        'https://api.openai.com/v1/audio/transcriptions', data=body,
        headers={'Authorization': f'Bearer {openai_key}',
                 'Content-Type': f'multipart/form-data; boundary={boundary}'},
        method='POST')
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode())
        return jsonify({'ok': True, 'text': result.get('text', '')})
    except urllib.error.HTTPError as e:
        return jsonify({'ok': False, 'error': f'OpenAI {e.code}: {e.read().decode()}'}), 502
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 502


if __name__ == '__main__':
    init_db()
    _ensure_auto_backup_thread()
    print("\n" + "="*50)
    print("  ðŸš€ RC DOMOTIC Cotizador v3.0")
    print("  http://localhost:5000")
    print("  Red local: http://0.0.0.0:5000")
    print("="*50 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)

